"""
SAA KPI Monitor — Streamlit App (80%+ Attendance Workflow)
==================================================================
Redesign of the combined pipeline app, focused ONLY on the >=80% attendance
workflow. Adds KPI 3 (New Programmes by Month), a more forgiving Course Reporting Namelist
loader (reads the "Name List format" sheet out of the full SharePoint
workbook), dual-format Certificate List support, and a reorganized set of
tabs (Results / Data Quality / Reference & Classification / Advanced).

Removed vs the prior combined version:
  - The "All Attendance %" branch (results, quick-summary, download sheets)
  - The "Threshold Impact" comparison tab
  - The old "Org Name QA" tab (replaced by "Reference & Classification")
  - Bulky raw-table dumps in the Advanced tab

HOW TO RUN:
    streamlit run Pipeline_app_v2.py
"""

import io
import re
import datetime
from pathlib import Path
import pandas as pd
import numpy as np
import streamlit as st
import plotly.graph_objects as go
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# ────────────────────────────────────────────────────────────────
# PAGE CONFIG
# ────────────────────────────────────────────────────────────────
st.set_page_config(page_title="SAA KPI Monitor", layout="wide")

st.title("📊 SAA KPI Monitor")
st.caption("Upload your 5 exported files, set your reporting period, and click Run.")


# ────────────────────────────────────────────────────────────────
# CLASSIFICATION REFERENCE FILE UPLOADS (optional — override built-in defaults)
# ────────────────────────────────────────────────────────────────
# These three are separate from the 5 reporting files below — they replace
# the reference data used to compute segments (Local/CAAS/International),
# senior designations, and public-holiday-aware trainee-day counts, instead
# of relying on what's hardcoded in this script.
# Declared here, early, so an upload takes effect immediately — including in
# the "Manage Organization Name Classifications" editor further down, which
# would otherwise show stale data for one extra rerun after a fresh upload.
st.sidebar.header("🗂️ Classification Reference Files")
st.sidebar.caption(
    "These are no longer hardcoded in this script. Organisation Classification can also be built up "
    "manually in the on-page editor further down (and saved locally). Director Designation "
    "Classification and Singapore Public Holidays have no built-in fallback at all — without an "
    "upload here, KPI 2 will read 0, and trainee-day counts won't exclude any public holidays."
)

uploaded_org_class_file = st.sidebar.file_uploader(
    "Organisation Classification List (.xlsx or .csv)", type=["xlsx", "csv"], key="org_class_upload"
)
st.sidebar.caption(
    "Required columns: **Organization Name**, **Classification** (Local / CAAS / International / Unknown). "
    "Organization Type and Country/Region of Organization are optional and carried through if present."
)

uploaded_designation_file = st.sidebar.file_uploader(
    "Director Designation Classification List (.xlsx or .csv) — required for KPI 2", type=["xlsx", "csv"], key="designation_upload"
)
st.sidebar.caption("Required columns: **Designation**, **Category** (Private Sector / Public Sector).")

uploaded_holidays_file = st.sidebar.file_uploader(
    "Singapore Public Holidays List (.xlsx or .csv) — affects trainee-day accuracy", type=["xlsx", "csv"], key="holidays_upload"
)
st.sidebar.caption("Required column: **Date**. Update and re-upload this file annually — a stale or missing "
                    "list will silently under-count trainee days for courses spanning a real public holiday.")

st.sidebar.divider()


def _read_classification_upload(file, header_keywords=None):
    """Reads an uploaded classification file as .xlsx or .csv. Normally the
    header is expected on row 1 (data from row 2), but this scans the first
    15 rows for one that actually looks like a header — containing one of
    `header_keywords` (e.g. 'organization name') as a cell value — and uses
    that row instead. This handles files that still have legend/instruction
    rows above the real header (e.g. an earlier export of this table)
    without the caller needing to manually re-save a 'clean' version first."""
    is_csv = file.name.lower().endswith(".csv")
    raw = pd.read_csv(file, header=None) if is_csv else pd.read_excel(file, header=None)

    header_row_idx = 0
    if header_keywords:
        for i in range(min(15, len(raw))):
            row_values = [str(v).strip().lower() for v in raw.iloc[i].tolist()]
            if any(kw in row_values for kw in header_keywords):
                header_row_idx = i
                break

    df = raw.iloc[header_row_idx + 1:].copy()
    df.columns = raw.iloc[header_row_idx]
    df = df.reset_index(drop=True)
    df = df.dropna(how="all")  # drop fully-blank rows (e.g. a spacer row under the header)
    return df


def _file_sig(f):
    return (f.name, f.size) if f is not None else None


# ────────────────────────────────────────────────────────────────
# ORGANIZATION-NAME CLASSIFICATION TABLE
# ──────────────────────────────────────────────────────────────
# No longer hardcoded in this script. Loaded from a local cache file
# (org_classification_data.csv, written whenever you click "Save changes"
# below) if one exists, else starts empty — populate it by uploading an
# Organisation Classification List (sidebar) or adding rows directly in the
# editor below.
# ──────────────────────────────────────────────────────────────
ORG_CLASS_COLUMNS = ['Organization Name', 'Classification', 'Needs Review']
ORG_CLASS_FILE = Path(__file__).resolve().parent / "org_classification_data.csv"


def _load_org_class_df_from_disk_or_default():
    """Loads the local cache file if it exists, else starts with an empty table."""
    if ORG_CLASS_FILE.exists():
        try:
            df = pd.read_csv(ORG_CLASS_FILE)
            if 'Organization Name' in df.columns and 'Classification' in df.columns:
                if 'Needs Review' not in df.columns:
                    df['Needs Review'] = False
                return df, f"local file ({ORG_CLASS_FILE.name})"
        except Exception:
            pass
    df = pd.DataFrame(columns=ORG_CLASS_COLUMNS)
    return df, "no data yet — upload an Organisation Classification List, or add rows manually below"


if "org_class_df" not in st.session_state:
    _df0, _src0 = _load_org_class_df_from_disk_or_default()
    st.session_state.org_class_df = _df0
    st.session_state.org_class_source = _src0

if "org_class_upload_sig" not in st.session_state:
    st.session_state.org_class_upload_sig = None

VALID_ORG_CLASSIFICATIONS = {"Local", "CAAS", "International", "Unknown"}


def _find_col(df, candidates):
    """Case/whitespace-insensitive column finder, e.g. so 'Organisation Name'
    (British spelling, matching this app's own UI labels) and 'Organization
    Name' (American spelling) are both recognised as the same column."""
    cols_lower = {re.sub(r'\s+', ' ', c).strip().lower(): c for c in df.columns}
    for cand in candidates:
        if cand in cols_lower:
            return cols_lower[cand]
    return None


def _load_uploaded_org_class(file):
    """Parses an uploaded Organisation Classification List and shapes it to
    match the app's working table (Needs Review added if missing). Returns
    None (and shows a sidebar error) if required columns can't be found —
    tolerant of British/American spelling, case, and stray whitespace in
    the header row."""
    try:
        df = _read_classification_upload(file, header_keywords=["organization name", "organisation name"])
    except Exception as e:
        st.sidebar.error(f"❌ Could not read Organisation Classification List: {e}")
        return None

    name_col = _find_col(df, ["organization name", "organisation name"])
    class_col = _find_col(df, ["classification"])
    type_col = _find_col(df, ["organization type", "organisation type"])
    country_col = _find_col(df, ["country/region of organization", "country/region of organisation",
                                  "country of organization", "country of organisation", "country"])
    review_col = _find_col(df, ["needs review"])

    missing = []
    if name_col is None:
        missing.append("Organization Name (or 'Organisation Name')")
    if class_col is None:
        missing.append("Classification")
    if missing:
        st.sidebar.error(
            f"❌ Organisation Classification List is missing column(s): {', '.join(missing)}.\n\n"
            f"Columns found in your file: {', '.join(str(c) for c in df.columns)}"
        )
        return None

    # Normalise to the app's internal column names regardless of what the
    # uploaded file called them.
    rename_map = {name_col: "Organization Name", class_col: "Classification"}
    if type_col:
        rename_map[type_col] = "Organization Type"
    if country_col:
        rename_map[country_col] = "Country/Region of Organization"
    if review_col:
        rename_map[review_col] = "Needs Review"
    df = df.rename(columns=rename_map)

    if "Needs Review" not in df.columns:
        df["Needs Review"] = False
    else:
        df["Needs Review"] = df["Needs Review"].fillna(False).astype(bool)

    bad_rows = ~df["Classification"].astype(str).str.strip().isin(VALID_ORG_CLASSIFICATIONS)
    if bad_rows.any():
        st.sidebar.warning(
            f"⚠️ {int(bad_rows.sum())} row(s) in the uploaded Organisation Classification List have a "
            "Classification value outside Local/CAAS/International/Unknown — they'll still load, "
            "but won't match any KPI segment correctly."
        )
    return df


_org_class_upload_sig = _file_sig(uploaded_org_class_file)
if uploaded_org_class_file is not None and _org_class_upload_sig != st.session_state.org_class_upload_sig:
    _uploaded_org_df = _load_uploaded_org_class(uploaded_org_class_file)
    if _uploaded_org_df is not None:
        st.session_state.org_class_df = _uploaded_org_df
        st.session_state.org_class_source = f"uploaded file ({uploaded_org_class_file.name})"
        st.session_state.org_class_upload_sig = _org_class_upload_sig
        st.sidebar.success(f"✅ Loaded {len(_uploaded_org_df):,} rows from {uploaded_org_class_file.name}")

with st.expander("✏️ Manage Organization Name Classifications", expanded=False):
    st.caption(
        "This table decides whether each organization name is classified as **Local**, "
        "**CAAS**, or **International**. Populate it by uploading an Organisation Classification "
        "List (see '🗂️ Classification Reference Files' in the sidebar), or edit any cell, add new "
        "rows at the bottom, or delete a row using the row menu right here. As soon as you make a "
        "change, every KPI, chart, and table on this page automatically recomputes using the updated "
        "classification (as long as you've clicked **Run Pipeline** at least once for your current "
        "files). Click **Save changes** to persist your edits to a local file so they're still there "
        "next time you run the app, without needing to re-upload. "
        f"Currently loaded from: **{st.session_state.org_class_source}** "
        f"({len(st.session_state.org_class_df):,} rows)."
    )

    _edited_org_df = st.data_editor(
        st.session_state.org_class_df,
        num_rows="dynamic",
        use_container_width=True,
        height=350,
        column_config={
            "Classification": st.column_config.SelectboxColumn(
                "Classification",
                options=["Local", "CAAS", "International", "Unknown"],
                required=True,
            ),
            "Needs Review": st.column_config.CheckboxColumn("Needs Review"),
        },
        key="org_class_editor",
    )
    # Live edits are used immediately by the pipeline below — no need to
    # click Save just to have this session's run reflect your changes.
    st.session_state.org_class_df = _edited_org_df

    oc1, oc2, oc3 = st.columns([1, 1, 2])
    if oc1.button("💾 Save changes"):
        try:
            _edited_org_df.to_csv(ORG_CLASS_FILE, index=False)
            st.session_state.org_class_source = f"local file ({ORG_CLASS_FILE.name})"
            st.success(f"Saved {len(_edited_org_df):,} rows to {ORG_CLASS_FILE.name}. This will load automatically next time.")
        except Exception as e:
            st.warning(f"Could not write to local file ({e}) — your edits are still active for this session.")
    if oc2.button("🗑️ Clear table"):
        st.session_state.org_class_df = pd.DataFrame(columns=ORG_CLASS_COLUMNS)
        st.session_state.org_class_source = "cleared — upload a file or add rows manually"
        st.rerun()
    oc3.download_button(
        "⬇️ Download table as CSV",
        data=_edited_org_df.to_csv(index=False).encode("utf-8"),
        file_name=f"Organisation_Classification_List_{datetime.date.today().strftime('%Y%m%d')}.csv",
        mime="text/csv",
    )


# ────────────────────────────────────────────────────────────────
# SIDEBAR — FILE UPLOAD + SETTINGS
# ────────────────────────────────────────────────────────────────
st.sidebar.header("📁 Step 1 — Upload Files")

with st.sidebar.expander("ℹ️ Where do I find each report?"):
    st.markdown(
        """
**Admitted List**
Report → Course, application and enrolment → Applicant details report
*Example file name: `Application_Details_Report_20260709153318`*

**Enrollment List**
Report → Course, application and enrolment → Insight report → Course intakes
*Example file name: `Detailed_Insights_Report_20260709153540`*

**Attendance List**
Report → Progress and timetabling → Attendance & session information report → Attendance information
*Example file name: `Attendance_Information_Report_20260526110649`*

**Certificate List**
Report → Progress and timetabling → Assessment and graduation → Course & module certificate report → Module Certificate
*Example file name: `Module_Certificate_Report_20260709153714`*

**Course Reporting Namelist**
Download the Excel workbook from SharePoint and upload it as-is — no need to delete other sheets.
The app automatically reads the **Name List format** sheet.
*Example file: `Course Reporting and Evaluation_NEW.xlsx`*
        """
    )

uploaded_admitted = st.sidebar.file_uploader("Admitted List (.xlsx)", type="xlsx")
st.sidebar.caption("e.g. Application_Details_Report_20260709153318.xlsx")

uploaded_enrollment = st.sidebar.file_uploader("Enrollment List (.xlsx)", type="xlsx")
st.sidebar.caption("e.g. Detailed_Insights_Report_20260709153540.xlsx")

uploaded_attendance = st.sidebar.file_uploader("Attendance List (.xlsx)", type="xlsx")
st.sidebar.caption("e.g. Attendance_Information_Report_20260526110649.xlsx")

uploaded_cert = st.sidebar.file_uploader("Certificate List (.xlsx)", type="xlsx")
st.sidebar.caption("e.g. Module_Certificate_Report_20260626143252.xlsx (standard Learn@SAA export)")

uploaded_namelist = st.sidebar.file_uploader("Course Reporting Namelist workbook (.xlsx)", type="xlsx")
st.sidebar.caption("e.g. Course Reporting and Evaluation_NEW.xlsx (full SharePoint workbook)")

st.sidebar.caption("💡 Org Name Classification and Senior Designations can be managed on-page (see the "
                    "'✏️ Manage Organization Name Classifications' panel above) or overridden via the "
                    "'🗂️ Classification Reference Files' upload section at the top of the sidebar.")

st.sidebar.divider()
st.sidebar.header("⚙️ Step 2 — Settings")

_today = datetime.date.today()
start_date = st.sidebar.date_input("Report start date", value=datetime.date(_today.year, 1, 1))
end_date = st.sidebar.date_input("Report end date", value=_today)

attendance_threshold = st.sidebar.slider("Minimum attendance % (by session)", 0, 100, 80)

KPI1_TARGET = st.sidebar.number_input("KPI 1 annual target (trainee days)", value=10000, step=500)
KPI2_TARGET = st.sidebar.number_input(
    "KPI 2 annual target (international trainee days for directors and above)", value=2000, step=100
)
KPI3_TARGET = st.sidebar.number_input("KPI 3 annual target (new programmes)", value=10, step=1)

st.sidebar.divider()
show_advanced = st.sidebar.checkbox("Show data preview", value=False)

start_date = pd.Timestamp(start_date)
end_date = pd.Timestamp(end_date)

all_uploaded = all([uploaded_admitted, uploaded_enrollment, uploaded_attendance,
                     uploaded_cert, uploaded_namelist])


# ────────────────────────────────────────────────────────────────
# HELPERS
# ────────────────────────────────────────────────────────────────
def safe_read_excel(file, label, expected_cols, **kwargs):
    """Reads an Excel file and gives a friendly error message instead of a
    scary Python crash if something is wrong with the file."""
    try:
        df = pd.read_excel(file, **kwargs)
    except ValueError as e:
        st.error(
            f"❌ Could not read **{label}**.\n\n"
            f"This usually means the sheet name has changed, or the wrong file "
            f"was uploaded into this slot.\n\n"
            f"Technical detail: {e}"
        )
        st.stop()
    except Exception as e:
        st.error(f"❌ Unexpected problem reading **{label}**: {e}")
        st.stop()

    missing = [c for c in expected_cols if c not in df.columns]
    if missing:
        st.error(
            f"❌ **{label}** is missing expected column(s): {', '.join(missing)}.\n\n"
            f"This usually means the wrong file was uploaded into this slot, "
            f"or the website export format has changed. Please check and re-upload."
        )
        st.stop()

    return df


def safe_step(step_name, func, *args, **kwargs):
    """Runs a pipeline step and catches any unexpected error, showing a
    friendly message instead of crashing the whole app."""
    try:
        return func(*args, **kwargs)
    except Exception as e:
        st.error(
            f"❌ Something went wrong during **{step_name}**.\n\n"
            f"This is likely a data issue rather than an app problem. "
            f"Please check your uploaded files match the expected format.\n\n"
            f"Technical detail: {type(e).__name__}: {e}"
        )
        st.stop()


def load_namelist_workbook(file, expected_cols):
    """Searches the uploaded workbook for a sheet called 'Name List format'
    (case-insensitive, trimmed). Falls back to the first sheet if not found."""
    try:
        xls = pd.ExcelFile(file)
    except Exception as e:
        st.error(f"❌ Could not open **Course Reporting Namelist workbook**: {e}")
        st.stop()

    target_sheet = None
    for sheet in xls.sheet_names:
        if sheet.strip().lower() == "name list format":
            target_sheet = sheet
            break

    if target_sheet is None:
        st.info(
            "ℹ️ Could not find a sheet named **'Name List format'** in the Course Reporting Namelist "
            f"workbook — falling back to the first sheet ('{xls.sheet_names[0]}')."
        )
        target_sheet = xls.sheet_names[0]

    df = pd.read_excel(xls, sheet_name=target_sheet, header=0)
    missing = [c for c in expected_cols if c not in df.columns]
    if missing:
        st.error(
            f"❌ Course Reporting Namelist sheet **'{target_sheet}'** is missing expected column(s): "
            f"{', '.join(missing)}. Please check the workbook."
        )
        st.stop()
    return df, target_sheet


# ---- Certificate loader (standard Learn@SAA export only) ----
CERT_FORMAT_A_COLS = [
    "Learner name", "Learner ID", "Branch", "School", "Course code", "Course name",
    "Course intake No.", "Enrollment status", "Withhold status", "Payment status",
    "Certificate name", "Certificate status", "Generation date",
    "Certificate serial No.", "Certificate number",
]


def load_certificate_file(file):
    """Loads the standard Learn@SAA module certificate export."""
    try:
        raw = pd.read_excel(file, header=0)
    except Exception as e:
        st.error(f"❌ Could not read **Certificate List**: {e}")
        st.stop()

    required_min = ["Learner name", "Course intake No.", "Certificate status"]
    missing = [c for c in required_min if c not in raw.columns]
    if missing:
        st.error(
            f"❌ **Certificate List** is missing expected column(s): {', '.join(missing)}."
        )
        st.stop()

    df = raw.copy()
    # 'Email Address' isn't part of the standard export, but is used
    # downstream as a dedup fallback key when Learner ID is missing — kept
    # as an always-present (blank) column so that logic doesn't break.
    if "Email Address" not in df.columns:
        df["Email Address"] = pd.NA

    return df, "Standard Learn@SAA export"


# ────────────────────────────────────────────────────────────────
# MAIN PIPELINE
# ────────────────────────────────────────────────────────────────
if not all_uploaded:
    st.info("⬆️ Please upload all 5 files in the sidebar to continue.")
    st.stop()

st.sidebar.success("✅ All 5 files uploaded")


# ────────────────────────────────────────────────────────────────
# RUN-BUTTON STATE
# ────────────────────────────────────────────────────────────────
# Tracks which uploaded files + settings the last successful run used, so
# that (a) results stay visible across unrelated reruns (e.g. editing the
# org classification table, clicking Download), and (b) re-uploading a new
# file automatically requires — and correctly re-enables — a fresh click of
# "Run Pipeline", with no need to refresh the whole page.
def _file_signature(f):
    return (f.name, f.size) if f is not None else None


current_signature = (
    _file_signature(uploaded_admitted), _file_signature(uploaded_enrollment),
    _file_signature(uploaded_attendance), _file_signature(uploaded_cert),
    _file_signature(uploaded_namelist),
    str(start_date), str(end_date), attendance_threshold,
    KPI1_TARGET, KPI2_TARGET, KPI3_TARGET,
)

if "pipeline_should_run" not in st.session_state:
    st.session_state.pipeline_should_run = False
if "pipeline_signature" not in st.session_state:
    st.session_state.pipeline_signature = None

run_clicked = st.sidebar.button("▶️  Run Pipeline", type="primary", use_container_width=True)

if run_clicked:
    st.session_state.pipeline_should_run = True
    st.session_state.pipeline_signature = current_signature

# Files or settings changed since the last run — require a fresh click.
if st.session_state.pipeline_signature != current_signature:
    st.session_state.pipeline_should_run = False

if not st.session_state.pipeline_should_run:
    if st.session_state.pipeline_signature is not None:
        st.info("🔄 New files or settings detected. Click **Run Pipeline** in the sidebar to refresh the results.")
    else:
        st.info("👈 Files are ready. Click **Run Pipeline** in the sidebar to begin.")
    st.stop()


with st.spinner("Running pipeline..."):

    # ============================================================
    # STEP 0 — LOAD FILES
    # ============================================================
    df_admitted = safe_read_excel(
        uploaded_admitted, "Admitted List",
        expected_cols=["Learner name", "Course intake No.", "Designation",
                        "Country/Region of organization", "Email address", "Learner ID"],
        sheet_name='Application details report', header=0
    )

    df_enrollment = safe_read_excel(
        uploaded_enrollment, "Enrollment List",
        expected_cols=["Course intake No.", "No. of enrollments", "Course intake end date"],
        sheet_name='Detailed Insights Report', header=0
    )

    df_attendance = safe_read_excel(
        uploaded_attendance, "Attendance List",
        expected_cols=["Learner name", "Course intake No.", "Learner email address",
                        "Module attendance percentage (by session number)",
                        "Course start date", "Course end date"],
        dtype={'Module attendance percentage (by session hours)': str}
    )

    df_cert, cert_format_label = safe_step("Load Certificate List", load_certificate_file, uploaded_cert)

    df_namelist, namelist_sheet_used = safe_step(
        "Load Course Reporting Namelist Workbook", load_namelist_workbook, uploaded_namelist,
        ["Course Intake Number", "Name", "Country", "No. of Trainee Days"]
    )

    # ------------------------------------------------------------
    # Load org-name classification table (from the in-page editor / session state)
    # ------------------------------------------------------------
    def step_load_org_class_table():
        raw = st.session_state.org_class_df.copy()
        if 'Needs Review' not in raw.columns:
            raw['Needs Review'] = False

        raw = raw.dropna(subset=['Organization Name']).copy()
        raw['_key'] = raw['Organization Name'].astype(str).str.strip().str.replace(r'\s+', ' ', regex=True).str.upper()
        raw = raw.drop_duplicates(subset='_key', keep='first')
        lookup_dict = {
            row['_key']: (row['Classification'], bool(row['Needs Review']))
            for _, row in raw.iterrows()
        }
        return lookup_dict, st.session_state.org_class_source, raw

    org_class_lookup, org_class_source, org_class_table_raw = safe_step(
        "Load Org Name Classification Table", step_load_org_class_table
    )

    # ============================================================
    # BLOCK 1.1: Date Filter (Admitted / Enrollment / Attendance)
    # Certificate rows are filtered later by intake membership, not by date,
    # since the standard Certificate export has no Course Start/End Date columns.
    # ============================================================
    def step_date_filter():
        df_admitted['Course end date'] = pd.to_datetime(df_admitted['Course end date'], dayfirst=True, errors='coerce')
        df_enrollment['Course intake end date'] = pd.to_datetime(df_enrollment['Course intake end date'], format='%d %b %Y', errors='coerce')
        df_attendance['Course end date'] = pd.to_datetime(df_attendance['Course end date'], dayfirst=True, errors='coerce')

        a = df_admitted[(df_admitted['Course end date'] >= start_date) & (df_admitted['Course end date'] <= end_date)]
        e = df_enrollment[(df_enrollment['Course intake end date'] >= start_date) & (df_enrollment['Course intake end date'] <= end_date)]
        att = df_attendance[(df_attendance['Course end date'] >= start_date) & (df_attendance['Course end date'] <= end_date)]
        return a, e, att

    df_admitted, df_enrollment, df_attendance = safe_step("Date Filter", step_date_filter)

    if len(df_admitted) == 0 or len(df_attendance) == 0:
        st.warning(
            "⚠️ No rows remain after applying the date filter. "
            "Check that your report start/end dates match the data in your files."
        )

    # ============================================================
    # BLOCK 1.2: Compute attendance % and apply the 80%+ gate
    # ============================================================
    def step_attendance_pct_and_gate():
        df_attendance['Attendance % (by session)'] = pd.to_numeric(
            df_attendance['Module attendance percentage (by session number)'].str.replace('%', '', regex=False),
            errors='coerce'
        )
        gated = df_attendance[df_attendance['Attendance % (by session)'] >= attendance_threshold].copy()
        return df_attendance, gated

    df_attendance, df_attendance_gated = safe_step("Attendance % Parsing + Threshold Gate", step_attendance_pct_and_gate)

    # Course intake numbers surviving the attendance gate — used to align Certificate rows
    surviving_intakes = set(df_attendance_gated['Course intake No.'].astype(str).str.strip().unique())

    # ============================================================
    # BLOCK 2.1 & 2.2: Deduplication (Admitted / Enrollment / Cert)
    # ============================================================
    PERSONAL_EMAIL_DOMAINS = {
        'gmail.com', 'googlemail.com', 'yahoo.com', 'yahoo.co.uk', 'yahoo.co.in',
        'yahoo.co.id', 'ymail.com', 'rocketmail.com', 'hotmail.com', 'hotmail.co.uk',
        'outlook.com', 'live.com', 'msn.com', 'icloud.com', 'me.com', 'mac.com',
        'aol.com', 'protonmail.com', 'proton.me', 'qq.com', '163.com', '126.com',
        'zoho.com', 'gmx.com', 'mail.com',
    }

    def is_personal_email(email):
        email = str(email).strip().lower()
        if '@' not in email:
            return False
        return email.rsplit('@', 1)[-1] in PERSONAL_EMAIL_DOMAINS

    def dedup_by_name_and_email(df, name_col, email_col, intake_col):
        # When the same learner (same name + intake) appears with more than one
        # email address, prefer keeping the row with the official/work email
        # over a personal one (gmail, yahoo, hotmail, etc.) — a stable sort
        # puts non-personal-email rows first so drop_duplicates(keep='first')
        # retains them; original file order still breaks ties within each group.
        d = df.copy()
        d['_is_personal_email'] = d[email_col].apply(is_personal_email)
        d = d.sort_values('_is_personal_email', kind='stable').drop(columns='_is_personal_email')
        step1 = d.drop_duplicates(subset=[name_col, intake_col], keep='first')
        step2 = step1.drop_duplicates(subset=[email_col, intake_col], keep='first')
        return step2.reset_index(drop=True)

    def step_dedup_non_attendance():
        a = dedup_by_name_and_email(df_admitted, 'Learner name', 'Email address', 'Course intake No.')
        e = df_enrollment.drop_duplicates(subset=['Course intake No.'], keep='first').reset_index(drop=True)

        # Certificate: align to intakes surviving the attendance gate first
        c_raw = df_cert.copy()
        c_raw['Course intake No.'] = c_raw['Course intake No.'].astype(str).str.strip()
        c_aligned = c_raw[c_raw['Course intake No.'].isin(surviving_intakes)].copy()

        # Primary match key: intake + Learner ID. Backup: intake + Email Address
        # (when Learner ID is missing) — implemented as a dedup key choice.
        has_id = c_aligned['Learner ID'].notna() & (c_aligned['Learner ID'].astype(str).str.strip() != '')
        c_with_id = c_aligned[has_id]
        c_without_id = c_aligned[~has_id]

        c_with_id_dedup = c_with_id.drop_duplicates(subset=['Course intake No.', 'Learner ID'], keep='first')
        c_without_id_dedup = c_without_id.drop_duplicates(subset=['Course intake No.', 'Email Address'], keep='first')
        c = pd.concat([c_with_id_dedup, c_without_id_dedup], ignore_index=True)
        return a, e, c

    df_admitted_clean, df_enrollment_clean, df_cert_clean = safe_step("Deduplication", step_dedup_non_attendance)

    # ============================================================
    # SHARED HELPERS
    # ============================================================
    def _load_public_holidays():
        """Singapore Public Holidays is no longer hardcoded in this script —
        it comes entirely from the uploaded Singapore Public Holidays List.
        Without a valid upload, trainee-day calculations proceed with NO
        holidays excluded (a pure business-day count), which will slightly
        UNDER-count trainee days for any course that spans a real public
        holiday — a silent accuracy risk, not a crash, so it's flagged here."""
        if uploaded_holidays_file is not None:
            try:
                _hol_df = _read_classification_upload(uploaded_holidays_file, header_keywords=["date"])
            except Exception as e:
                st.warning(f"⚠️ Could not read Singapore Public Holidays List ({e}) — "
                           "trainee-day calculations will proceed with NO public holidays excluded.")
                return np.array([], dtype='datetime64[D]'), "no data (upload failed)"

            _date_col = _find_col(_hol_df, ["date"])
            if _date_col is None:
                st.warning(
                    f"⚠️ Singapore Public Holidays List is missing a 'Date' column — "
                    f"columns found: {', '.join(str(c) for c in _hol_df.columns)}. "
                    "Trainee-day calculations will proceed with NO public holidays excluded."
                )
                return np.array([], dtype='datetime64[D]'), "no data (upload invalid)"

            _parsed_dates = pd.to_datetime(_hol_df[_date_col], dayfirst=True, errors='coerce').dropna()
            _n_bad = len(_hol_df) - len(_parsed_dates)
            if _n_bad > 0:
                st.warning(f"⚠️ {_n_bad} row(s) in the Singapore Public Holidays List had an unparseable Date and were skipped.")
            _holidays_np = np.array([d.strftime('%Y-%m-%d') for d in _parsed_dates], dtype='datetime64[D]')
            return _holidays_np, f"uploaded file ({uploaded_holidays_file.name}, {len(_holidays_np):,} dates)"

        st.warning("⚠️ No Singapore Public Holidays List uploaded — trainee-day calculations will proceed "
                   "with NO public holidays excluded (a pure business-day count), which will slightly "
                   "under-count trainee days for any course spanning a real public holiday.")
        return np.array([], dtype='datetime64[D]'), "no data — upload a Singapore Public Holidays List"

    sg_holidays_np, public_holidays_source = _load_public_holidays()

    def calculate_course_duration(start, end):
        if pd.isna(start) or pd.isna(end):
            return 0
        if start.weekday() >= 5:
            return (end - start).days + 1
        else:
            return np.busday_count(start.date(), (end + pd.Timedelta(days=1)).date(), holidays=sg_holidays_np)

    def _find_namelist_col(candidates, cols):
        cols_lower = {c.lower().strip(): c for c in cols}
        for cand in candidates:
            if cand in cols_lower:
                return cols_lower[cand]
        return None

    def _build_namelist_school_lookup():
        """Builds an intake-number -> School lookup from the Course Reporting
        Namelist, used only as a fallback for intake numbers that don't match
        the standard AES/AVS/ATS/AMS prefix pattern — replaces what used to
        be a hardcoded 3-entry MANUAL_SCHOOL_MAPPING for known exceptions.
        Degrades gracefully: returns an empty lookup (non-standard intakes
        fall through to 'Others', same as before) if the Namelist doesn't
        have a recognisable School column."""
        school_col = _find_namelist_col(['school'], df_namelist.columns)
        intake_col = _find_namelist_col(['course intake number', 'course intake no.'], df_namelist.columns)
        if school_col is None or intake_col is None:
            return {}, "No 'School' column found in the Course Reporting Namelist — non-standard intake numbers will show as 'Others'."
        nl = df_namelist[[intake_col, school_col]].dropna(subset=[intake_col]).copy()
        nl['_key'] = nl[intake_col].astype(str).str.strip().str.upper()
        nl = nl.drop_duplicates(subset='_key', keep='first')
        lookup = dict(zip(nl['_key'], nl[school_col].astype(str).str.strip()))
        return lookup, f"Loaded {len(lookup):,} intake-to-School mappings from the Course Reporting Namelist ('{school_col}' column) as a fallback for non-standard intake numbers."

    NAMELIST_SCHOOL_LOOKUP, namelist_school_source_note = _build_namelist_school_lookup()

    def map_school(intake):
        if pd.isna(intake):
            return 'Others'
        intake = str(intake).strip()
        for prefix in ('AES', 'AVS', 'ATS', 'AMS'):
            if intake.upper().startswith(prefix):
                return prefix
        # Non-standard intake number format (e.g. a numeric CAAS course ID
        # instead of the usual school-prefixed code) — check the Namelist
        # before giving up and calling it 'Others'.
        key = intake.upper()
        if key in NAMELIST_SCHOOL_LOOKUP:
            return NAMELIST_SCHOOL_LOOKUP[key]
        return 'Others'

    CAAS_NAMES = ['civil aviation authority of singapore']

    def classify_segment_country(row):
        """Country-of-organisation-based segment — used only for the Country
        lookup match-rate check in Data Quality (not for QA comparison)."""
        country = str(row.get('Country of Organisation', '')).strip().lower()
        org = str(row.get('Company name', '')).strip().lower()
        if country in ('', 'nan'):
            return 'Unknown'
        if country != 'singapore':
            return 'International'
        if any(c in org for c in CAAS_NAMES):
            return 'CAAS'
        return 'Local'

    # ---- Organization-name classifier (curated table only — no heuristic
    # fallback. Anything not in the uploaded Organisation Classification
    # table is Unknown, flagged for review, rather than silently guessed.) ----
    def _norm_org_key(name):
        return re.sub(r'\s+', ' ', str(name).strip()).upper()

    def classify_org_name(name, lookup_dict):
        if pd.isna(name) or str(name).strip() == '':
            return 'Unknown', False, 'none', 'empty organization name'
        key = _norm_org_key(name)
        if key in lookup_dict:
            cls, needs_review = lookup_dict[key]
            return cls, bool(needs_review), 'table', 'matched curated classification table'
        return 'Unknown', True, 'none', 'not found in Organisation Classification table — needs to be added'

    def get_admitted_org_column(df_admitted_clean):
        candidates = ['organization name', 'organisation name', 'organization', 'organisation',
                      'company name', 'company']
        cols_lower = {c.lower().strip(): c for c in df_admitted_clean.columns}
        for cand in candidates:
            if cand in cols_lower:
                return cols_lower[cand]
        return None

    ADMITTED_ORG_COL = get_admitted_org_column(df_admitted_clean)

    def build_org_name_lookup(df_att, df_admitted_clean, lookup_dict, org_col):
        df = df_att.copy()
        if org_col:
            admitted_org = df_admitted_clean[['Course intake No.', 'Learner ID', org_col]].copy()
            admitted_org['Course intake No.'] = admitted_org['Course intake No.'].astype(str).str.strip()
            admitted_org['Learner ID'] = admitted_org['Learner ID'].astype(str).str.strip()
            admitted_org = admitted_org.rename(columns={org_col: 'Organization Name'})
            df['Course intake No.'] = df['Course intake No.'].astype(str).str.strip()
            df['Learner ID'] = df['Learner ID'].astype(str).str.strip()
            df = df.merge(admitted_org, on=['Course intake No.', 'Learner ID'], how='left')
        elif 'Company name' in df.columns:
            df['Organization Name'] = df['Company name']
        else:
            df['Organization Name'] = ''

        results = df['Organization Name'].apply(lambda x: classify_org_name(x, lookup_dict))
        df['Segment (Org Name)'] = results.apply(lambda r: r[0])
        df['Org Name Needs Review'] = results.apply(lambda r: r[1])
        df['Org Name Match Source'] = results.apply(lambda r: r[2])
        df['Org Name Match Reason'] = results.apply(lambda r: r[3])
        df['Trainee Type (Org Name)'] = df['Segment (Org Name)']
        return df

    if ADMITTED_ORG_COL:
        org_name_source_note = f"Organization name pulled from Admitted List column '{ADMITTED_ORG_COL}'."
    elif 'Company name' in df_attendance.columns:
        org_name_source_note = "No organization-name column found in the Admitted List — used the Attendance List's 'Company name' column as a fallback."
    else:
        org_name_source_note = "No organization-name column found in either file — Trainee Type (Org Name) defaulted to 'Unknown' for all rows."

    def step_country(df_att, df_admitted_clean):
        admitted_lookup = df_admitted_clean[['Course intake No.', 'Learner ID', 'Country/Region of organization']].copy()
        admitted_lookup['Course intake No.'] = admitted_lookup['Course intake No.'].str.strip()
        admitted_lookup['Learner ID'] = admitted_lookup['Learner ID'].astype(str).str.strip()

        df_att['Course intake No.'] = df_att['Course intake No.'].str.strip()
        df_att['Learner ID'] = df_att['Learner ID'].astype(str).str.strip()

        merged = df_att.merge(admitted_lookup, on=['Course intake No.', 'Learner ID'], how='left') \
            .rename(columns={'Country/Region of organization': 'Country of Organisation'})
        merged = merged.loc[:, ~merged.columns.duplicated()]
        merged['Country of Organisation'] = merged['Country of Organisation'].str.strip().str.upper()
        return merged

    def step_designation(df_att, df_admitted_clean):
        df = df_att.drop(columns=['Designation'], errors='ignore')
        designation_lookup = df_admitted_clean[['Course intake No.', 'Learner ID', 'Learner name', 'Email address', 'Designation']].copy()

        df = df.merge(designation_lookup[['Course intake No.', 'Learner ID', 'Designation']],
                       on=['Course intake No.', 'Learner ID'], how='left')

        unmatched_mask = df['Designation'].isna()
        if unmatched_mask.sum() > 0:
            name_lookup = designation_lookup[['Course intake No.', 'Learner name', 'Designation']].rename(columns={'Designation': 'Designation_name'})
            df = df.merge(name_lookup, on=['Course intake No.', 'Learner name'], how='left')
            df['Designation'] = df['Designation'].fillna(df['Designation_name'])
            df = df.drop(columns=['Designation_name'])

        unmatched_mask = df['Designation'].isna()
        if unmatched_mask.sum() > 0:
            email_lookup = designation_lookup[['Course intake No.', 'Email address', 'Designation']].rename(columns={'Designation': 'Designation_email'})
            df = df.merge(email_lookup, left_on=['Course intake No.', 'Learner email address'],
                           right_on=['Course intake No.', 'Email address'], how='left')
            df['Designation'] = df['Designation'].fillna(df['Designation_email'])
            df = df.drop(columns=['Designation_email', 'Email address'], errors='ignore')
        return df

    REPORT_YEAR = end_date.year
    month_map = {1: 'Jan', 2: 'Feb', 3: 'Mar', 4: 'Apr', 5: 'May', 6: 'Jun',
                 7: 'Jul', 8: 'Aug', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dec'}
    REPORT_MONTHS = list(range(start_date.month, end_date.month + 1)) if start_date.year == end_date.year else list(range(1, 13))

    def _load_designation_classification():
        """Senior Designation Classification is no longer hardcoded in this
        script — it comes entirely from the uploaded Director Designation
        Classification List. Without a valid upload, KPI 2 (senior trainee
        days) reads 0 for every segment, since no designation can match."""
        if uploaded_designation_file is not None:
            try:
                _desig_df = _read_classification_upload(uploaded_designation_file, header_keywords=["designation"])
            except Exception as e:
                st.warning(f"⚠️ Could not read Director Designation Classification List ({e}) — "
                           "KPI 2 will read 0 for all segments until a valid file is uploaded.")
                return set(), "no data (upload failed)"

            _designation_col = _find_col(_desig_df, ["designation"])
            if _designation_col is None:
                st.warning(
                    f"⚠️ Director Designation Classification List is missing a 'Designation' column — "
                    f"columns found: {', '.join(str(c) for c in _desig_df.columns)}. "
                    "KPI 2 will read 0 for all segments until a valid file is uploaded."
                )
                return set(), "no data (upload invalid)"

            _designations = set(_desig_df[_designation_col].dropna().astype(str).str.strip().str.lower())
            return _designations, f"uploaded file ({uploaded_designation_file.name}, {len(_designations):,} designations)"

        st.warning("⚠️ No Director Designation Classification List uploaded — KPI 2 (senior trainee days) "
                   "will read 0 for all segments until one is provided in the sidebar.")
        return set(), "no data — upload a Director Designation Classification List"

    SENIOR_DESIGNATIONS, designation_source = _load_designation_classification()

    def segment_breakdown(df_seg):
        by_school = df_seg.groupby('School')['No of Trainee Days'].sum().reset_index() \
            .rename(columns={'No of Trainee Days': 'Trainee Days'}).sort_values('Trainee Days', ascending=False)
        by_month_school = df_seg.groupby(['Month', 'Month Name', 'School'])['No of Trainee Days'].sum().reset_index() \
            .rename(columns={'No of Trainee Days': 'Trainee Days'}).sort_values(['Month', 'School']).drop(columns='Month')
        by_month_course = df_seg.groupby(['Month', 'Month Name', 'School', 'Course name'])['No of Trainee Days'].sum().reset_index() \
            .rename(columns={'No of Trainee Days': 'Trainee Days'}).sort_values(['Month', 'Trainee Days'], ascending=[True, False]).drop(columns='Month')
        return by_school, by_month_school, by_month_course

    # ============================================================
    # NEW-PROGRAMME (KPI 3) HELPERS
    # ============================================================
    def extract_course_code(intake_no):
        """AES2601-260001 -> AES2601"""
        if pd.isna(intake_no):
            return None
        s = str(intake_no).strip()
        return s.split('-')[0].strip() if '-' in s else s

    def is_new_programme_code(course_code):
        if not course_code:
            return False
        return str(course_code).strip().endswith('01')

    # ============================================================
    # SINGLE-BRANCH PIPELINE (>= threshold% attendance only):
    # dedup -> trainee days -> country -> org-name -> school ->
    # designation -> KPI 1 -> KPI 2 -> KPI 3
    # ============================================================
    def run_attendance_branch(df_att_input):
        b = {}

        def _dedup():
            return dedup_by_name_and_email(df_att_input, 'Learner name', 'Learner email address', 'Course intake No.')
        df_clean = safe_step("Deduplication (Attendance)", _dedup)
        b['rows_before_dedup'] = len(df_att_input)
        b['rows_after_dedup'] = len(df_clean)

        def _trainee_days():
            d = df_clean.copy()
            d['Course start date'] = pd.to_datetime(d['Course start date'])
            d['Course end date'] = pd.to_datetime(d['Course end date'])
            d['Attendance % (by session)'] = pd.to_numeric(
                d['Module attendance percentage (by session number)'].str.replace('%', '', regex=False),
                errors='coerce'
            ) / 100
            d['Course duration (days)'] = d.apply(
                lambda r: calculate_course_duration(r['Course start date'], r['Course end date']), axis=1
            )
            # Attendance is all-or-nothing: rows reaching this point already have
            # attendance >= the user-set threshold (the attendance gate applied
            # upstream), so each one counts as 100% attendance — the full course
            # duration in trainee days. Rows below the threshold never reach here
            # at all (excluded upstream), which is equivalent to counting them as
            # 0% attendance / 0 trainee days.
            d['No of Trainee Days'] = d['Course duration (days)'].round(2)
            return d
        df_clean = safe_step("Trainee Days Calculation", _trainee_days)
        b['null_trainee_days'] = int(df_clean['No of Trainee Days'].isna().sum())

        df_clean = safe_step("Country of Organisation Lookup", step_country, df_clean, df_admitted_clean)
        b['country_matched'] = int(df_clean['Country of Organisation'].notna().sum() -
                                    (df_clean['Country of Organisation'] == '').sum())
        b['country_unmatched'] = int(len(df_clean) - b['country_matched'])
        b['country_unfilled'] = df_clean[
            df_clean['Country of Organisation'].isna() | (df_clean['Country of Organisation'] == '')
        ]

        df_clean = safe_step("Organization Name Lookup + Classification",
                              build_org_name_lookup, df_clean, df_admitted_clean, org_class_lookup, ADMITTED_ORG_COL)

        df_clean['Segment (Country)'] = df_clean.apply(classify_segment_country, axis=1)
        df_clean['Segment'] = df_clean['Segment (Org Name)']

        def _school():
            d = df_clean.drop(columns=['School'], errors='ignore')
            d['School'] = d['Course intake No.'].apply(map_school)
            return d
        df_clean = safe_step("School Mapping", _school)
        b['others_intakes'] = df_clean[df_clean['School'] == 'Others']['Course intake No.'].unique()

        df_clean = safe_step("Designation Lookup", step_designation, df_clean, df_admitted_clean)
        b['designation_matched'] = int(df_clean['Designation'].notna().sum())
        b['designation_unmatched'] = int(df_clean['Designation'].isna().sum())

        # ---- month/year prep ----
        df_clean['Course end date'] = pd.to_datetime(df_clean['Course end date'], dayfirst=True, errors='coerce')
        df_clean['Month'] = df_clean['Course end date'].dt.month
        df_clean['Year'] = df_clean['Course end date'].dt.year
        df_clean['Month Name'] = df_clean['Month'].map(month_map)

        df_filtered = df_clean[(df_clean['Year'] == REPORT_YEAR) & (df_clean['Month'].isin(REPORT_MONTHS))].copy()

        # ---- KPI 1 (org-name-based segment) ----
        kpi1_overall = df_filtered.groupby('Segment')['No of Trainee Days'].sum().reset_index() \
            .rename(columns={'No of Trainee Days': 'Trainee Days'}).sort_values('Trainee Days', ascending=False)
        kpi1_intl_total = df_filtered.loc[df_filtered['Segment'] == 'International', 'No of Trainee Days'].sum()
        kpi1_local_total = df_filtered.loc[df_filtered['Segment'] == 'Local', 'No of Trainee Days'].sum()
        kpi1_caas_total = df_filtered.loc[df_filtered['Segment'] == 'CAAS', 'No of Trainee Days'].sum()
        kpi1_unknown_total = df_filtered.loc[df_filtered['Segment'] == 'Unknown', 'No of Trainee Days'].sum()

        # ---- KPI 2 (senior, org-name-based segment) ----
        df2 = df_clean.copy()
        df2['Designation_lower'] = df2['Designation'].str.strip().str.lower()
        df2['Is Senior'] = df2['Designation_lower'].isin(SENIOR_DESIGNATIONS)
        df2_period = df2[(df2['Year'] == REPORT_YEAR) & (df2['Month'].isin(REPORT_MONTHS))]
        df2_filtered = df2_period[df2_period['Is Senior']].copy()

        kpi2_total_rows = df2_period.shape[0]
        kpi2_senior_rows = df2_filtered.shape[0]
        kpi2_unmatched_designations = df2_period[~df2_period['Is Senior']]['Designation'].dropna().str.strip().str.lower().value_counts().reset_index()
        kpi2_unmatched_designations.columns = ['Designation', 'Count']

        kpi2_overall = df2_filtered.groupby('Segment')['No of Trainee Days'].sum().reset_index() \
            .rename(columns={'No of Trainee Days': 'Trainee Days'}).sort_values('Trainee Days', ascending=False)
        kpi2_intl_total = df2_filtered.loc[df2_filtered['Segment'] == 'International', 'No of Trainee Days'].sum()
        kpi2_local_total = df2_filtered.loc[df2_filtered['Segment'] == 'Local', 'No of Trainee Days'].sum()
        kpi2_caas_total = df2_filtered.loc[df2_filtered['Segment'] == 'CAAS', 'No of Trainee Days'].sum()
        kpi2_unknown_total = df2_filtered.loc[df2_filtered['Segment'] == 'Unknown', 'No of Trainee Days'].sum()

        # ---- KPI 3: New Programmes by Month (course code ends in '01') ----
        df3 = df_filtered.copy()
        df3['Course Code'] = df3['Course intake No.'].apply(extract_course_code)
        df3['Is New Programme'] = df3['Course Code'].apply(is_new_programme_code)
        df3_new = df3[df3['Is New Programme']].copy()

        kpi3_monthly = df3_new.drop_duplicates(subset=['Month', 'Course intake No.']) \
            .groupby(['Month', 'Month Name'])['Course intake No.'].nunique().reset_index() \
            .rename(columns={'Course intake No.': 'New Programmes'}).sort_values('Month').drop(columns='Month')
        kpi3_total_new_programmes = df3_new['Course intake No.'].nunique()

        course_name_col = 'Course name' if 'Course name' in df3_new.columns else None
        date_cols = [c for c in ['Course start date', 'Course end date'] if c in df3_new.columns]
        unique_cols = ['Course Code', 'Course intake No.', 'School'] + ([course_name_col] if course_name_col else []) + date_cols
        kpi3_unique_courses = df3_new[unique_cols].drop_duplicates().sort_values('Course intake No.')

        # ---- QA data (kept for the Reference & Classification tab) ----
        qa_cols = ['Learner name', 'Course intake No.', 'Learner email address',
                   'Country of Organisation', 'Organization Name', 'Segment (Country)',
                   'Segment (Org Name)', 'Org Name Needs Review', 'Org Name Match Source', 'Org Name Match Reason']
        qa_cols = [c for c in qa_cols if c in df_clean.columns]
        qa_df = df_clean[qa_cols].copy()

        b.update(dict(
            df_clean=df_clean, df_filtered=df_filtered, df2_filtered=df2_filtered,
            kpi1_overall=kpi1_overall, kpi1_intl_total=kpi1_intl_total, kpi1_local_total=kpi1_local_total, kpi1_caas_total=kpi1_caas_total, kpi1_unknown_total=kpi1_unknown_total,
            kpi2_overall=kpi2_overall, kpi2_intl_total=kpi2_intl_total, kpi2_local_total=kpi2_local_total, kpi2_caas_total=kpi2_caas_total, kpi2_unknown_total=kpi2_unknown_total,
            kpi2_total_rows=kpi2_total_rows, kpi2_senior_rows=kpi2_senior_rows, kpi2_unmatched_designations=kpi2_unmatched_designations,
            kpi3_monthly=kpi3_monthly, kpi3_total_new_programmes=kpi3_total_new_programmes, kpi3_unique_courses=kpi3_unique_courses,
            qa_df=qa_df,
        ))
        return b

    branch = safe_step(f"Run '≥ {attendance_threshold}%' Attendance Branch", run_attendance_branch, df_attendance_gated)

    # ---- Convenience bindings ----
    df_attendance_clean = branch['df_clean']
    df_filtered = branch['df_filtered']
    df2_filtered = branch['df2_filtered']
    kpi1_overall = branch['kpi1_overall']
    kpi1_intl_total = branch['kpi1_intl_total']
    kpi1_local_total = branch['kpi1_local_total']
    kpi1_caas_total = branch['kpi1_caas_total']
    kpi1_unknown_total = branch['kpi1_unknown_total']
    kpi2_overall = branch['kpi2_overall']
    kpi2_intl_total = branch['kpi2_intl_total']
    kpi2_local_total = branch['kpi2_local_total']
    kpi2_caas_total = branch['kpi2_caas_total']
    kpi2_unknown_total = branch['kpi2_unknown_total']
    kpi2_total_rows = branch['kpi2_total_rows']
    kpi2_senior_rows = branch['kpi2_senior_rows']
    kpi2_unmatched_designations = branch['kpi2_unmatched_designations']
    kpi3_monthly = branch['kpi3_monthly']
    kpi3_total_new_programmes = branch['kpi3_total_new_programmes']
    kpi3_unique_courses = branch['kpi3_unique_courses']
    null_trainee_days = branch['null_trainee_days']
    country_unfilled = branch['country_unfilled']
    country_matched = branch['country_matched']
    country_unmatched = branch['country_unmatched']
    others_intakes = branch['others_intakes']
    designation_matched = branch['designation_matched']
    designation_unmatched = branch['designation_unmatched']

    # ============================================================
    # BLOCK 3: Data Quality Checks
    # ============================================================
    def is_email(val):
        return bool(re.match(r'^[\w\.\+\-]+@[\w\.\-]+\.\w+$', str(val).strip()))

    def run_dq_checks(df, label, name_col, email_col, id_col):
        def _subset(d):
            out = pd.DataFrame(index=d.index)
            out['Source'] = label
            if name_col in d.columns:
                out['Learner Name'] = d[name_col]
            if 'Course intake No.' in d.columns:
                out['Course intake No.'] = d['Course intake No.']
            if email_col in d.columns:
                out['Email'] = d[email_col]
            if id_col in d.columns:
                out['Learner ID'] = d[id_col]
            return out.reset_index(drop=True)

        email_as_name = df[df[name_col].apply(is_email)]
        missing_email = df[df[email_col].isna() | (df[email_col].astype(str).str.strip() == '')]
        email_as_id = df[df[id_col].apply(is_email)]
        missing_id = df[df[id_col].isna() | (df[id_col].astype(str).str.strip() == '')]
        duplicate_rows = df[df.duplicated(keep=False)]
        counts = {
            'Data Source': label,
            'Email as Name': email_as_name[name_col].nunique(),
            'Missing Email': missing_email[name_col].nunique(),
            'Email as ID': email_as_id[name_col].nunique(),
            'Missing ID': missing_id[name_col].nunique(),
            'Duplicate Rows': len(duplicate_rows),
        }
        detail = {
            'Email as Name': _subset(email_as_name),
            'Missing Email': _subset(missing_email),
            'Email as ID': _subset(email_as_id),
            'Missing ID': _subset(missing_id),
            'Duplicate Rows': _subset(duplicate_rows),
        }
        return counts, detail

    def step_dq():
        counts_list = []
        detail_lists = {'Email as Name': [], 'Missing Email': [], 'Email as ID': [], 'Missing ID': [], 'Duplicate Rows': []}
        for df, label, name_col, email_col, id_col in [
            (df_attendance_clean, 'Attendance List (data cleaned)', 'Learner name', 'Learner email address', 'Learner ID'),
            (df_admitted_clean, 'Admitted List (data cleaned)', 'Learner name', 'Email address', 'Learner ID'),
            (df_cert_clean, 'Certificate List (data cleaned)', 'Learner name', 'Email Address', 'Learner ID'),
        ]:
            counts, detail = run_dq_checks(df, label, name_col, email_col, id_col)
            counts_list.append(counts)
            for k in detail_lists:
                detail_lists[k].append(detail[k])
        detail_combined = {
            k: (pd.concat(v, ignore_index=True) if v else pd.DataFrame()) for k, v in detail_lists.items()
        }
        return counts_list, detail_combined

    dq_results, dq_detail = safe_step("Data Quality Checks", step_dq)
    summary_dq = pd.DataFrame(dq_results)

    # ============================================================
    # Duplicate detection preview (across raw files)
    # ============================================================
    def step_dedup_preview():
        def _sequential_dedup_display(df, passes, source_label, email_col=None):
            """Mirrors the real sequential dedup logic (pass 1 by name+intake,
            pass 2 by email+intake among survivors of pass 1), applying the
            same 'prefer official email over personal email' tiebreak, and
            returns ONE row per physical record — not one row per check — so
            a learner who matches on both name and email doesn't appear as
            four separate duplicate records."""
            d = df.copy()
            if email_col is not None and email_col in d.columns:
                d['_is_personal_email'] = d[email_col].apply(is_personal_email)
                d = d.sort_values('_is_personal_email', kind='stable')

            active = pd.Series(True, index=d.index)
            status = pd.Series('Kept', index=d.index)
            dup_key = pd.Series('', index=d.index)
            included = pd.Series(False, index=d.index)

            for cols in passes:
                key_label = ' + '.join(cols)
                active_idx = d.index[active]
                sub = d.loc[active_idx]
                dup_all = sub.duplicated(subset=cols, keep=False)
                dup_removed = sub.duplicated(subset=cols, keep='first')
                grouped_idx = active_idx[dup_all]
                newly_grouped = grouped_idx[~included.loc[grouped_idx]]
                dup_key.loc[newly_grouped] = key_label
                included.loc[grouped_idx] = True
                removed_idx = active_idx[dup_removed]
                status.loc[removed_idx] = 'Removed'
                active.loc[removed_idx] = False

            d['Status'] = status
            d['Duplicate Key'] = dup_key
            d['Data Source'] = source_label
            return d[included].copy()

        admitted_dups = _sequential_dedup_display(
            df_admitted, [['Learner name', 'Course intake No.'], ['Email address', 'Course intake No.']],
            'Admitted List', email_col='Email address'
        )

        enrollment_dups = _sequential_dedup_display(df_enrollment, [['Course intake No.']], 'Enrollment List')
        enrollment_dups['Learner name'] = None
        enrollment_dups['Email address'] = None

        attendance_dups = _sequential_dedup_display(
            df_attendance_gated.rename(columns={'Learner email address': 'Email address'}),
            [['Learner name', 'Course intake No.'], ['Email address', 'Course intake No.']],
            'Attendance List (>= threshold)', email_col='Email address'
        )

        cert_dups = _sequential_dedup_display(df_cert, [['Learner name', 'Course intake No.']], 'Certificate List')

        dup_cols = ['Data Source', 'Duplicate Key', 'Course intake No.', 'Learner name', 'Email address', 'Status']
        for d in [admitted_dups, enrollment_dups, attendance_dups, cert_dups]:
            for c in dup_cols:
                if c not in d.columns:
                    d[c] = None

        all_dups_display = pd.concat([
            admitted_dups[dup_cols], enrollment_dups[dup_cols],
            attendance_dups[dup_cols], cert_dups[dup_cols],
        ], ignore_index=True).sort_values(
            ['Data Source', 'Duplicate Key', 'Course intake No.', 'Learner name', 'Status']
        ).reset_index(drop=True)

        dup_removed_summary = (
            all_dups_display[all_dups_display['Status'] == 'Removed']
            .groupby(['Data Source', 'Duplicate Key']).size().reset_index(name='Rows Removed')
        )
        return dup_removed_summary, all_dups_display

    dup_removed_summary, all_duplicates_display = safe_step("Duplicate Detection", step_dedup_preview)

    summary_dedup = pd.DataFrame({
        'Data Source': ['Admitted List', 'Enrollment List', 'Attendance List (>= threshold)', 'Certificate List'],
        'Trainee Count (before data clean)': [len(df_admitted), len(df_enrollment), len(df_attendance_gated), len(df_cert)],
        'Trainee Count (after data clean)': [len(df_admitted_clean), len(df_enrollment_clean), len(df_attendance_clean), len(df_cert_clean)],
    })
    summary_dedup['Rows Removed (during data clean)'] = (
        summary_dedup['Trainee Count (before data clean)'] - summary_dedup['Trainee Count (after data clean)']
    )

    # ============================================================
    # BLOCK 4: Overview Summary + Enrolment Funnel Check
    # ============================================================
    def step_overview():
        total_unique_intakes = pd.concat([
            df_admitted_clean['Course intake No.'], df_enrollment_clean['Course intake No.'],
            df_attendance_clean['Course intake No.'], df_cert_clean['Course intake No.']
        ]).nunique()

        summary_overview = pd.DataFrame({
            'Metric': ['Unique Course Intake No.', 'Total Learners (Admitted)', 'Total Learners (Enrolled)',
                       'Total Learners (Attended, >= threshold)', 'Total Learners (Cert Issued)'],
            'Count': [total_unique_intakes, len(df_admitted_clean), df_enrollment_clean['No. of enrollments'].sum(),
                      len(df_attendance_clean), len(df_cert_clean)]
        })

        # ---- Funnel check, per Course Intake No.: ----
        #   Admitted   >= Enrolled     (every enrolled learner should have been admitted)
        #   Enrolled   >= Attended     (every attendee should have been enrolled)
        #   Attended   >= Certificate  (every certificate should go to someone who attended)
        # Only intakes that break one or more of these rules are "Flagged", and the
        # specific rule(s) broken are named in 'Offending Rule(s)'.
        admitted_counts = df_admitted_clean.groupby('Course intake No.').size().reset_index(name='Admitted')
        enrolled_counts = df_enrollment_clean[['Course intake No.', 'No. of enrollments']].rename(
            columns={'No. of enrollments': 'Enrolled'})
        attended_counts = df_attendance_clean.groupby('Course intake No.').size().reset_index(name='Attended')
        cert_counts = df_cert_clean.groupby('Course intake No.').size().reset_index(name='Certificate')

        discrepancy = (
            admitted_counts
            .merge(enrolled_counts, on='Course intake No.', how='outer')
            .merge(attended_counts, on='Course intake No.', how='outer')
            .merge(cert_counts, on='Course intake No.', how='outer')
            .fillna(0)
        )
        for col in ['Admitted', 'Enrolled', 'Attended', 'Certificate']:
            discrepancy[col] = discrepancy[col].astype(int)

        def _offending_rules(row):
            issues = []
            if row['Admitted'] < row['Enrolled']:
                issues.append('Admitted < Enrolled')
            if row['Enrolled'] < row['Attended']:
                issues.append('Enrolled < Attended')
            if row['Attended'] < row['Certificate']:
                issues.append('Attended < Certificate')
            return '; '.join(issues)

        discrepancy['Offending Rule(s)'] = discrepancy.apply(_offending_rules, axis=1)
        discrepancy['Flagged'] = discrepancy['Offending Rule(s)'] != ''
        discrepancy = discrepancy.sort_values(
            ['Flagged', 'Course intake No.'], ascending=[False, True]
        ).reset_index(drop=True)

        return summary_overview, discrepancy

    summary_overview, discrepancy = safe_step("Overview Summary", step_overview)

    # ============================================================
    # BLOCK 5.5: Non-Course Events Check
    # ============================================================
    def step_non_course_events():
        target = ['CAS-0001', 'AMS-0001', 'AMS-0002']
        df_check = df_namelist[df_namelist['Course Intake Number'].isin(target)].copy()

        # Apply the same reporting-period date filter used elsewhere, if the
        # Namelist has a recognisable start-date column for these rows.
        nce_start_date_col = _find_namelist_col(['course start date', 'start date'], df_check.columns)
        if nce_start_date_col:
            df_check[nce_start_date_col] = pd.to_datetime(df_check[nce_start_date_col], dayfirst=True, errors='coerce')
            df_check = df_check[(df_check[nce_start_date_col] >= start_date) & (df_check[nce_start_date_col] <= end_date)]

        df_check['Type'] = df_check['Country'].apply(lambda x: 'Local' if str(x).strip().lower() == 'singapore' else 'International')
        headcount = df_check.groupby(['Course Intake Number', 'Type']).size().unstack(fill_value=0).reset_index()
        trainee_days = df_check.groupby(['Course Intake Number', 'Type'])['No. of Trainee Days'].sum().unstack(fill_value=0).reset_index()

        non_course_local_total = float(trainee_days['Local'].sum()) if 'Local' in trainee_days.columns else 0.0
        non_course_intl_total = float(trainee_days['International'].sum()) if 'International' in trainee_days.columns else 0.0

        return headcount, trainee_days, non_course_local_total, non_course_intl_total, nce_start_date_col

    non_course_headcount, non_course_days, non_course_local_total, non_course_intl_total, _nce_start_date_col = safe_step(
        "Non-Course Events Check", step_non_course_events
    )

    # ============================================================
    # BLOCK 6: Namelist-Only KPI Cross-Check
    # ============================================================
    # Interim-measure support: while Course Managers double-key data into
    # both the Namelist and Learn@SAA, this recomputes KPI 1 / KPI 2 / KPI 3
    # using ONLY the Namelist file — completely independent of the
    # Admitted/Enrollment/Attendance/Certificate pipeline used everywhere
    # else in this app — so the two can be reconciled side by side. Large
    # gaps usually mean the Namelist and Learn@SAA have drifted apart for
    # some intakes, which is exactly the gap this app is meant to surface
    # on the way to Learn@SAA becoming the single source of truth.
    def compute_namelist_kpis():
        df_nl = df_namelist.copy()
        notes = []

        org_col = _find_namelist_col(
            ['organisation', 'organization', 'organisation name', 'organization name', 'company', 'company name'],
            df_nl.columns
        )
        designation_col = _find_namelist_col(
            ['designation', 'position', 'job title', 'title', 'rank'], df_nl.columns
        )
        start_date_col = _find_namelist_col(['course start date', 'start date'], df_nl.columns)

        df_nl['No. of Trainee Days'] = pd.to_numeric(df_nl['No. of Trainee Days'], errors='coerce').fillna(0)

        # ---- Date filter (reporting period), if the column exists ----
        if start_date_col:
            df_nl[start_date_col] = pd.to_datetime(df_nl[start_date_col], dayfirst=True, errors='coerce')
            before_n = len(df_nl)
            df_nl = df_nl[(df_nl[start_date_col] >= start_date) & (df_nl[start_date_col] <= end_date)]
            notes.append(f"Filtered to the reporting period using '{start_date_col}': {len(df_nl):,} of {before_n:,} Course Reporting Namelist rows kept.")
        else:
            notes.append("No 'Course Start Date' column found in the Course Reporting Namelist — the reporting-period date filter "
                          "could NOT be applied here; all Course Reporting Namelist rows are included regardless of date.")

        # ---- Classification: org-name table if possible, else Country fallback ----
        if org_col:
            results = df_nl[org_col].apply(lambda x: classify_org_name(x, org_class_lookup))
            df_nl['Segment (Course Reporting Namelist)'] = results.apply(lambda r: r[0])
            notes.append(f"Classified using the Course Reporting Namelist's '{org_col}' column against the same "
                         "Organization Name Classification table used in the Results tab.")
        else:
            def _country_fallback(country):
                c = str(country).strip().lower()
                if c in ('', 'nan'):
                    return 'Unknown'
                return 'Local' if c == 'singapore' else 'International'
            df_nl['Segment (Course Reporting Namelist)'] = df_nl['Country'].apply(_country_fallback)
            notes.append("No organisation-name column found in the Course Reporting Namelist — classification fell back to the "
                         "'Country' column (Singapore → Local, else International). CAAS cannot be separately "
                         "identified this way, so CAAS trainee days from the Course Reporting Namelist will read as 0.")

        # ---- KPI 1: trainee days by segment ----
        kpi1_nl = df_nl.groupby('Segment (Course Reporting Namelist)')['No. of Trainee Days'].sum()
        kpi1_nl_intl = float(kpi1_nl.get('International', 0))
        kpi1_nl_local = float(kpi1_nl.get('Local', 0))
        kpi1_nl_caas = float(kpi1_nl.get('CAAS', 0))
        kpi1_nl_unknown = float(kpi1_nl.get('Unknown', 0))

        # ---- KPI 2: senior trainee days, only if a designation-like column exists ----
        kpi2_available = designation_col is not None
        if kpi2_available:
            designation_lower = df_nl[designation_col].astype(str).str.strip().str.lower()
            is_senior = designation_lower.isin(SENIOR_DESIGNATIONS)
            kpi2_nl = df_nl[is_senior].groupby('Segment (Course Reporting Namelist)')['No. of Trainee Days'].sum()
            kpi2_nl_intl = float(kpi2_nl.get('International', 0))
            kpi2_nl_local = float(kpi2_nl.get('Local', 0))
            kpi2_nl_caas = float(kpi2_nl.get('CAAS', 0))
            notes.append(f"Senior learners identified using the Course Reporting Namelist's '{designation_col}' column, "
                         "matched against the same senior-designation list used in the Results tab.")
        else:
            kpi2_nl_intl = kpi2_nl_local = kpi2_nl_caas = None
            notes.append("No designation/position column found in the Course Reporting Namelist — KPI 2 cannot be computed "
                         "from Course Reporting Namelist data alone.")

        # ---- KPI 3: new programmes (course code ends in '01') ----
        df_nl['_course_code'] = df_nl['Course Intake Number'].apply(extract_course_code)
        df_nl['_is_new_programme'] = df_nl['_course_code'].apply(is_new_programme_code)
        kpi3_nl_total = df_nl.loc[df_nl['_is_new_programme'], 'Course Intake Number'].nunique()

        return dict(
            df_nl=df_nl, notes=notes, org_col=org_col, designation_col=designation_col, start_date_col=start_date_col,
            kpi1_nl_intl=kpi1_nl_intl, kpi1_nl_local=kpi1_nl_local, kpi1_nl_caas=kpi1_nl_caas, kpi1_nl_unknown=kpi1_nl_unknown,
            kpi2_available=kpi2_available, kpi2_nl_intl=kpi2_nl_intl, kpi2_nl_local=kpi2_nl_local, kpi2_nl_caas=kpi2_nl_caas,
            kpi3_nl_total=kpi3_nl_total,
        )

    namelist_kpis = safe_step("Course Reporting Namelist KPI Cross-Check", compute_namelist_kpis)


# ────────────────────────────────────────────────────────────────
# QUICK SUMMARY BANNER
# ────────────────────────────────────────────────────────────────
st.success("✅ Pipeline completed successfully")
st.caption(
    f"Organization-name classification source: {org_class_source}. {org_name_source_note} "
    f"Certificate file detected as: {cert_format_label}. Course Reporting Namelist sheet used: '{namelist_sheet_used}'."
)

st.markdown(f"### 🎯 Quick Summary — ≥ {attendance_threshold}% Attendance")
b1, b3, b4 = st.columns(3)
b1.metric("KPI 1 — Intl Trainee Days", f"{kpi1_intl_total:,.0f}",
          f"{(kpi1_intl_total / KPI1_TARGET) * 100:.1f}% of target", delta_color="normal")
b3.metric("KPI 2 — Intl Trainee Days for Directors and above", f"{kpi2_intl_total:,.0f}",
          f"{(kpi2_intl_total / KPI2_TARGET) * 100:.1f}% of target", delta_color="normal")
b4.metric("KPI 3 — New Programmes", f"{kpi3_total_new_programmes:,.0f}",
          f"{(kpi3_total_new_programmes / KPI3_TARGET) * 100:.1f}% of target", delta_color="normal")

if null_trainee_days > 0:
    st.warning(f"⚠️ {null_trainee_days} rows have missing trainee day calculations — check for blank dates in your Attendance List.")
if len(country_unfilled) > 5:
    st.warning(f"⚠️ {len(country_unfilled)} rows have no Country of Organisation — higher than usual, worth checking.")

st.divider()


# ────────────────────────────────────────────────────────────────
# DOWNLOAD BUTTON — same CAAS styling as the Discrepancy Checker's export
# ────────────────────────────────────────────────────────────────
_NAVY = "192F55"
_GREY_BORDER = "BFBFBF"
_RED = "C00000"
_BLACK = "000000"

_thin_side = Side(style="thin", color=_GREY_BORDER)
_thin_border = Border(left=_thin_side, right=_thin_side, top=_thin_side, bottom=_thin_side)
_header_fill = PatternFill(start_color=_NAVY, end_color=_NAVY, fill_type="solid")
_header_font = Font(name="Arial", bold=True, color="FFFFFF")
_body_font = Font(name="Arial", color=_BLACK)
_red_font = Font(name="Arial", bold=True, color=_RED)


def _style_export_sheet(ws, df, red_cols=None, red_if_nonempty_cols=None):
    """Applies the navy header / thin grey grid / autofilter / frozen header
    styling to a worksheet already populated via df.to_excel(), matching the
    Discrepancy Checker's export. red_cols: 0-indexed numeric columns that
    turn red+bold when non-zero. red_if_nonempty_cols: 0-indexed text columns
    that turn red+bold when non-empty (used for problem/flag descriptions)."""
    red_cols = red_cols or []
    red_if_nonempty_cols = red_if_nonempty_cols or []
    n_rows, n_cols = df.shape

    for c in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = _header_font
        cell.fill = _header_fill
        cell.border = _thin_border
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for r in range(2, n_rows + 2):
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = _thin_border
            value = cell.value
            flagged = False
            if (c - 1) in red_cols and isinstance(value, (int, float)) and value != 0:
                flagged = True
            if (c - 1) in red_if_nonempty_cols and isinstance(value, str) and value.strip() != '':
                flagged = True
            cell.font = _red_font if flagged else _body_font
            if isinstance(value, (int, float)):
                cell.alignment = Alignment(horizontal="center")

    last_col_letter = get_column_letter(n_cols)
    ws.auto_filter.ref = f"A1:{last_col_letter}{n_rows + 1}"
    ws.freeze_panes = "A2"


output = io.BytesIO()
with pd.ExcelWriter(output, engine='openpyxl') as writer:
    kpi1_overall.to_excel(writer, sheet_name="KPI1_TraineeDays", index=False)
    kpi2_overall.to_excel(writer, sheet_name="KPI2_SeniorDays", index=False)
    kpi3_monthly.to_excel(writer, sheet_name="KPI3_NewProgrammes", index=False)
    kpi3_unique_courses.to_excel(writer, sheet_name="KPI3_UniqueCourses", index=False)
    summary_dq.to_excel(writer, sheet_name="Data_Quality", index=False)
    summary_dedup.to_excel(writer, sheet_name="Dedup_Summary", index=False)
    discrepancy.to_excel(writer, sheet_name="Enrolment_Funnel_Check", index=False)

    # Data Quality: every count column represents a problem — red if non-zero.
    _style_export_sheet(
        writer.sheets["Data_Quality"], summary_dq,
        red_cols=list(range(1, len(summary_dq.columns))),
    )
    # Dedup Summary: only "Rows Removed" is worth flagging.
    _style_export_sheet(
        writer.sheets["Dedup_Summary"], summary_dedup,
        red_cols=[summary_dedup.columns.get_loc("Rows Removed (during data clean)")],
    )
    # Enrolment Funnel Check: red the "Offending Rule(s)" text when non-empty.
    _style_export_sheet(
        writer.sheets["Enrolment_Funnel_Check"], discrepancy,
        red_if_nonempty_cols=[discrepancy.columns.get_loc("Offending Rule(s)")],
    )
    # The remaining sheets (KPI totals, course lists) aren't "problem" tables —
    # style them plainly (navy header + grid) with no red column.
    for _sheet_name, _df in [
        ("KPI1_TraineeDays", kpi1_overall), ("KPI2_SeniorDays", kpi2_overall),
        ("KPI3_NewProgrammes", kpi3_monthly), ("KPI3_UniqueCourses", kpi3_unique_courses),
    ]:
        _style_export_sheet(writer.sheets[_sheet_name], _df)

st.download_button(
    label="⬇️  Download Full Report (Excel)",
    data=output.getvalue(),
    file_name=f"SAA_Pipeline_Output_{datetime.date.today().strftime('%Y%m%d')}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True,
)

st.divider()


# ────────────────────────────────────────────────────────────────
# TABS
# ────────────────────────────────────────────────────────────────
tab_results, tab_namelist, tab_dq, tab_reference, tab_advanced = st.tabs([
    f"📊 Results (≥{attendance_threshold}%)", "🔀 KPI Cross-check with Course Reporting Namelist", "🔎 Data Quality",
    "📚 Reference & Classification", "📄 Data Preview"
])

# ── RESULTS TAB ────────────────────────────────────────────────
with tab_results:
    st.caption(
        f"Reporting period: **{start_date:%d %b %Y} – {end_date:%d %b %Y}** "
        f"(courses filtered by this date range). Rows with attendance % (by session) ≥ {attendance_threshold}%. "
        f"Segment = org-name-based classification."
    )

    # --- KPI 1 ---
    st.subheader("KPI 1 — International Trainee Days")
    st.caption(f"KPI 1 annual target: **{KPI1_TARGET:,.0f} trainee days**")
    k1_1, k1_2, k1_3, k1_4 = st.columns(4)
    k1_1.metric("KPI 1 International", f"{kpi1_intl_total:,.0f}",
                f"{(kpi1_intl_total / KPI1_TARGET) * 100:.1f}% of target", delta_color="normal")
    k1_2.metric("Local", f"{kpi1_local_total:,.0f}")
    k1_3.metric("CAAS", f"{kpi1_caas_total:,.0f}")
    k1_4.metric("⚠️ Unknown", f"{kpi1_unknown_total:,.0f}")
    st.dataframe(kpi1_overall, use_container_width=True)

    def show_unknown_learners_by_course(seg_df, by_month_course, key_prefix):
        """For the Unknown segment: under the By Month and Course table, give
        each course its own toggle to reveal just that course's learners.
        Uses individual checkboxes (one per course) rather than a nested
        expander (Streamlit doesn't allow expanders inside expanders) and
        rather than one combined dropdown, so only the course(s) you
        actually open get filtered and rendered."""
        st.write("**Learners by Course** — tick a course to view its learners")
        if len(by_month_course) == 0:
            st.info("No courses to show.")
            return
        for i, course_row in by_month_course.reset_index(drop=True).iterrows():
            label = f"👤 {course_row['Month Name']} · {course_row['Course name']} ({course_row['School']})"
            show_it = st.checkbox(label, key=f"{key_prefix}_unknown_course_{i}")
            if show_it:
                course_mask = (
                    (seg_df['Month Name'] == course_row['Month Name']) &
                    (seg_df['School'] == course_row['School']) &
                    (seg_df['Course name'] == course_row['Course name'])
                )
                course_learners = seg_df.loc[course_mask, ['Learner name', 'Learner ID', 'Learner email address']]
                st.dataframe(course_learners, use_container_width=True)

    for label, total in [("International", kpi1_intl_total), ("Local", kpi1_local_total), ("CAAS", kpi1_caas_total), ("Unknown", kpi1_unknown_total)]:
        icon = "⚠️" if label == "Unknown" else "📂"
        with st.expander(f"{icon} {label} — detail breakdown"):
            seg_df = df_filtered[df_filtered['Segment'] == label]
            if len(seg_df) == 0:
                st.info(f"No {label} records in this reporting period.")
            else:
                if label == "Unknown":
                    st.warning(
                        f"⚠️ {seg_df['Learner name'].nunique():,} learner(s) could not be classified "
                        "(usually a blank organization name). Review these rows and either fix the "
                        "organization name in your source file, or add it to the classification table."
                    )
                by_school, by_month_school, by_month_course = segment_breakdown(seg_df)
                st.write(f"Total: **{total:,.2f} trainee days**")
                st.write("**By School**")
                st.dataframe(by_school, use_container_width=True)
                st.write("**By Month and School**")
                st.dataframe(by_month_school, use_container_width=True)
                st.write("**By Month and Course**")
                st.dataframe(by_month_course, use_container_width=True)
                if label == "Unknown":
                    show_unknown_learners_by_course(seg_df, by_month_course, "kpi1")

    st.divider()

    # --- KPI 2 ---
    st.subheader("KPI 2 — International Trainee Days for Director & Above")
    st.caption(f"KPI 2 annual target: **{KPI2_TARGET:,.0f} international senior trainee days**")
    k2_1, k2_2, k2_3, k2_4 = st.columns(4)
    k2_1.metric("International Trainee Days", f"{kpi2_intl_total:,.0f}",
                f"{(kpi2_intl_total / KPI2_TARGET) * 100:.1f}% of target", delta_color="normal")
    k2_2.metric("Local Trainee Days", f"{kpi2_local_total:,.0f}")
    k2_3.metric("CAAS Trainee Days", f"{kpi2_caas_total:,.0f}")
    k2_4.metric("⚠️ Unknown Senior Days", f"{kpi2_unknown_total:,.0f}")
    st.dataframe(kpi2_overall, use_container_width=True)

    for label in ["International", "Local", "CAAS", "Unknown"]:
        icon = "⚠️" if label == "Unknown" else "📂"
        with st.expander(f"{icon} {label} — senior detail breakdown"):
            seg_df = df2_filtered[df2_filtered['Segment'] == label]
            if len(seg_df) == 0:
                st.info(f"No senior {label} records in this reporting period.")
            else:
                if label == "Unknown":
                    st.warning(
                        f"⚠️ {seg_df['Learner name'].nunique():,} senior learner(s) could not be classified "
                        "(usually a blank organization name). Review these rows and either fix the "
                        "organization name in your source file, or add it to the classification table."
                    )
                by_school, by_month_school, by_month_course = segment_breakdown(seg_df)
                st.write("**By School**")
                st.dataframe(by_school, use_container_width=True)
                st.write("**By Month and School**")
                st.dataframe(by_month_school, use_container_width=True)
                st.write("**By Month and Course**")
                st.dataframe(by_month_course, use_container_width=True)
                if label == "Unknown":
                    show_unknown_learners_by_course(seg_df, by_month_course, "kpi2")

    st.divider()

    # --- KPI 3 ---
    st.subheader("KPI 3 — New Programmes by Month")
    st.caption("A programme is treated as new when its course code (the part of the Course intake No. before the hyphen) ends in '01'.")
    k3_1, k3_2 = st.columns(2)
    k3_1.metric("Total New Programmes", f"{kpi3_total_new_programmes:,.0f}",
                f"{(kpi3_total_new_programmes / KPI3_TARGET) * 100:.1f}% of target", delta_color="normal")
    k3_2.metric("KPI 3 Target", f"{KPI3_TARGET:,.0f}")
    st.write("**New Programmes by Month**")
    st.dataframe(kpi3_monthly, use_container_width=True)
    with st.expander("📂 Unique new programme courses"):
        st.dataframe(kpi3_unique_courses, use_container_width=True)

    st.divider()

    # --- Non-Course Events ---
    st.subheader("Non-Course Events (from Course Reporting Namelist)")
    st.write("**Headcount by Intake and Type**")
    st.dataframe(non_course_headcount, use_container_width=True)
    st.write("**Trainee Days by Intake and Type**")
    st.dataframe(non_course_days, use_container_width=True)


# ── DATA QUALITY TAB ──────────────────────────────────────────
with tab_dq:
    st.caption(f"Based on the ≥ {attendance_threshold}% attendance branch.")

    # 6.1 Overview Summary
    st.subheader("Overview Summary")
    st.dataframe(summary_overview, use_container_width=True)
    with st.expander("🔍 View duplicate records found"):
        if len(all_duplicates_display) > 0:
            st.caption("Every row involved in a duplicate group is shown, with a **Status** of Kept or Removed "
                       "(the first occurrence is kept; later occurrences are removed during data cleaning).")
            st.dataframe(all_duplicates_display, use_container_width=True)
            st.write("**Duplicates removed, by list**")
            if len(dup_removed_summary) > 0:
                st.dataframe(dup_removed_summary, use_container_width=True)
            else:
                st.success("No rows were removed as duplicates.")
        else:
            st.success("No duplicate records found.")

    st.divider()

    # 6.2 Data Inconsistency
    st.subheader("Data Inconsistency")
    st.caption("Email as Name / Missing Email / Email as ID / Missing ID / Duplicate Rows, by distinct learner.")
    st.dataframe(summary_dq, use_container_width=True)

    for _dq_label in ['Email as Name', 'Missing Email', 'Email as ID', 'Missing ID', 'Duplicate Rows']:
        _detail_df = dq_detail.get(_dq_label, pd.DataFrame())
        with st.expander(f"🔍 {_dq_label} ({len(_detail_df)} rows) — view and rectify"):
            if len(_detail_df) == 0:
                st.success("No rows found.")
            else:
                st.dataframe(_detail_df, use_container_width=True)

    st.divider()

    # 6.2b Enrolment Funnel Check
    st.subheader("Enrolment Funnel Check")
    st.caption("Per course intake, each stage should have at least as many learners as the next: "
               "**Admitted ≥ Enrolled ≥ Attended ≥ Certificate**. Only intakes that break one or more "
               "of these rules are flagged below, along with exactly which rule(s) they broke.")
    _intakes_flagged = discrepancy[discrepancy['Flagged']]
    e1, e2, e3 = st.columns(3)
    e1.metric("Total Course Intakes", f"{len(discrepancy):,}")
    e2.metric("Intakes OK", f"{int((~discrepancy['Flagged']).sum()):,}")
    e3.metric("Intakes Flagged", f"{len(_intakes_flagged):,}")
    with st.expander(f"🔍 View flagged intakes ({len(_intakes_flagged)}) — view and rectify"):
        if len(_intakes_flagged) == 0:
            st.success("No flagged intakes — Admitted ≥ Enrolled ≥ Attended ≥ Certificate holds for every intake.")
        else:
            st.dataframe(
                _intakes_flagged[['Course intake No.', 'Admitted', 'Enrolled', 'Attended', 'Certificate', 'Offending Rule(s)']],
                use_container_width=True
            )
    with st.expander("📂 View full funnel table (all intakes)"):
        st.dataframe(
            discrepancy[['Course intake No.', 'Admitted', 'Enrolled', 'Attended', 'Certificate', 'Offending Rule(s)']],
            use_container_width=True
        )

    st.divider()

    # 6.3 Identifying Trainee Designation
    st.subheader("Identifying Trainee Designation")
    d1, d2 = st.columns(2)
    d1.metric("Matched Rows", f"{designation_matched:,}")
    d2.metric("Unmatched Rows", f"{designation_unmatched:,}")
    unmatched_designation_df = df_attendance_clean[df_attendance_clean['Designation'].isna()]
    unmatched_designation_cols = [c for c in ['Learner name', 'Learner ID', 'Course intake No.', 'Learner email address', 'Designation']
                                   if c in unmatched_designation_df.columns]
    with st.expander(f"🔍 View unmatched designation rows ({designation_unmatched}) — view and rectify"):
        if designation_unmatched == 0:
            st.success("No unmatched designation rows.")
        else:
            st.dataframe(unmatched_designation_df[unmatched_designation_cols], use_container_width=True)

    st.divider()

    # 6.4 Identifying Trainee Country of Organisation
    st.subheader("Identifying Trainee Country of Organisation")
    f1, f2 = st.columns(2)
    f1.metric("Matched Rows", f"{country_matched:,}")
    f2.metric("Unmatched Rows", f"{country_unmatched:,}")
    country_unfilled_cols = [c for c in ['Learner name', 'Learner ID', 'Course intake No.', 'Learner email address', 'Country of Organisation']
                              if c in country_unfilled.columns]
    with st.expander(f"🔍 View unmatched country rows ({country_unmatched}) — view and rectify"):
        if country_unmatched == 0:
            st.success("No unmatched country rows.")
        else:
            st.dataframe(country_unfilled[country_unfilled_cols], use_container_width=True)

    st.divider()

    # 6.5 School mapping
    st.subheader("School Mapping")
    st.caption(f"ℹ️ {namelist_school_source_note}")
    school_order = sorted(['AES', 'AMS', 'AVS', 'ATS']) + ['Others']
    school_counts = df_attendance_clean['School'].value_counts().reindex(school_order).fillna(0).astype(int).reset_index()
    school_counts.columns = ['School', 'Trainee Count']
    st.dataframe(school_counts, use_container_width=True)


# ── REFERENCE & CLASSIFICATION TAB ────────────────────────────
with tab_reference:
    st.info("💡 To add, edit, or remove organization classifications, use the "
            "**'✏️ Manage Organization Name Classifications'** panel near the top of the page "
            "(above the file upload sidebar), then re-run the pipeline.")
    st.subheader("Organisation Classification by Country")
    st.caption("Distinct organisations found in this run's attendance data (≥ threshold, reporting period), "
               "with the country and classification actually used for KPI 1 / KPI 2.")

    _org_country_cols = [c for c in ['Organization Name', 'Country of Organisation', 'Segment'] if c in df_filtered.columns]
    if 'Organization Name' in _org_country_cols:
        org_by_country = (
            df_filtered[_org_country_cols]
            .dropna(subset=['Organization Name'])
            .loc[lambda d: d['Organization Name'].astype(str).str.strip() != '']
            .drop_duplicates(subset=['Organization Name'])
            .rename(columns={'Segment': 'Classification'})
            .sort_values(['Classification', 'Organization Name'])
            .reset_index(drop=True)
        )
    else:
        org_by_country = pd.DataFrame(columns=['Organization Name', 'Country of Organisation', 'Classification'])

    # If the org's entry in the classification list itself has no country
    # asserted (blank, or the column doesn't exist at all in the upload),
    # there's nothing to check against — skip flagging entirely rather than
    # inferring an "expected" country purely from Classification.
    _org_class_country_col = 'Country/Region of Organization'
    if _org_class_country_col in org_class_table_raw.columns:
        _blank_country_keys = set(
            org_class_table_raw.loc[
                org_class_table_raw[_org_class_country_col].isna()
                | (org_class_table_raw[_org_class_country_col].astype(str).str.strip() == ''),
                '_key'
            ]
        )
    else:
        _blank_country_keys = set(org_class_table_raw['_key']) if len(org_class_table_raw) > 0 else set()

    def _country_check(row):
        org_key = re.sub(r'\s+', ' ', str(row.get('Organization Name', '')).strip()).upper()
        if org_key in _blank_country_keys:
            return '⬜ Not checked (no country on classification list)'
        country = str(row.get('Country of Organisation', '')).strip().upper()
        cls = row.get('Classification', '')
        if country in ('', 'NAN', 'NONE'):
            return '❔ No country on record'
        if cls in ('CAAS', 'Local'):
            return '✅ Consistent' if country == 'SINGAPORE' else '⚠️ Mismatch (expected Singapore)'
        elif cls == 'International':
            return '✅ Consistent' if country != 'SINGAPORE' else '⚠️ Mismatch (expected non-Singapore)'
        return '—'

    if len(org_by_country) > 0:
        org_by_country['Country Check'] = org_by_country.apply(_country_check, axis=1)
    else:
        org_by_country['Country Check'] = pd.Series(dtype=str)

    seg_counts = org_by_country['Classification'].value_counts().reindex(['International', 'Local', 'CAAS']).fillna(0).astype(int) \
        if 'Classification' in org_by_country.columns else pd.Series({'International': 0, 'Local': 0, 'CAAS': 0})
    mismatch_count = int(org_by_country['Country Check'].astype(str).str.startswith('⚠️').sum())
    r1, r2, r3, r4 = st.columns(4)
    r1.metric("International", int(seg_counts.get('International', 0)))
    r2.metric("Local", int(seg_counts.get('Local', 0)))
    r3.metric("CAAS", int(seg_counts.get('CAAS', 0)))
    r4.metric("⚠️ Country Mismatches", mismatch_count)
    st.dataframe(org_by_country, use_container_width=True)

    if mismatch_count > 0:
        with st.expander(f"⚠️ View {mismatch_count} classification/country mismatch(es) — view and rectify"):
            st.caption("CAAS and Local should have Country of Organisation = Singapore; International should not. "
                       "Fix these in the '✏️ Manage Organization Name Classifications' panel near the top of the page.")
            st.dataframe(
                org_by_country[org_by_country['Country Check'].astype(str).str.startswith('⚠️')],
                use_container_width=True
            )

    st.divider()

    st.subheader("Senior Designation List (used for KPI 2)")
    st.caption(f"Source: **{designation_source}**")
    senior_df = pd.DataFrame(sorted(SENIOR_DESIGNATIONS), columns=['Designation'])
    st.dataframe(senior_df, use_container_width=True)

    st.divider()

    st.subheader("Singapore Public Holidays (used for trainee-day calculations)")
    st.caption(f"Source: **{public_holidays_source}**")
    if len(sg_holidays_np) > 0:
        holidays_df = pd.DataFrame(sorted(sg_holidays_np.astype(str)), columns=['Date'])
        st.dataframe(holidays_df, use_container_width=True)
    else:
        st.warning("⚠️ No public holidays loaded — trainee-day counts currently exclude weekends only, "
                   "not public holidays. This will under-count trainee days for any course spanning a "
                   "real Singapore public holiday.")


# ── Course Reporting Namelist CROSS-CHECK TAB ────────────────────────────────────
with tab_namelist:
    st.info(
        "📋 **Why this tab exists.** The end goal is for **Learn@SAA to be the single source of truth** — this "
        "app is a first step toward that. As an interim measure, Course Managers currently maintain **both** the "
        "Course Reporting Namelist and Learn@SAA in parallel, which means double work. This tab recomputes KPI 1 / KPI 2 / KPI 3 "
        "using **only the Course Reporting Namelist file** — completely independent of the Admitted / Enrollment / Attendance / "
        "Certificate exports used in the Results tab — so the two can be compared side by side. Large gaps below "
        "usually mean an intake has drifted out of sync between the Course Reporting Namelist and Learn@SAA."
    )

    for _note in namelist_kpis['notes']:
        st.caption(f"ℹ️ {_note}")

    st.divider()

    def _pct_vs_namelist(official_val, namelist_val):
        """% difference using the Course Reporting Namelist figure as the base."""
        if namelist_val in (0, None) or pd.isna(namelist_val):
            return None
        return (official_val - namelist_val) / namelist_val * 100

    # ---- "At a glance" variance charts (KPI 1 and KPI 2, side by side) ----
    st.subheader("Where the Numbers Disagree")
    st.caption("Attendance Report (Learn@SAA) vs Course Reporting Namelist, by KPI. % difference uses the Course Reporting Namelist as the base.")

    CHART_OFFICIAL_COLOR = "#1B3A6B"  # dark blue
    CHART_NAMELIST_COLOR = "#E67E22"  # orange

    def _build_variance_chart(categories, official_vals, namelist_vals, height=340):
        fig = go.Figure()
        fig.add_trace(go.Bar(
            name="Attendance Report (Learn@SAA)", x=categories, y=official_vals,
            marker_color=CHART_OFFICIAL_COLOR, text=[f"{v:,.0f}" for v in official_vals], textposition="outside",
            cliponaxis=False,
        ))
        fig.add_trace(go.Bar(
            name="Course Reporting Namelist", x=categories, y=namelist_vals,
            marker_color=CHART_NAMELIST_COLOR, text=[f"{v:,.0f}" for v in namelist_vals], textposition="outside",
            cliponaxis=False,
        ))
        max_val = max(list(official_vals) + list(namelist_vals) + [0])
        y_upper = max_val * 1.20 if max_val > 0 else 1
        fig.update_layout(
            barmode="group",
            yaxis_title="Trainee Days",
            legend=dict(orientation="h", yanchor="bottom", y=-0.3, xanchor="center", x=0.5),
            margin=dict(t=10, b=10, l=10, r=10),
            height=height,
            plot_bgcolor="rgba(0,0,0,0)",
            paper_bgcolor="rgba(0,0,0,0)",
        )
        return fig

    def _render_pct_boxes(categories, official_vals, namelist_vals):
        _cols = st.columns(len(categories))
        for _col, _label, _off, _nl in zip(_cols, categories, official_vals, namelist_vals):
            _pct = _pct_vs_namelist(_off, _nl)
            _pct_text = f"{_pct:+.1f}%" if _pct is not None else "N/A"
            _col.markdown(
                f"""
                <div style="background-color:#FBEAEA; border:1px solid #C0392B; border-radius:6px; padding:8px; text-align:center;">
                    <div style="font-size:0.7rem; color:#5B6B85;">{_label}</div>
                    <div style="font-size:1.1rem; font-weight:700; color:#C0392B;">{_pct_text}</div>
                </div>
                """,
                unsafe_allow_html=True,
            )

    chart_col1, divider_col, chart_col2 = st.columns([10, 1, 10])

    with chart_col1:
        st.markdown("**KPI 1 — International Trainee Days**")
        _kpi1_categories = ["International", "Local", "CAAS"]
        _kpi1_official = [kpi1_intl_total, kpi1_local_total, kpi1_caas_total]
        _kpi1_namelist = [namelist_kpis['kpi1_nl_intl'], namelist_kpis['kpi1_nl_local'], namelist_kpis['kpi1_nl_caas']]
        st.plotly_chart(_build_variance_chart(_kpi1_categories, _kpi1_official, _kpi1_namelist), use_container_width=True)
        _render_pct_boxes(_kpi1_categories, _kpi1_official, _kpi1_namelist)

    with divider_col:
        st.markdown(
            """
            <div style="width:1px; height:460px; background-color:rgba(0,0,0,0.15); margin:36px auto 0 auto;"></div>
            """,
            unsafe_allow_html=True,
        )

    with chart_col2:
        st.markdown("**KPI 2 — International Trainee Days for Directors and above**")
        if namelist_kpis['kpi2_available']:
            _kpi2_categories = ["International", "Local", "CAAS"]
            _kpi2_official = [kpi2_intl_total, kpi2_local_total, kpi2_caas_total]
            _kpi2_namelist = [namelist_kpis['kpi2_nl_intl'], namelist_kpis['kpi2_nl_local'], namelist_kpis['kpi2_nl_caas']]
            st.plotly_chart(_build_variance_chart(_kpi2_categories, _kpi2_official, _kpi2_namelist), use_container_width=True)
            _render_pct_boxes(_kpi2_categories, _kpi2_official, _kpi2_namelist)
        else:
            st.info("KPI 2 could not be computed from the Course Reporting Namelist — no designation/position column was found.")

    st.divider()

    # ---- Reconciliation summary ----
    st.subheader("Reconciliation: Attendance Report (Learn@SAA) vs Course Reporting Namelist (compiled by SP in Excel)")
    st.caption("% Difference uses the **Course Reporting Namelist figure as the base** — e.g. +10% means the official figure is "
               "10% higher than the Course Reporting Namelist figure.")
    _recon_rows = [
        {'KPI': 'KPI 1 — International Trainee Days', 'Attendance Report (Learn@SAA)': kpi1_intl_total, 'Course Reporting Namelist': namelist_kpis['kpi1_nl_intl']},
        {'KPI': 'Local Trainee Days', 'Attendance Report (Learn@SAA)': kpi1_local_total, 'Course Reporting Namelist': namelist_kpis['kpi1_nl_local']},
        {'KPI': 'CAAS Trainee Days', 'Attendance Report (Learn@SAA)': kpi1_caas_total, 'Course Reporting Namelist': namelist_kpis['kpi1_nl_caas']},
    ]
    if namelist_kpis['kpi2_available']:
        _recon_rows.append({'KPI': 'KPI 2 — International Trainee Days for Directors and above', 'Attendance Report (Learn@SAA)': kpi2_intl_total, 'Course Reporting Namelist': namelist_kpis['kpi2_nl_intl']})
        _recon_rows.append({'KPI': 'Local Senior Days', 'Attendance Report (Learn@SAA)': kpi2_local_total, 'Course Reporting Namelist': namelist_kpis['kpi2_nl_local']})
        _recon_rows.append({'KPI': 'CAAS Senior Days', 'Attendance Report (Learn@SAA)': kpi2_caas_total, 'Course Reporting Namelist': namelist_kpis['kpi2_nl_caas']})
    _recon_rows.append({'KPI': 'KPI 3 — New Programmes', 'Attendance Report (Learn@SAA)': kpi3_total_new_programmes, 'Course Reporting Namelist': namelist_kpis['kpi3_nl_total']})

    recon_df = pd.DataFrame(_recon_rows)
    recon_df['Difference (Learn@SAA − Course Reporting Namelist)'] = recon_df['Attendance Report (Learn@SAA)'] - recon_df['Course Reporting Namelist']
    recon_df['% Difference (base = Course Reporting Namelist)'] = recon_df.apply(
        lambda r: _pct_vs_namelist(r['Attendance Report (Learn@SAA)'], r['Course Reporting Namelist']), axis=1
    )
    recon_df['Match?'] = recon_df['Difference (Learn@SAA − Course Reporting Namelist)'].abs() < 0.01
    recon_df['Attendance Report (Learn@SAA)'] = recon_df['Attendance Report (Learn@SAA)'].round(2)
    recon_df['Course Reporting Namelist'] = recon_df['Course Reporting Namelist'].round(2)
    recon_df['Difference (Learn@SAA − Course Reporting Namelist)'] = recon_df['Difference (Learn@SAA − Course Reporting Namelist)'].round(2)
    recon_df['% Difference (base = Course Reporting Namelist)'] = recon_df['% Difference (base = Course Reporting Namelist)'].apply(
        lambda v: f"{v:+.1f}%" if v is not None else "N/A (Course Reporting Namelist = 0)"
    )
    st.dataframe(recon_df, use_container_width=True)

    _n_mismatch = int((~recon_df['Match?']).sum())
    if _n_mismatch == 0:
        st.success("✅ All computed KPIs match between the Course Reporting Namelist and the official pipeline for this reporting period.")
    else:
        st.warning(f"⚠️ {_n_mismatch} of {len(recon_df)} KPI figures differ between the Course Reporting Namelist and the official pipeline — "
                   "worth investigating which intakes are out of sync.")

    st.divider()

    # ---- Reconciliation summary, including Non-Course Events ----
    st.subheader("Reconciliation (Including Non-Course Events): Attendance Report (Learn@SAA) vs Course Reporting Namelist")
    st.caption(
        "The Results tab's KPI 1 / KPI 2 totals are computed only from the Attendance-based pipeline, and don't "
        "include the manually-tracked **Non-Course Events (from Course Reporting Namelist)** trainee days shown at the bottom of "
        "the Results tab — since those events never appear in Learn@SAA at all. This table adds those Non-Course "
        "Event trainee days into the Learn@SAA KPI 1 figures below (filtered to the same reporting period as "
        "everything else), for a fairer comparison against the Course Reporting Namelist, which already includes them naturally. "
        "% Difference still uses the **Course Reporting Namelist figure as the base**."
    )
    if _nce_start_date_col is None:
        st.caption("ℹ️ No 'Course Start Date' column found on the Non-Course Event rows — those trainee days "
                   "are included **without** date filtering below.")

    kpi1_intl_total_adj = kpi1_intl_total + non_course_intl_total
    kpi1_local_total_adj = kpi1_local_total + non_course_local_total

    _recon_rows_adj = [
        {'KPI': 'KPI 1 — International Trainee Days', 'Attendance Report (Learn@SAA)': kpi1_intl_total_adj, 'Course Reporting Namelist': namelist_kpis['kpi1_nl_intl']},
        {'KPI': 'Local Trainee Days', 'Attendance Report (Learn@SAA)': kpi1_local_total_adj, 'Course Reporting Namelist': namelist_kpis['kpi1_nl_local']},
        {'KPI': 'CAAS Trainee Days', 'Attendance Report (Learn@SAA)': kpi1_caas_total, 'Course Reporting Namelist': namelist_kpis['kpi1_nl_caas']},
    ]
    if namelist_kpis['kpi2_available']:
        _recon_rows_adj.append({'KPI': 'KPI 2 — International Trainee Days for Directors and above', 'Attendance Report (Learn@SAA)': kpi2_intl_total, 'Course Reporting Namelist': namelist_kpis['kpi2_nl_intl']})
        _recon_rows_adj.append({'KPI': 'Local Trainee Days for Directors and above', 'Attendance Report (Learn@SAA)': kpi2_local_total, 'Course Reporting Namelist': namelist_kpis['kpi2_nl_local']})
        _recon_rows_adj.append({'KPI': 'CAAS Senior Days for Directors and above', 'Attendance Report (Learn@SAA)': kpi2_caas_total, 'Course Reporting Namelist': namelist_kpis['kpi2_nl_caas']})
    _recon_rows_adj.append({'KPI': 'KPI 3 — New Programmes', 'Attendance Report (Learn@SAA)': kpi3_total_new_programmes, 'Course Reporting Namelist': namelist_kpis['kpi3_nl_total']})

    recon_df_adj = pd.DataFrame(_recon_rows_adj)
    recon_df_adj['Difference (Learn@SAA − Course Reporting Namelist)'] = recon_df_adj['Attendance Report (Learn@SAA)'] - recon_df_adj['Course Reporting Namelist']
    recon_df_adj['% Difference (base = Course Reporting Namelist)'] = recon_df_adj.apply(
        lambda r: _pct_vs_namelist(r['Attendance Report (Learn@SAA)'], r['Course Reporting Namelist']), axis=1
    )
    recon_df_adj['Match?'] = recon_df_adj['Difference (Learn@SAA − Course Reporting Namelist)'].abs() < 0.01
    recon_df_adj['Attendance Report (Learn@SAA)'] = recon_df_adj['Attendance Report (Learn@SAA)'].round(2)
    recon_df_adj['Course Reporting Namelist'] = recon_df_adj['Course Reporting Namelist'].round(2)
    recon_df_adj['Difference (Learn@SAA − Course Reporting Namelist)'] = recon_df_adj['Difference (Learn@SAA − Course Reporting Namelist)'].round(2)
    recon_df_adj['% Difference (base = Course Reporting Namelist)'] = recon_df_adj['% Difference (base = Course Reporting Namelist)'].apply(
        lambda v: f"{v:+.1f}%" if v is not None else "N/A (Course Reporting Namelist = 0)"
    )
    st.dataframe(recon_df_adj, use_container_width=True)

    _n_mismatch_adj = int((~recon_df_adj['Match?']).sum())
    if _n_mismatch_adj == 0:
        st.success("✅ All computed KPIs match between the Course Reporting Namelist and the official pipeline (with Non-Course Events included) for this reporting period.")
    else:
        st.warning(f"⚠️ {_n_mismatch_adj} of {len(recon_df_adj)} KPI figures still differ even after including Non-Course Events — "
                   "worth investigating which intakes are out of sync.")

    st.divider()

    with st.expander("🔍 View Course Reporting Namelist rows used in this computation"):
        st.dataframe(namelist_kpis['df_nl'], use_container_width=True)


# ── ADVANCED TAB ──────────────────────────────────────────────
with tab_advanced:
    if not show_advanced:
        st.info("☑️ Tick **'Show data preview'** in the sidebar to view this section.")
    else:
        st.caption("Quick sanity-check previews only (first 2 rows of each uploaded file).")
        st.write("**Admitted List**")
        st.dataframe(df_admitted.head(2), use_container_width=True)
        st.write("**Enrollment List**")
        st.dataframe(df_enrollment.head(2), use_container_width=True)
        st.write("**Attendance List (≥ threshold, cleaned)**")
        st.dataframe(df_attendance_clean.head(2), use_container_width=True)
        st.write("**Certificate List**")
        st.dataframe(df_cert.head(2), use_container_width=True)
        st.write("**Course Reporting Namelist**")
        st.dataframe(df_namelist.head(2), use_container_width=True)
