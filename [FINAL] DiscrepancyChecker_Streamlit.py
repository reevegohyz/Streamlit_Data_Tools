"""
Course Attendance vs Course Report — Discrepancy Checker
==========================================================
Streamlit app that replicates the data-cleaning pipeline from the
original notebook (Steps 1-3), then surfaces ONLY the results of:

  Step 5 — Summary of the discrepancy check
  Step 4 — Discrepancy check for the user-selected date range

Results are organised by School (AES / AMS / ATS / AVS / Others), using
the same Course-Intake-No.-based school mapping as the main SAA Data
Pipeline app.

Run with:
    streamlit run discrepancy_checker.py
"""

import datetime as dt
import io
import re

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

ATT_PCT_COL = "Module attendance percentage (by session hours)"

# ------------------------------------------------------------------
# School mapping (same logic as the main SAA Data Pipeline app)
# ------------------------------------------------------------------
SCHOOL_ORDER = sorted(['AES', 'AMS', 'AVS', 'ATS']) + ['Others']


def map_school(intake, namelist_school_lookup=None):
    """AES/AVS/ATS/AMS prefix rule first; for non-standard intake numbers
    (e.g. a numeric CAAS course ID), falls back to whatever School the
    Course Report itself has for that intake — no hardcoded exception list."""
    if pd.isna(intake):
        return 'Others'
    intake = str(intake).strip()
    for prefix in ('AES', 'AVS', 'ATS', 'AMS'):
        if intake.upper().startswith(prefix):
            return prefix
    if namelist_school_lookup:
        return namelist_school_lookup.get(intake.upper(), 'Others')
    return 'Others'


# ------------------------------------------------------------------
# Page setup
# ------------------------------------------------------------------
st.set_page_config(page_title="SAA Discrepancy Checker", layout="wide")

# ------------------------------------------------------------------
# Sidebar — file uploads & settings
# ------------------------------------------------------------------
with st.sidebar:
    st.header("📁 Step 1 — Upload Files")

    with st.expander("ℹ️ Where do I find each report?"):
        st.markdown(
            """
**Attendance List**
Report → Progress and timetabling → Attendance & session information report → Attendance information
*Example file name: `Attendance_Information_Report_20260526110649`*

**Namelist**
Download the Excel workbook from SharePoint and upload it as-is — no need to delete other sheets.
The app automatically reads the **Name List format** sheet.
*Example file: `Course Reporting and Evaluation_NEW.xlsx`*
            """
        )

    attendance_file = st.file_uploader(
        "Attendance Information Report (.xlsx)", type=["xlsx"], key="attendance"
    )
    course_file = st.file_uploader(
        "Course Reporting and Evaluation_NEW (.xlsx)", type=["xlsx"], key="course",
        help="You can upload the full original workbook — the app will "
             "automatically locate and use the 'Name List format' sheet.",
    )

    st.header("⚙️ Step 2 — Settings")
    _today = dt.date.today()
    start_date = st.date_input(
        "Report start date",
        value=dt.date(_today.year, 1, 1),
        help="Only courses whose start date falls on or after this date are compared.",
    )
    end_date = st.date_input(
        "Report end date",
        value=_today,
        help="Only courses whose start date falls on or before this date are compared.",
    )
    min_attendance = st.number_input(
        "Minimum attendance % to include", min_value=0, max_value=100, value=80
    )
    exclude_intakes_raw = st.text_area(
        "Intakes to exclude (comma-separated, optional)", value=""
    )
    # Normalized the same way as "Course intake No." itself (see
    # _normalize_intake_no) so an exclusion typed in a different case or
    # spacing than the source file still matches correctly.
    exclude_intakes = [
        re.sub(r'\s+', ' ', x.strip().upper()) for x in exclude_intakes_raw.split(",") if x.strip()
    ]

    run_button = st.button("Run discrepancy check", type="primary")

st.title(f"📊 SAA Discrepancy Checker ({start_date:%d %b %Y} – {end_date:%d %b %Y})")
st.caption("Upload the Attendance Report and the Namelist (Course Report). "
           "Results shown are limited to the discrepancy check for the "
           "selected date range and its summary, organised by School.")


# ------------------------------------------------------------------
# Cleaning helpers (Steps 1-3 from the original pipeline)
# ------------------------------------------------------------------
def _safe_str_col(series: pd.Series) -> pd.Series:
    """Coerces a column to string dtype safely before any .str accessor
    call. Handles columns pandas/Excel inferred as numeric (e.g. an intake
    number or ID stored as a pure number) or entirely blank (inferred as
    float64 NaN) — both of which make .str.strip() etc. raise 'Can only use
    .str accessor with string values!' otherwise. Real blanks/NaN are kept
    as empty strings rather than turned into the literal text 'nan', and a
    whole-number float (e.g. 260001.0, from a numeric-inferred ID column)
    is converted to '260001' rather than '260001.0', so ID/intake-number
    matching elsewhere in the app isn't silently corrupted."""
    def _to_str(x):
        if pd.isna(x):
            return ''
        if isinstance(x, float) and x.is_integer():
            return str(int(x))
        return str(x)
    return series.apply(_to_str)


def _normalize_intake_no(series: pd.Series) -> pd.Series:
    """Normalizes 'Course intake No.' before it's used as a join key between
    the Attendance and Course reports. Plain .str.strip() only removes
    leading/trailing whitespace — it does NOT fix case differences (e.g.
    'AES-1001' vs 'aes-1001') or internal double-spaces, either of which
    makes pandas treat the same real intake as two different join keys,
    causing every row to appear as a discrepancy even when the underlying
    data actually matches. Collapses internal whitespace and uppercases."""
    return series.str.strip().str.upper().str.replace(r'\s+', ' ', regex=True)


def clean_attendance(df_attendance: pd.DataFrame, exclude_intakes: list) -> pd.DataFrame:
    df = df_attendance.copy()

    df[ATT_PCT_COL] = (
        df[ATT_PCT_COL]
        .astype(str)
        .str.replace("%", "", regex=False)
        .str.strip()
        .replace("nan", None)
        .astype(float)
    )

    df["Learner name"] = _safe_str_col(df["Learner name"]).str.strip().str.title()
    df["Learner email address"] = _safe_str_col(df["Learner email address"]).str.strip().str.lower()
    df["Course intake No."] = _normalize_intake_no(_safe_str_col(df["Course intake No."]))

    # format="mixed" lets pandas infer the date format per-row instead of
    # locking onto one pattern from the first value — needed because these
    # reports mix abbreviated ("1 Jun 2026 00:00") and full ("1 June 2026
    # 00:00") month names in the same column. errors="coerce" turns any
    # genuinely unparseable value into NaT instead of crashing the app.
    df["Course start date"] = pd.to_datetime(
        df["Course start date"], format="mixed", dayfirst=True, errors="coerce"
    )
    df["Course end date"] = pd.to_datetime(
        df["Course end date"], format="mixed", dayfirst=True, errors="coerce"
    )

    df = df[~df["Course intake No."].isin(exclude_intakes)].reset_index(drop=True)

    df_clean = (
        df.sort_values(ATT_PCT_COL, ascending=False)
        .drop_duplicates(subset=["Course intake No.", "Learner email address"], keep="first")
        .reset_index(drop=True)
    )
    return df_clean


def clean_course(df_course: pd.DataFrame, exclude_intakes: list) -> pd.DataFrame:
    df = df_course.rename(
        columns={
            "Course Intake Number": "Course intake No.",
            "Course Start Date": "Course start date",
            "Course End Date": "Course end date",
            "Name": "Learner name",
            "Email Address": "Learner email address",
            "Organisation": "Company name",
            "Sponsor": "Sponsorship status",
        }
    ).copy()

    df["Learner name"] = _safe_str_col(df["Learner name"]).str.strip().str.title()
    df["Learner email address"] = _safe_str_col(df["Learner email address"]).str.strip().str.lower()
    df["Course intake No."] = _normalize_intake_no(_safe_str_col(df["Course intake No."]))

    # Same mixed-format fix as clean_attendance above.
    df["Course start date"] = pd.to_datetime(
        df["Course start date"], format="mixed", dayfirst=True, errors="coerce"
    )
    df["Course end date"] = pd.to_datetime(
        df["Course end date"], format="mixed", dayfirst=True, errors="coerce"
    )

    df = df[~df["Course intake No."].isin(exclude_intakes)].reset_index(drop=True)

    df_clean = (
        df.drop_duplicates(
            subset=["Course intake No.", "Learner email address", "Learner name"],
            keep="first",
        )
        .reset_index(drop=True)
    )
    return df_clean


def run_pipeline(df_attendance_raw, df_course_raw, exclude_intakes, min_attendance,
                  start_date, end_date):
    # Step 1: clean attendance
    df_attendance_clean = clean_attendance(df_attendance_raw, exclude_intakes)

    # Step 2: clean course report
    df_course_clean = clean_course(df_course_raw, exclude_intakes)

    # Warn (rather than silently drop) if any dates couldn't be parsed at all.
    unparsed_attendance = int(
        df_attendance_clean["Course start date"].isna().sum()
        + df_attendance_clean["Course end date"].isna().sum()
    )
    unparsed_course = int(
        df_course_clean["Course start date"].isna().sum()
        + df_course_clean["Course end date"].isna().sum()
    )
    if unparsed_attendance > 0:
        st.warning(
            f"⚠️ {unparsed_attendance} date value(s) in the Attendance Report could not be "
            "parsed and were treated as blank — those rows will be excluded from the date filter."
        )
    if unparsed_course > 0:
        st.warning(
            f"⚠️ {unparsed_course} date value(s) in the Course Report could not be parsed "
            "and were treated as blank — those rows will be excluded from the date filter."
        )

    # Step 3: filter by chosen date range + attendance threshold
    start_ts = pd.Timestamp(start_date)
    end_ts = pd.Timestamp(end_date)

    # A blank Course start date (common on real Course Report exports — see
    # your AES-1001 sample) would otherwise make a row fail the date filter
    # unconditionally, since NaT never satisfies >=/<= comparisons — silently
    # dropping an otherwise-valid course from every comparison regardless of
    # date range chosen. Falling back to Course end date when start is blank
    # avoids that.
    _attendance_filter_date = df_attendance_clean["Course start date"].fillna(df_attendance_clean["Course end date"])
    df_attendance_range = df_attendance_clean[
        (_attendance_filter_date >= start_ts)
        & (_attendance_filter_date <= end_ts)
    ].reset_index(drop=True)

    _course_filter_date = df_course_clean["Course start date"].fillna(df_course_clean["Course end date"])
    df_course_range = df_course_clean[
        (_course_filter_date >= start_ts)
        & (_course_filter_date <= end_ts)
    ].reset_index(drop=True)

    df_attendance_filtered = df_attendance_range[
        df_attendance_range[ATT_PCT_COL] >= min_attendance
    ].copy().reset_index(drop=True)

    # Step 4: discrepancy check for the selected date range
    #
    # dropna=False is essential here — pandas groupby() silently DROPS every
    # row where a grouping key is blank/NaN by default. Course Title (and,
    # less commonly, Course name) is often blank on real exports, which
    # would otherwise make entire courses vanish from this comparison
    # without any error or warning — the row count would just quietly not
    # match "Learners in Course Report" above, which is computed separately.
    attendance_count = (
        df_attendance_filtered.groupby(["Course intake No.", "Course name"], dropna=False)["Learner name"]
        .count()
        .reset_index()
    )
    attendance_count.columns = [
        "Course Intake No.",
        "Course Name (Attendance)",
        "Count in Attendance Report",
    ]

    course_count = (
        df_course_range.groupby(["Course intake No.", "Course Title"], dropna=False)["Learner name"]
        .count()
        .reset_index()
    )
    course_count.columns = [
        "Course Intake No.",
        "Course Name (Course Report)",
        "Count in Course Report",
    ]

    df_comparison = attendance_count.merge(
        course_count, on="Course Intake No.", how="outer"
    )
    # Fill each column type appropriately — a blanket .fillna(0) on the whole
    # dataframe would put the integer 0 into the "Course Name" text columns
    # wherever an intake only appears in one report, mixing str and int in
    # the same column. That silently breaks Streamlit's Arrow serialization
    # (visible as "Serialization of dataframe to Arrow table was
    # unsuccessful..." warnings in the terminal) even though the app doesn't
    # crash. Fill counts with 0 and names with "" instead.
    df_comparison["Course Name (Attendance)"] = df_comparison["Course Name (Attendance)"].fillna("")
    df_comparison["Course Name (Course Report)"] = df_comparison["Course Name (Course Report)"].fillna("")
    df_comparison["Count in Attendance Report"] = df_comparison["Count in Attendance Report"].fillna(0).astype(int)
    df_comparison["Count in Course Report"] = df_comparison["Count in Course Report"].fillna(0).astype(int)
    df_comparison["Variance"] = (
        df_comparison["Count in Attendance Report"] - df_comparison["Count in Course Report"]
    )
    df_comparison["Discrepancy"] = df_comparison["Variance"] != 0

    # Organise by School. Standard AES/AVS/ATS/AMS-prefixed intakes are
    # mapped directly; anything else falls back to whatever School the
    # Course Report itself already has for that intake, if it has a School
    # column at all — no hardcoded exception list to maintain.
    _namelist_school_lookup = {}
    if "School" in df_course_clean.columns:
        _nl_school = df_course_clean[["Course intake No.", "School"]].dropna(subset=["Course intake No."])
        _namelist_school_lookup = dict(zip(_nl_school["Course intake No."], _nl_school["School"]))
    df_comparison["School"] = df_comparison["Course Intake No."].apply(
        lambda x: map_school(x, _namelist_school_lookup)
    )
    df_comparison = df_comparison[
        ["School", "Course Intake No.", "Course Name (Attendance)", "Count in Attendance Report",
         "Course Name (Course Report)", "Count in Course Report", "Variance", "Discrepancy"]
    ]

    df_discrepancy = df_comparison[df_comparison["Discrepancy"]].reset_index(drop=True)

    # Per-school breakdown (all intakes checked vs how many have a discrepancy)
    school_summary = (
        df_comparison.groupby("School")
        .agg(
            Total_Intakes=("Course Intake No.", "nunique"),
            Intakes_With_Discrepancy=("Discrepancy", "sum"),
        )
        .reindex(SCHOOL_ORDER)
        .fillna(0)
        .astype(int)
        .reset_index()
        .rename(columns={"Total_Intakes": "Total Intakes", "Intakes_With_Discrepancy": "Intakes with Discrepancy"})
    )

    # Step 5: summary
    summary = {
        "learners_range_attendance": len(df_attendance_filtered),
        "learners_range_course": len(df_course_range),
        "total_intakes": len(df_comparison),
        "total_intakes_with_discrepancy": len(df_discrepancy),
        "sum_abs_variance": df_comparison["Variance"].abs().sum(),
    }

    return df_discrepancy, df_comparison, school_summary, summary


# ------------------------------------------------------------------
# Styled Excel export (replaces the plain CSV download)
# ------------------------------------------------------------------
NAVY = "192F55"
GREY_BORDER = "BFBFBF"
RED = "C00000"
BLACK = "000000"

_thin_side = Side(style="thin", color=GREY_BORDER)
_thin_border = Border(left=_thin_side, right=_thin_side, top=_thin_side, bottom=_thin_side)
_header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
_header_font = Font(name="Arial", bold=True, color="FFFFFF")
_body_font = Font(name="Arial", color=BLACK)
_red_font = Font(name="Arial", bold=True, color=RED)


def _write_styled_sheet(ws, headers, rows, red_cols=None, col_widths=None):
    """Writes one worksheet: navy bold header, thin grey grid, no row banding,
    autofilter + frozen header row, and red bold text in `red_cols` (0-indexed)
    wherever that cell's value is truthy/non-zero (i.e. a discrepancy exists)."""
    red_cols = red_cols or []

    for c, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=header)
        cell.font = _header_font
        cell.fill = _header_fill
        cell.border = _thin_border
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for r, row in enumerate(rows, start=2):
        for c, value in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.border = _thin_border
            is_flagged = (c - 1) in red_cols and isinstance(value, (int, float)) and value != 0
            cell.font = _red_font if is_flagged else _body_font
            if isinstance(value, (int, float)):
                cell.alignment = Alignment(horizontal="center")

    last_col_letter = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col_letter}{len(rows) + 1}"
    ws.freeze_panes = "A2"

    if col_widths:
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w


def build_styled_workbook(df_discrepancy, school_summary, summary, date_label):
    """Builds the 3-sheet workbook (Summary, By School, Discrepancy Detail)
    with the CAAS-style navy/grey/red formatting, and returns it as bytes
    ready for st.download_button."""
    wb = Workbook()

    # ---- Summary sheet ----
    ws_summary = wb.active
    ws_summary.title = "Summary"
    summary_rows = [
        [f"Total intakes ({date_label})", summary["total_intakes"]],
        ["Intakes with discrepancy", summary["total_intakes_with_discrepancy"]],
        ["Sum of absolute variance", int(summary["sum_abs_variance"])],
        [f"Learners in Attendance Report ({date_label})", summary["learners_range_attendance"]],
        [f"Learners in Course Report ({date_label})", summary["learners_range_course"]],
    ]
    _write_styled_sheet(
        ws_summary, ["Metric", "Value"], summary_rows,
        red_cols=[], col_widths=[42, 16],
    )
    # Only the "Intakes with discrepancy" row (row 3: header=1, then rows in
    # summary_rows order) should turn red when it's non-zero — every other
    # numeric metric here (totals, learner counts) isn't itself a problem.
    if summary["total_intakes_with_discrepancy"] > 0:
        ws_summary.cell(row=3, column=2).font = _red_font

    # ---- By School sheet ----
    ws_school = wb.create_sheet("By School")
    _write_styled_sheet(
        ws_school,
        list(school_summary.columns),
        school_summary.values.tolist(),
        red_cols=[2],  # "Intakes with Discrepancy" column
        col_widths=[12, 16, 22],
    )

    # ---- Discrepancy Detail sheet ----
    ws_detail = wb.create_sheet("Discrepancy Detail")
    detail_cols = [c for c in df_discrepancy.columns if c != "Discrepancy"]
    _write_styled_sheet(
        ws_detail,
        detail_cols,
        df_discrepancy[detail_cols].values.tolist(),
        red_cols=[detail_cols.index("Variance")],
        col_widths=[10, 18, 30, 16, 30, 16, 12],
    )

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ------------------------------------------------------------------
# Main
# ------------------------------------------------------------------
if run_button:
    if attendance_file is None or course_file is None:
        st.error("Please upload both the Attendance Report and the Course Report (Namelist) files.")
        st.stop()

    try:
        df_attendance_raw = pd.read_excel(
            attendance_file, dtype={ATT_PCT_COL: str}
        )
    except Exception as e:
        st.error(f"Could not read the Attendance Report file: {e}")
        st.stop()

    try:
        course_xl = pd.ExcelFile(course_file)
    except Exception as e:
        st.error(f"Could not read the Course Report file: {e}")
        st.stop()

    target_sheet = None
    for sheet_name in course_xl.sheet_names:
        if sheet_name.strip().lower() == "name list format":
            target_sheet = sheet_name
            break

    if target_sheet is None:
        st.error(
            "Could not find a 'Name List format' sheet in the uploaded Course "
            f"Report file. Sheets found: {course_xl.sheet_names}"
        )
        st.stop()

    try:
        df_course_raw = course_xl.parse(target_sheet, header=0)
    except Exception as e:
        st.error(f"Could not read the '{target_sheet}' sheet: {e}")
        st.stop()

    required_att_cols = {"Learner name", "Learner email address", "Course intake No.",
                          "Course start date", "Course end date", "Course name", ATT_PCT_COL}
    missing_att = required_att_cols - set(df_attendance_raw.columns)
    if missing_att:
        st.error(f"Attendance file is missing expected columns: {sorted(missing_att)}")
        st.stop()

    required_course_cols = {"Course Intake Number", "Course Title", "Name",
                             "Email Address", "Course Start Date", "Course End Date"}
    missing_course = required_course_cols - set(df_course_raw.columns)
    if missing_course:
        st.error(
            f"Course Report (sheet '{target_sheet}') is missing expected columns: {sorted(missing_course)}.\n\n"
            f"Columns found: {sorted(df_course_raw.columns)}"
        )
        st.stop()

    if start_date > end_date:
        st.error("Start date must be on or before the end date.")
        st.stop()

    try:
        with st.spinner("Cleaning data and computing discrepancies..."):
            df_discrepancy, df_comparison, school_summary, summary = run_pipeline(
                df_attendance_raw, df_course_raw, exclude_intakes, min_attendance,
                start_date, end_date,
            )
    except Exception as e:
        st.error(f"Something went wrong while processing the files: {e}")
        st.stop()

    date_label = f"{start_date:%d %b %Y} – {end_date:%d %b %Y}"

    st.divider()

    # ---------------- Download button (centered, between the divider and Summary) ----------------
    workbook_bytes = build_styled_workbook(df_discrepancy, school_summary, summary, date_label)
    _, dl_col, _ = st.columns([1, 2, 1])
    dl_col.download_button(
        "Download discrepancy report (Excel)",
        data=workbook_bytes,
        file_name=f"discrepancy_{start_date:%Y%m%d}_{end_date:%Y%m%d}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

    st.divider()

    # ---------------- Step 5 output (summary, shown first) ----------------
    st.header("Summary")
    s1, s2, s3 = st.columns(3)
    s1.metric(f"Total intakes ({date_label})", summary["total_intakes"])
    s2.metric("Intakes with discrepancy", summary["total_intakes_with_discrepancy"])
    _pct_discrepancy = (
        summary["total_intakes_with_discrepancy"] / summary["total_intakes"] * 100
        if summary["total_intakes"] else 0
    )
    s2.markdown(
        f'<div style="margin-top:-14px; color:#FF0000; font-weight:700; font-size:0.85rem;">'
        f'{_pct_discrepancy:.1f}% of total intakes</div>',
        unsafe_allow_html=True,
    )
    s3.metric("Sum of absolute variance", int(summary["sum_abs_variance"]))

    c1, c2 = st.columns(2)
    c1.metric(f"Learners in Attendance Report ({date_label})", summary["learners_range_attendance"])
    c2.metric(f"Learners in Course Report ({date_label})", summary["learners_range_course"])

    st.divider()

    # ---------------- By School summary ----------------
    st.header("By School")
    st.caption("School is derived from the Course Intake No. prefix (AES / AMS / ATS / AVS), "
               "with a few manually-mapped intakes and everything else grouped under 'Others' — "
               "the same logic used in the main SAA Data Pipeline app.")
    st.dataframe(school_summary, use_container_width=True)

    st.divider()

    # ---------------- Step 4 output, organised by School ----------------
    st.header(f"Discrepancy Check ({date_label})")
    for school in SCHOOL_ORDER:
        school_df = df_discrepancy[df_discrepancy["School"] == school]
        with st.expander(f"📂 {school} — {len(school_df)} discrepant intake(s)"):
            if len(school_df) == 0:
                st.info(f"No discrepancies for {school} in this date range.")
            else:
                st.dataframe(school_df.drop(columns="School"), use_container_width=True)


else:
    st.info("Upload both files in the sidebar, adjust settings if needed, then click "
            "**Run discrepancy check**.")
