"""
2026 Trainee Achievement & Projection Dashboard
==================================================================
Focused, single-year (2026) view: what's actually happened so far this
year, what's projected for the rest of it (from an optional Course
Calendar upload), and how the two combine — by school and by segment
(Local / CAAS / International).

HISTORICAL DATA — pick ONE, via a sidebar toggle:
  A) Admitted List + Attendance List (full pipeline: dedup, attendance-%
     gating, business-day trainee-day calc, org-name classification).
     More rigorous, and cross-checkable against Learn@SAA.
  B) A single Course Reporting and Evaluation_NEW export, which already
     carries Trainee Type (Local/CAAS/International) and No. of Trainee
     Days pre-computed. Quicker, but trusts that file as-is. The data
     lives on its 'Name List format' sheet, not the first sheet.
Both are normalized into the SAME internal shape before anything else
runs, so every chart below is agnostic to which source fed it.

PROJECTION — optional Course Calendar / Projection List upload:
  projected trainee days = business-day duration x class size, computed
  for Min / Max / Average(Min,Max) — shown as a bar (at the Average) with
  a min-max whisker on top, rather than three separate bars.
  Segment split uses each course's own historical Local/CAAS/International
  mix (falling back to the school's mix, then 'Unknown', if the course
  has never run before).

SCHOOL MAPPING:
  Historical data -> derived from the Course Code prefix (AES/AVS/AMS/ATS).
  Course Calendar -> derived from its own School column, where SAVS/SAES/
  SAMS/SATS map to AVS/AES/AMS/ATS, and TE maps to Others (SharePoint's
  own school-code convention, different from the course-code prefix).

HOW TO RUN:
    streamlit run Trainee_Achievement_2026.py
"""

import re
import datetime
import traceback
import io
from pathlib import Path

import numpy as np
import pandas as pd
import streamlit as st
import plotly.graph_objects as go
from openpyxl.styles import Font, PatternFill
from openpyxl.drawing.image import Image as XLImage

st.set_page_config(page_title="SAA Course Performance & Projection", layout="wide")
st.title("🎯 SAA Course Performance & Projection")
st.caption(
    "Actual trainee days so far this year, alongside a projection for the rest of it, "
    "by school and by segment (Local / CAAS / International)."
)

REPORT_YEAR = 2026
month_map = {1: 'Jan', 2: 'Feb', 3: 'Mar', 4: 'Apr', 5: 'May', 6: 'Jun',
             7: 'Jul', 8: 'Aug', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dec'}

SCHOOL_ORDER = ['AES', 'AMS', 'AVS', 'ATS', 'Others']
SCHOOL_COLORS = {'AES': '#0F2447', 'AMS': '#1E5C3A', 'AVS': '#4A2364', 'ATS': '#0F5A68', 'Others': '#3A3A3A'}
SCHOOL_ICONS = {'AES': 'AES', 'AMS': 'AMS', 'AVS': 'AVS', 'ATS': 'ATS', 'Others': 'Others'}
GREEN_SHADES = {'Local': '#1B5E20', 'CAAS': '#43A047', 'International': '#A5D6A7', 'Unknown': '#DCEDC8'}
SEGMENTS = ['Local', 'CAAS', 'International', 'Unknown']
SEG_COLORS = {'Local': '#1B3A6B', 'CAAS': '#5B9BD5', 'International': '#AFC9E8', 'Unknown': '#B0B0B0'}
# For Year-on-Year comparisons: prior (non-current) years use this genuinely
# different, warm palette rather than a faded/opacity version of SEG_COLORS —
# an opacity difference alone doesn't read as clearly distinct at a glance,
# and it also can't produce separate legend entries per year.
PRIOR_YEAR_SEG_COLORS = {'Local': '#8C3B00', 'CAAS': '#C2620A', 'International': '#E8A54D', 'Unknown': '#F0CFA0'}


# ────────────────────────────────────────────────────────────────
# SHARED HELPERS
# ────────────────────────────────────────────────────────────────
# _sg_holidays_np is set later, from the sidebar's Singapore Public Holidays
# upload (falling back to a built-in 2026-only list if nothing's uploaded).
# Declared here as a placeholder so calculate_course_duration (defined next)
# has something to reference — it's reassigned in the global scope before
# any duration calculation actually runs, and Python functions resolve
# globals at call time, not at definition time, so this is safe.
_sg_holidays_np = np.array([], dtype='datetime64[D]')


def calculate_course_duration(start, end):
    if pd.isna(start) or pd.isna(end):
        return 0
    if start.weekday() >= 5:
        return (end - start).days + 1
    return np.busday_count(start.date(), (end + pd.Timedelta(days=1)).date(), holidays=_sg_holidays_np)


def extract_course_code(intake_no):
    if pd.isna(intake_no):
        return None
    s = str(intake_no).strip()
    return s.split('-')[0].strip() if '-' in s else s


def extract_course_code_v2(intake_no):
    """Performance Dashboard's course-code rule, distinct from extract_course_code:
    - 'AMS2001-0001' -> 'AMS2001' (prefix carries its own batch/year digits — the
      part before the first '-' IS the course code, same as extract_course_code).
    - 'AMS-0001' -> 'AMS-0001' kept whole (the prefix alone, with no digits, isn't
      a real course code by itself — these are treated as special/non-standard
      codes and not truncated)."""
    if pd.isna(intake_no):
        return None
    s = str(intake_no).strip()
    if '-' not in s:
        return s
    prefix = s.split('-')[0].strip()
    return s if prefix.isalpha() else prefix


def map_school_from_code(course_code_or_intake):
    """For historical data: derive School from the Course Code prefix."""
    if pd.isna(course_code_or_intake):
        return 'Others'
    s = str(course_code_or_intake).strip().upper()
    for prefix in ('AES', 'AVS', 'ATS', 'AMS'):
        if s.startswith(prefix):
            return prefix
    return 'Others'


CALENDAR_SCHOOL_MAP = {
    'SAVS': 'AVS', 'SAES': 'AES', 'SAMS': 'AMS', 'SATS': 'ATS',
    'AVS': 'AVS', 'AES': 'AES', 'AMS': 'AMS', 'ATS': 'ATS',
    'TE': 'Others',
}


def map_school_calendar(raw):
    """For the Course Calendar file: its own School column uses a
    different S-prefixed convention (SAVS/SAES/SAMS/SATS/TE)."""
    if pd.isna(raw):
        return 'Others'
    return CALENDAR_SCHOOL_MAP.get(str(raw).strip().upper(), 'Others')


def normalize_segment(val):
    s = str(val).strip().lower()
    if s == 'local':
        return 'Local'
    if s == 'caas':
        return 'CAAS'
    if s in ('international', 'intl'):
        return 'International'
    return 'Unknown'


PERSONAL_EMAIL_DOMAINS = {
    'gmail.com', 'googlemail.com', 'yahoo.com', 'yahoo.co.uk', 'yahoo.co.in',
    'yahoo.co.id', 'ymail.com', 'rocketmail.com', 'hotmail.com', 'hotmail.co.uk',
    'outlook.com', 'live.com', 'msn.com', 'icloud.com', 'me.com', 'mac.com',
    'aol.com', 'protonmail.com', 'proton.me', 'qq.com', '163.com', '126.com',
    'zoho.com', 'gmx.com', 'mail.com',
}


def is_personal_email(email):
    email = str(email).strip().lower()
    if '@' not in email:
        return False
    return email.rsplit('@', 1)[-1] in PERSONAL_EMAIL_DOMAINS


def dedup_by_name_and_email(df, name_col, email_col, intake_col):
    d = df.copy()
    d['_is_personal_email'] = d[email_col].apply(is_personal_email)
    d = d.sort_values('_is_personal_email', kind='stable').drop(columns='_is_personal_email')
    step1 = d.drop_duplicates(subset=[name_col, intake_col], keep='first')
    step2 = step1.drop_duplicates(subset=[email_col, intake_col], keep='first')
    return step2.reset_index(drop=True)


SG_NAME_PATTERNS = [
    r'\bpte\.?\s*ltd\.?\b', r'\bsingapore\b', r'\bsia engineering\b',
    r'\bst engineering\b', r'\bchangi airport\b', r'\brepublic of singapore navy\b',
]
CAAS_NAME_PATTERNS = [r'\bcivil aviation authority of singapore\b', r'\bcaas\b']
COUNTRY_HINTS = [
    'thailand', 'vietnam', 'congo', 'malaysia', 'nigeria', 'macau', 'brunei', 'hong kong',
    'namibia', 'oman', 'fiji', 'korea', 'rwanda', 'togo', 'philippines', 'indonesia',
    'myanmar', 'cambodia', 'laos', 'india', 'pakistan', 'bangladesh', 'sri lanka', 'nepal',
    'china', 'japan', 'mongolia', 'kazakhstan', 'uzbekistan', 'russia', 'ukraine', 'poland',
    'germany', 'france', 'united kingdom', 'united states', 'canada', 'australia',
    'new zealand', 'papua new guinea', 'brazil', 'argentina', 'mexico', 'egypt', 'kenya',
    'tanzania', 'uganda', 'ethiopia', 'ghana', 'south africa', 'morocco', 'zambia',
    'zimbabwe', 'saudi arabia', 'united arab emirates', 'qatar', 'kuwait', 'jordan',
    'turkey', 'italy', 'spain', 'netherlands', 'switzerland', 'sweden', 'norway', 'denmark',
    'mauritius', 'maldives', 'bhutan', 'taiwan', 'south korea', 'barbados', 'uganda',
]


def _norm_org_key(name):
    return re.sub(r'\s+', ' ', str(name).strip()).upper()


def heuristic_classify_org(key_upper):
    for pat in CAAS_NAME_PATTERNS:
        if re.search(pat, key_upper, re.IGNORECASE):
            return 'CAAS'
    for pat in SG_NAME_PATTERNS:
        if re.search(pat, key_upper, re.IGNORECASE):
            return 'Local'
    for country in COUNTRY_HINTS:
        if country.upper() in key_upper:
            return 'International'
    return 'International'


def classify_org_name(name, lookup_dict):
    if pd.isna(name) or str(name).strip() == '':
        return 'Unknown'
    key = _norm_org_key(name)
    if key in lookup_dict:
        return lookup_dict[key]
    return heuristic_classify_org(key)


def _find_col_in(cols, candidates):
    cols_lower = {re.sub(r'\s+', ' ', str(c)).strip().lower(): c for c in cols}
    for cand in candidates:
        if cand in cols_lower:
            return cols_lower[cand]
    return None


def _read_classification_upload(file, header_keywords=None):
    is_csv = file.name.lower().endswith(".csv")
    raw = pd.read_csv(file, header=None) if is_csv else pd.read_excel(file, header=None)
    header_row_idx = 0
    if header_keywords:
        for i in range(min(15, len(raw))):
            row_values = [str(v).strip().lower() for v in raw.iloc[i].tolist()]
            if any(kw in row_values for kw in header_keywords):
                header_row_idx = i
                break
    df = raw.iloc[header_row_idx + 1:].copy()
    df.columns = raw.iloc[header_row_idx]
    return df.reset_index(drop=True).dropna(how="all")


def _truthy(v):
    return str(v).strip().upper() in ('TRUE', 'YES', '1')


def month_is_elapsed(y, m, as_of):
    month_end = pd.Timestamp(int(y), int(m), 1) + pd.offsets.MonthEnd(0)
    return month_end <= as_of


def smart_parse_dates(series):
    """Auto-detects DD/MM vs MM/DD for a whole date column, rather than assuming
    one convention. Different generations of the Course Calendar export have
    used different conventions (e.g. '30/11/2026' = DD/MM, vs '11/30/2026' =
    MM/DD) — a hardcoded dayfirst= setting silently mis-parses whichever one
    doesn't match, shifting rows into the wrong month without any error.
    If any value's first number is >12, it can only be a day (forces DD/MM).
    If any value's second number is >12, it can only be a day (forces MM/DD).
    If both/neither signal appears, falls back to whichever setting parses
    more rows successfully."""
    s = series.astype(str)
    first_num = pd.to_numeric(s.str.extract(r'^\s*(\d{1,4})')[0], errors='coerce')
    second_num = pd.to_numeric(s.str.extract(r'^\s*\d{1,4}[\/\-](\d{1,4})')[0], errors='coerce')
    looks_day_first = bool((first_num > 12).any())
    looks_month_first = bool((second_num > 12).any())
    if looks_month_first and not looks_day_first:
        return pd.to_datetime(series, dayfirst=False, errors='coerce')
    if looks_day_first and not looks_month_first:
        return pd.to_datetime(series, dayfirst=True, errors='coerce')
    a = pd.to_datetime(series, dayfirst=True, errors='coerce')
    b = pd.to_datetime(series, dayfirst=False, errors='coerce')
    return a if a.notna().sum() >= b.notna().sum() else b


def hex_to_rgba(hex_color, alpha):
    hex_color = hex_color.lstrip('#')
    r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
    return f'rgba({r},{g},{b},{alpha})'


def safe_read(file, label, expected_cols, **kwargs):
    try:
        df = pd.read_excel(file, **kwargs)
    except Exception as e:
        st.error(f"❌ Could not read **{label}**: {e}")
        st.stop()
    missing = [c for c in expected_cols if c not in df.columns]
    if missing:
        st.error(f"❌ **{label}** is missing expected column(s): {', '.join(missing)}.")
        st.stop()
    return df


def read_excel_flexible_sheet(file, preferred_sheet_names, label, expected_cols):
    try:
        xls = pd.ExcelFile(file)
    except Exception as e:
        st.error(f"❌ Could not open **{label}**: {e}")
        st.stop()
    target = None
    for name in xls.sheet_names:
        if name.strip().lower() in [p.lower() for p in preferred_sheet_names]:
            target = name
            break
    if target is None:
        target = xls.sheet_names[0]
    df = pd.read_excel(xls, sheet_name=target, header=0)
    missing = [c for c in expected_cols if c not in df.columns]
    if missing:
        st.error(f"❌ **{label}** (sheet '{target}') is missing expected column(s): {', '.join(missing)}.")
        st.stop()
    return df


# ────────────────────────────────────────────────────────────────
# SIDEBAR — 📚 CLASSIFICATION (Organisation, Designation, Public Holidays)
# ────────────────────────────────────────────────────────────────
st.sidebar.header("📚 Classification")

st.sidebar.subheader("Organisation Classification")
st.sidebar.caption("Only used in 'Admitted + Attendance' mode — the Course Reporting file already has its own Trainee Type.")
uploaded_org_class_file = st.sidebar.file_uploader("Organisation Classification List (.xlsx or .csv)", type=["xlsx", "csv"])

ORG_CLASS_FILE = Path(__file__).resolve().parent / "org_classification_data.csv"
if "org_class_df" not in st.session_state:
    if ORG_CLASS_FILE.exists():
        try:
            st.session_state.org_class_df = pd.read_csv(ORG_CLASS_FILE)
            st.session_state.org_class_source = f"local file ({ORG_CLASS_FILE.name})"
        except Exception:
            st.session_state.org_class_df = pd.DataFrame(columns=['Organization Name', 'Classification'])
            st.session_state.org_class_source = "no data yet"
    else:
        st.session_state.org_class_df = pd.DataFrame(columns=['Organization Name', 'Classification'])
        st.session_state.org_class_source = "no data yet"

if uploaded_org_class_file is not None:
    try:
        _df = _read_classification_upload(uploaded_org_class_file, header_keywords=["organization name", "organisation name"])
        name_col = _find_col_in(_df.columns, ["organization name", "organisation name"])
        class_col = _find_col_in(_df.columns, ["classification"])
        if name_col and class_col:
            _df = _df.rename(columns={name_col: "Organization Name", class_col: "Classification"})
            st.session_state.org_class_df = _df
            st.session_state.org_class_source = f"uploaded file ({uploaded_org_class_file.name})"
        else:
            st.sidebar.error("❌ Missing 'Organization Name' or 'Classification' column.")
    except Exception as e:
        st.sidebar.error(f"❌ Could not read classification list: {e}")

with st.expander(f"✏️ Edit classifications ({len(st.session_state.org_class_df):,} rows)"):
    _edited = st.data_editor(
        st.session_state.org_class_df, num_rows="dynamic", use_container_width=True, height=220,
        column_config={"Classification": st.column_config.SelectboxColumn(options=["Local", "CAAS", "International", "Unknown"])},
        key="org_editor",
    )
    st.session_state.org_class_df = _edited
    if st.button("💾 Save"):
        try:
            _edited.to_csv(ORG_CLASS_FILE, index=False)
            st.session_state.org_class_source = f"local file ({ORG_CLASS_FILE.name})"
            st.success("Saved.")
        except Exception as e:
            st.warning(f"Could not save ({e}) — still active for this session.")

_org_raw = st.session_state.org_class_df.dropna(subset=['Organization Name']).copy()
if len(_org_raw):
    _org_raw['_key'] = _org_raw['Organization Name'].astype(str).str.strip().str.replace(r'\s+', ' ', regex=True).str.upper()
    org_class_lookup = dict(zip(_org_raw['_key'], _org_raw['Classification']))
else:
    org_class_lookup = {}

# Reserved here (top of page, below the classification editor) — filled in
# further down the script once the historical/calendar data is actually
# ready, since the export needs data that isn't computed yet at this point.
_export_button_slot = st.empty()

st.divider()

st.sidebar.divider()
st.sidebar.subheader("Director Designation Classification")
st.sidebar.caption("Needed for KPI 2 (international senior trainee days). Without this, KPI 2 reads 0.")
uploaded_designation_file = st.sidebar.file_uploader("Director Designation Classification List (.xlsx or .csv)", type=["xlsx", "csv"])

if uploaded_designation_file is not None:
    try:
        _desig_df = _read_classification_upload(uploaded_designation_file, header_keywords=["designation"])
        _designation_col = _find_col_in(_desig_df.columns, ["designation"])
        if _designation_col:
            SENIOR_DESIGNATIONS = set(_desig_df[_designation_col].dropna().astype(str).str.strip().str.lower())
            designation_source = f"uploaded file ({uploaded_designation_file.name}, {len(SENIOR_DESIGNATIONS):,} designations)"
        else:
            SENIOR_DESIGNATIONS = set()
            designation_source = "upload invalid — missing a Designation column"
            st.sidebar.error("❌ Missing a 'Designation' column.")
    except Exception as e:
        SENIOR_DESIGNATIONS = set()
        designation_source = f"upload failed ({e})"
        st.sidebar.error(f"❌ Could not read Director Designation Classification List: {e}")
else:
    SENIOR_DESIGNATIONS = set()
    designation_source = "no data — KPI 2 will read 0 until a Director Designation Classification List is uploaded"

st.sidebar.divider()
st.sidebar.subheader("Singapore Public Holidays")
st.sidebar.caption(
    "Needed for correct business-day duration calculations in 'Admitted + Attendance' mode, for ANY year "
    "(the app no longer assumes a fixed year's holidays). Without this, falls back to a 2026-only built-in list."
)
uploaded_holidays_file = st.sidebar.file_uploader("Singapore Public Holidays List (.xlsx or .csv)", type=["xlsx", "csv"])

_FALLBACK_SG_HOLIDAYS_2026 = [
    '2026-01-01', '2026-02-17', '2026-02-18', '2026-03-21', '2026-04-03',
    '2026-05-01', '2026-05-27', '2026-05-31', '2026-06-01', '2026-08-09',
    '2026-08-10', '2026-11-08', '2026-11-09', '2026-12-25',
]
if uploaded_holidays_file is not None:
    try:
        _hol_df = _read_classification_upload(uploaded_holidays_file, header_keywords=["date", "holiday date"])
        _hol_date_col = _find_col_in(_hol_df.columns, ["date", "holiday date", "public holiday"])
        if _hol_date_col:
            _hol_dates = pd.to_datetime(_hol_df[_hol_date_col], dayfirst=True, errors='coerce').dropna()
            _sg_holidays_np = np.array([d.strftime('%Y-%m-%d') for d in _hol_dates], dtype='datetime64[D]')
            holidays_source = f"uploaded file ({uploaded_holidays_file.name}, {len(_hol_dates):,} dates, years {_hol_dates.dt.year.min()}–{_hol_dates.dt.year.max()})"
        else:
            _sg_holidays_np = np.array(_FALLBACK_SG_HOLIDAYS_2026, dtype='datetime64[D]')
            holidays_source = "upload invalid — missing a Date column — falling back to the built-in 2026-only list"
            st.sidebar.error("❌ Missing a 'Date' column.")
    except Exception as e:
        _sg_holidays_np = np.array(_FALLBACK_SG_HOLIDAYS_2026, dtype='datetime64[D]')
        holidays_source = f"upload failed ({e}) — falling back to the built-in 2026-only list"
        st.sidebar.error(f"❌ Could not read Singapore Public Holidays List: {e}")
else:
    _sg_holidays_np = np.array(_FALLBACK_SG_HOLIDAYS_2026, dtype='datetime64[D]')
    holidays_source = "no data — using the built-in 2026-only list (business-day duration for other years will be slightly off)"

st.sidebar.divider()

# ────────────────────────────────────────────────────────────────
# SIDEBAR — 📁 COURSE DATA
# ────────────────────────────────────────────────────────────────
st.sidebar.header("📁 Course Data")

with st.sidebar.expander("ℹ️ Where do I find each report?"):
    st.markdown(
        """
**Method 1 — Admitted + Attendance**

*Admitted List*
Report → Course, application and enrolment → Applicant details report
*Example file name: `Application_Details_Report_20260709153318`*

*Attendance List*
Report → Progress and timetabling → Attendance & session information report → Attendance information
*Example file name: `Attendance_Information_Report_20260526110649`*

**Method 2 — Course Reporting workbook**

Download the Excel workbook from SharePoint and upload it as-is — no need to delete other sheets.
The app automatically reads the **Name List format** or **Course Evaluation** sheet (whichever is
present). You can upload more than one year's workbook at once — e.g. both the 2025 and 2026
files together — and the app will bucket each row into its own year automatically.
*Example files: `Course Reporting and Evaluation_NEW.xlsx` (2026), `Course Reporting and Evaluation_CY2025.xlsx` (2025)*
        """
    )

hist_mode = st.sidebar.radio(
    "Source", ["Admitted + Attendance (full pipeline)", "Course Reporting & Evaluation only (quick)"],
    label_visibility="collapsed",
)

uploaded_admitted = uploaded_attendance = None
uploaded_course_reporting_files = []
if hist_mode.startswith("Admitted"):
    uploaded_admitted = st.sidebar.file_uploader("Admitted List / Applicant details report (.xlsx)", type="xlsx")
    uploaded_attendance = st.sidebar.file_uploader("Attendance List (.xlsx)", type="xlsx")
    st.sidebar.caption("Single-year only for now — multi-year support in this mode is a bigger redesign (a matched Admitted+Attendance pair per year).")
else:
    uploaded_course_reporting_files = st.sidebar.file_uploader(
        "Course Reporting and Evaluation workbook(s) (.xlsx)", type="xlsx", accept_multiple_files=True,
    )
    st.sidebar.caption("Upload one file per year — e.g. both 2025 and 2026 workbooks together — for a multi-year view.")

st.sidebar.divider()
st.sidebar.header("📅 Projection")
uploaded_calendar = st.sidebar.file_uploader("Course Calendar / Projection List (.xlsx or .csv)", type=["xlsx", "csv"])

st.sidebar.divider()
st.sidebar.header("⚙️ Step 2 — Settings")
metric_choice = st.sidebar.selectbox("Metric", ["Trainee Days", "Trainee Count"], index=0)
st.sidebar.caption(
    "Trainee Count = one per attendee row (headcount). Trainee Days = course duration × attendee count "
    "(same logic as the pipeline app) — summed per attendee, this equals duration × headcount per intake."
)
metric_label = "Trainee Days" if metric_choice == "Trainee Days" else "Trainees"
_year_options = list(range(REPORT_YEAR - 3, REPORT_YEAR + 2))
selected_years = st.sidebar.multiselect(
    "Years to include", options=_year_options, default=[REPORT_YEAR],
    help="Pick one year for the single combined chart, or two+ years (incl. the current/future year) to also see a Year-on-Year comparison.",
)
if not selected_years:
    selected_years = [REPORT_YEAR]
    st.sidebar.info(f"No year selected — defaulting to {REPORT_YEAR}.")
selected_years = sorted(selected_years)
as_of_date = pd.Timestamp(st.sidebar.date_input("As of date (actual vs projected cutoff)", value=datetime.date.today()))

attendance_threshold = 80
if hist_mode.startswith("Admitted"):
    attendance_threshold = st.sidebar.slider("Minimum attendance % (by session)", 0, 100, 80)

KPI1_TARGET = st.sidebar.number_input("KPI 1 annual target (trainee days)", value=10000, step=500)
KPI2_TARGET = st.sidebar.number_input("KPI 2 annual target (international trainee days for directors and above)", value=2000, step=100)
st.sidebar.caption(
    "KPI 1 is tracked against the International segment's actual + projected total (shown below the combined "
    "chart). KPI 2 needs Designation/seniority data this app doesn't currently collect, so the target is captured "
    "here for reference only — no progress figure is computed against it yet."
)

include_chart_images = st.sidebar.checkbox(
    "Include chart images in Excel export", value=True,
    help="Requires Chrome to be installed (via the kaleido package). If unavailable, the export "
         "automatically falls back to data-only sheets instead of failing.",
)

hist_ready = (uploaded_admitted is not None and uploaded_attendance is not None) or len(uploaded_course_reporting_files) > 0
if not hist_ready:
    st.info("⬆️ Upload your historical data in the sidebar to begin — either **Admitted + Attendance**, or one or more **Course Reporting** workbooks.")
    st.stop()


# ────────────────────────────────────────────────────────────────
# BUILD CANONICAL HISTORICAL DATAFRAME
# Columns: Course Code, Course Intake Number, School, Segment, Year, Month, No of Trainee Days
# ────────────────────────────────────────────────────────────────
if uploaded_admitted is not None and uploaded_attendance is not None:
    df_admitted = read_excel_flexible_sheet(
        uploaded_admitted, ['Application details report', 'Applicant details report'],
        "Admitted List", ["Learner name", "Course intake No.", "Email address", "Learner ID"]
    )
    org_col = _find_col_in(df_admitted.columns, ['organization name', 'organisation name'])
    title_col = _find_col_in(df_admitted.columns, ['course name', 'course title', 'title'])
    df_admitted_clean = dedup_by_name_and_email(df_admitted, 'Learner name', 'Email address', 'Course intake No.')

    df_attendance = safe_read(
        uploaded_attendance, "Attendance List",
        ["Learner name", "Course intake No.", "Learner email address",
         "Module attendance percentage (by session number)", "Course start date", "Course end date", "Learner ID"],
    )
    df_attendance['Course start date'] = pd.to_datetime(df_attendance['Course start date'], dayfirst=True, errors='coerce')
    df_attendance['Course end date'] = pd.to_datetime(df_attendance['Course end date'], dayfirst=True, errors='coerce')
    df_att_clean = dedup_by_name_and_email(df_attendance, 'Learner name', 'Learner email address', 'Course intake No.')
    df_att_clean['Attendance %'] = pd.to_numeric(
        df_att_clean['Module attendance percentage (by session number)'].astype(str).str.replace('%', '', regex=False), errors='coerce'
    )
    df_att_gated = df_att_clean[df_att_clean['Attendance %'] >= attendance_threshold].copy()
    df_att_gated['Duration (days)'] = df_att_gated.apply(
        lambda r: calculate_course_duration(r['Course start date'], r['Course end date']), axis=1
    )
    # Column name stays 'No of Trainee Days' throughout the app regardless of metric choice
    # (minimizes downstream changes) — its VALUE switches between duration-weighted
    # trainee-days and a flat 1-per-row headcount based on the sidebar Metric selector.
    df_att_gated['No of Trainee Days'] = df_att_gated['Duration (days)'] if metric_choice == "Trainee Days" else 1.0
    df_att_gated['Headcount'] = 1.0  # always a plain headcount, independent of the Metric toggle — used by the Performance Dashboard's load factor, which is fundamentally a headcount-vs-capacity measure.
    df_att_gated['Course Code'] = df_att_gated['Course intake No.'].apply(extract_course_code)
    df_att_gated['Course Code (Performance)'] = df_att_gated['Course intake No.'].apply(extract_course_code_v2)
    df_att_gated['School'] = df_att_gated['Course Code'].apply(map_school_from_code)
    df_att_gated['Year'] = df_att_gated['Course end date'].dt.year
    df_att_gated['Month'] = df_att_gated['Course end date'].dt.month

    if org_col:
        lookup = df_admitted_clean[['Course intake No.', 'Learner ID', org_col]].copy()
        lookup['Course intake No.'] = lookup['Course intake No.'].astype(str).str.strip()
        lookup['Learner ID'] = lookup['Learner ID'].astype(str).str.strip()
        lookup = lookup.rename(columns={org_col: 'Organization Name'})
        df_att_gated['Course intake No.'] = df_att_gated['Course intake No.'].astype(str).str.strip()
        df_att_gated['Learner ID'] = df_att_gated['Learner ID'].astype(str).str.strip()
        df_att_gated = df_att_gated.merge(lookup, on=['Course intake No.', 'Learner ID'], how='left')
        df_att_gated['Segment'] = df_att_gated['Organization Name'].apply(lambda x: classify_org_name(x, org_class_lookup))
    else:
        df_att_gated['Segment'] = 'Unknown'

    if 'Designation' in df_admitted_clean.columns:
        desig_lookup = df_admitted_clean[['Course intake No.', 'Learner ID', 'Designation']].copy()
        desig_lookup['Course intake No.'] = desig_lookup['Course intake No.'].astype(str).str.strip()
        desig_lookup['Learner ID'] = desig_lookup['Learner ID'].astype(str).str.strip()
        df_att_gated['Course intake No.'] = df_att_gated['Course intake No.'].astype(str).str.strip()
        df_att_gated['Learner ID'] = df_att_gated['Learner ID'].astype(str).str.strip()
        df_att_gated = df_att_gated.merge(desig_lookup, on=['Course intake No.', 'Learner ID'], how='left')
        df_att_gated['Is Senior'] = df_att_gated['Designation'].astype(str).str.strip().str.lower().isin(SENIOR_DESIGNATIONS)
    else:
        df_att_gated['Is Senior'] = False

    if title_col:
        title_lookup = df_admitted_clean[['Course intake No.', title_col]].drop_duplicates(subset='Course intake No.').copy()
        title_lookup['Course intake No.'] = title_lookup['Course intake No.'].astype(str).str.strip()
        title_lookup = title_lookup.rename(columns={title_col: 'Course Title'})
        df_att_gated['Course intake No.'] = df_att_gated['Course intake No.'].astype(str).str.strip()
        df_att_gated = df_att_gated.merge(title_lookup, on='Course intake No.', how='left')
    else:
        df_att_gated['Course Title'] = df_att_gated['Course Code (Performance)']

    hist_df = df_att_gated.dropna(subset=['Year', 'Month'])[
        ['Course Code', 'Course Code (Performance)', 'Course Title', 'Course intake No.', 'School', 'Segment', 'Year', 'Month',
         'No of Trainee Days', 'Headcount', 'Is Senior', 'Course end date']
    ].rename(columns={'Course intake No.': 'Course Intake Number', 'Course end date': 'Course End Date'})
    hist_source_note = f"Admitted + Attendance pipeline (≥{attendance_threshold}% attendance gate)."
else:
    _cr_dfs = []
    _cr_notes = []
    for _cr_file in uploaded_course_reporting_files:
        _cr_i = read_excel_flexible_sheet(
            _cr_file, ['Name List format', 'Name List Format', 'Course Evaluation', 'course evaluation'],
            f"Course Reporting workbook ({_cr_file.name})",
            ["Course Intake Number", "No. of Trainee Days", "Trainee Type", "School"],
        )
        _end_col = _find_col_in(_cr_i.columns, ['course end date'])
        _start_col = _find_col_in(_cr_i.columns, ['course start date'])
        if _end_col:
            _cr_i['Course End Date'] = pd.to_datetime(_cr_i[_end_col], dayfirst=True, errors='coerce')
        elif _start_col:
            _cr_i['Course End Date'] = pd.to_datetime(_cr_i[_start_col], dayfirst=True, errors='coerce')
            _cr_notes.append(f"'{_cr_file.name}' has no Course End Date column — used Course Start Date instead for month/year bucketing.")
        else:
            st.error(f"❌ '{_cr_file.name}' has neither a Course End Date nor a Course Start Date column — cannot process.")
            st.stop()
        _cr_dfs.append(_cr_i)
    df_cr = pd.concat(_cr_dfs, ignore_index=True, sort=False)
    for _note in _cr_notes:
        st.sidebar.info(f"ℹ️ {_note}")

    # Real exports sometimes have a stray non-numeric value in this column
    # (e.g. a typo like 'c' instead of a number) — coerce to numeric so one
    # bad cell doesn't crash every downstream .sum(), and surface how many
    # rows were affected rather than silently dropping them to 0.
    _cr_days_raw = df_cr['No. of Trainee Days']
    df_cr['No. of Trainee Days'] = pd.to_numeric(_cr_days_raw, errors='coerce')
    _n_bad_days = int(df_cr['No. of Trainee Days'].isna().sum())
    if _n_bad_days > 0:
        st.sidebar.warning(
            f"⚠️ {_n_bad_days:,} row(s) in the Course Reporting file(s) had a non-numeric 'No. of Trainee Days' "
            "value (e.g. a typo) — treated as 0 rather than crashing. Check the source file(s) if this seems high."
        )
        df_cr['No. of Trainee Days'] = df_cr['No. of Trainee Days'].fillna(0)

    df_cr['Year'] = df_cr['Course End Date'].dt.year
    df_cr['Month'] = df_cr['Course End Date'].dt.month
    df_cr['Course Code'] = df_cr['Course Intake Number'].apply(extract_course_code)
    df_cr['Course Code (Performance)'] = df_cr['Course Intake Number'].apply(extract_course_code_v2)
    df_cr['School'] = df_cr['School'].apply(map_school_calendar)
    df_cr['Headcount'] = 1.0
    if 'Designation' in df_cr.columns:
        df_cr['Is Senior'] = df_cr['Designation'].astype(str).str.strip().str.lower().isin(SENIOR_DESIGNATIONS)
    else:
        df_cr['Is Senior'] = False
    _cr_title_col = _find_col_in(df_cr.columns, ['course title', 'course name', 'title'])
    df_cr['Course Title'] = df_cr[_cr_title_col] if _cr_title_col else df_cr['Course Code (Performance)']
    if metric_choice == "Trainee Count":
        df_cr['No. of Trainee Days'] = 1.0  # one per attendee row (headcount), overriding the file's own duration-based value
    hist_df = df_cr.dropna(subset=['Year', 'Month']).rename(
        columns={'No. of Trainee Days': 'No of Trainee Days', 'Trainee Type': 'Segment'}
    )[['Course Code', 'Course Code (Performance)', 'Course Title', 'Course Intake Number', 'School', 'Segment', 'Year', 'Month',
       'No of Trainee Days', 'Headcount', 'Is Senior', 'Course End Date']]
    hist_source_note = f"Course Reporting workbook(s) — {len(uploaded_course_reporting_files)} file(s) uploaded ({', '.join(f.name for f in uploaded_course_reporting_files)}), pre-computed Trainee Type / Trainee Days."

hist_df['Year'] = hist_df['Year'].astype(int)
hist_df['Month'] = hist_df['Month'].astype(int)
hist_df['Segment'] = hist_df['Segment'].apply(normalize_segment)
hist_df = hist_df[hist_df['Year'].isin(selected_years)].copy()

if len(hist_df) == 0:
    st.warning(f"⚠️ No rows found in your historical data for {', '.join(str(y) for y in selected_years)} — check the source file(s) and your Year selection.")


# ────────────────────────────────────────────────────────────────
# COURSE CALENDAR / PROJECTION — Min / Avg / Max, segment-split by historical mix
# ────────────────────────────────────────────────────────────────
calendar_ready = False
calendar_used_direct_segments = False
calendar_avg_df = pd.DataFrame(columns=['Course Code', 'School', 'Year', 'Month', 'Segment', 'No of Trainee Days'])
calendar_range_df = pd.DataFrame(columns=['School', 'Year', 'Month', 'Total_Min', 'Total_Max'])
calendar_capacity_df = pd.DataFrame(columns=['Course Intake Number', 'Course Code (Performance)', 'School', 'MaxClassSize', 'MinClassSize', 'Course End Date'])

if uploaded_calendar is not None:
    try:
        _is_csv = uploaded_calendar.name.lower().endswith(".csv")
        if _is_csv:
            _cal = pd.read_csv(uploaded_calendar, header=0)
        else:
            _xls = pd.ExcelFile(uploaded_calendar)
            _named_sheet = next((sn for sn in _xls.sheet_names if sn.strip().lower() == 'microsoft list projection'), None)
            if _named_sheet is not None:
                _cal = pd.read_excel(_xls, sheet_name=_named_sheet, header=0)
                _cal.columns = [str(c).strip() for c in _cal.columns]
            else:
                _best_cal, _best_score = None, -1
                for sn in _xls.sheet_names:
                    _trial = pd.read_excel(_xls, sheet_name=sn, header=0)
                    _trial.columns = [str(c).strip() for c in _trial.columns]
                    _tc = _find_col_in(_trial.columns, ['course intake number', 'course intake no.', 'course intake no'])
                    _score = len(_trial) if _tc is not None else -1
                    if _score > _best_score:
                        _best_cal, _best_score = _trial, _score
                _cal = _best_cal

        if _cal is not None:
            _cal.columns = [re.sub(r'\s+', ' ', str(c)).strip() for c in _cal.columns]
            colmap = {
                '_intake': _find_col_in(_cal.columns, ['course intake number', 'course intake no.', 'course intake no']),
                '_school': _find_col_in(_cal.columns, ['school']),
                '_start': _find_col_in(_cal.columns, ['coursestart', 'course start', 'course start date']),
                '_end': _find_col_in(_cal.columns, ['courseend', 'course end', 'course end date']),
                '_min': _find_col_in(_cal.columns, ['minclasssize', 'min class size']),
                '_max': _find_col_in(_cal.columns, ['maxclasssize', 'max class size']),
                '_pub': _find_col_in(_cal.columns, ['published?', 'published']),
                '_created': _find_col_in(_cal.columns, ['intakecreated?', 'intakecreated', 'intake created?']),
                # Pre-computed per-segment figures — present in the newer "Microsoft
                # List Projection" export. When available, these are used directly
                # instead of estimating from historical Local/CAAS/Intl mix.
                '_exp_caas': _find_col_in(_cal.columns, ['expected caas']),
                '_exp_local': _find_col_in(_cal.columns, ['expected local']),
                '_exp_int': _find_col_in(_cal.columns, ['expected int', 'expected intl', 'expected international']),
                '_days_caas': _find_col_in(_cal.columns, ['caas trainee days']),
                '_days_local': _find_col_in(_cal.columns, ['local trainee days']),
                '_days_int': _find_col_in(_cal.columns, ['int trainee days', 'intl trainee days', 'international trainee days']),
                '_min_days': _find_col_in(_cal.columns, ['mintraineedays', 'min trainee days']),
                '_max_days': _find_col_in(_cal.columns, ['maxtraineedays', 'max trainee days']),
            }
            missing_core = [n for n, c in [('Course Intake Number', colmap['_intake']), ('CourseStart', colmap['_start']), ('CourseEnd', colmap['_end'])] if c is None]

            with st.sidebar.expander("🔍 Calendar column detection", expanded=bool(missing_core)):
                st.write("Detected columns:", list(_cal.columns))
                st.write(colmap)

            if missing_core:
                st.sidebar.error(f"❌ Course Calendar is missing: {', '.join(missing_core)} — projections unavailable.")
            else:
                # Extract each identified column's VALUES directly into a brand-new
                # frame right now, while colmap and _cal are guaranteed to agree —
                # rather than renaming _cal in place and hoping a later '_intake'
                # lookup still matches. This sidesteps duplicate-label and hidden-
                # character issues in the source file's own header row entirely,
                # since nothing downstream ever looks a column up by name again.
                _n = len(_cal)
                _new_seg_cols = ['_exp_caas', '_exp_local', '_exp_int', '_days_caas', '_days_local', '_days_int', '_min_days', '_max_days']
                _cal = pd.DataFrame({
                    '_intake': _cal[colmap['_intake']].to_numpy(),
                    '_school': _cal[colmap['_school']].to_numpy() if colmap['_school'] else np.full(_n, np.nan),
                    '_start': _cal[colmap['_start']].to_numpy(),
                    '_end': _cal[colmap['_end']].to_numpy(),
                    '_min': _cal[colmap['_min']].to_numpy() if colmap['_min'] else np.full(_n, np.nan),
                    '_max': _cal[colmap['_max']].to_numpy() if colmap['_max'] else np.full(_n, np.nan),
                    '_pub': _cal[colmap['_pub']].to_numpy() if colmap['_pub'] else np.full(_n, True),
                    '_created': _cal[colmap['_created']].to_numpy() if colmap['_created'] else np.full(_n, True),
                    **{c: (_cal[colmap[c]].to_numpy() if colmap[c] else np.full(_n, np.nan)) for c in _new_seg_cols},
                })

                # Pre-computed segment figures use blank/dash to mean "0 expected",
                # not "unknown" — coerce to numeric and treat missing as 0 rather
                # than dropping the row or leaving a gap.
                _has_expected_counts = all(colmap[c] is not None for c in ['_exp_caas', '_exp_local', '_exp_int'])
                _has_expected_days = all(colmap[c] is not None for c in ['_days_caas', '_days_local', '_days_int'])
                for c in _new_seg_cols:
                    _cal[c] = pd.to_numeric(_cal[c], errors='coerce').fillna(0)

                _cal['_start'] = smart_parse_dates(_cal['_start'])
                _cal['_end'] = smart_parse_dates(_cal['_end'])
                _cal = _cal.dropna(subset=['_start', '_end']).reset_index(drop=True)

                _cal['Course Code'] = _cal['_intake'].apply(extract_course_code)
                _cal['Course Code (Performance)'] = _cal['_intake'].apply(extract_course_code_v2)
                _cal['School'] = _cal['_school'].apply(map_school_calendar)
                _cal['Year'] = _cal['_end'].dt.year
                _cal['Month'] = _cal['_end'].dt.month
                _cal['Duration (days)'] = _cal.apply(lambda r: calculate_course_duration(r['_start'], r['_end']), axis=1)

                # Per-intake capacity, for the Performance Dashboard's load factor —
                # independent of the Trainee Days / Trainee Count projection logic below.
                calendar_capacity_df = _cal[['_intake', 'Course Code (Performance)', 'School', '_min', '_max', '_end']].rename(
                    columns={'_intake': 'Course Intake Number', '_min': 'MinClassSize', '_max': 'MaxClassSize', '_end': 'Course End Date'}
                ).copy()
                calendar_capacity_df['MaxClassSize'] = pd.to_numeric(calendar_capacity_df['MaxClassSize'], errors='coerce')
                calendar_capacity_df['MinClassSize'] = pd.to_numeric(calendar_capacity_df['MinClassSize'], errors='coerce')
                calendar_capacity_df = calendar_capacity_df.dropna(subset=['MaxClassSize'])

                _min = pd.to_numeric(_cal['_min'], errors='coerce')
                _max = pd.to_numeric(_cal['_max'], errors='coerce')
                # Trainee Days: duration x class size (same logic as the pipeline app).
                # Trainee Count: class size alone — a projected headcount, no duration weighting.
                _duration_factor = _cal['Duration (days)'] if metric_choice == "Trainee Days" else 1.0
                _cal['Total_Avg'] = _duration_factor * (_min + _max) / 2
                _cal['Total_Min'] = (_duration_factor * _min).fillna(_cal['Total_Avg'])
                _cal['Total_Max'] = (_duration_factor * _max).fillna(_cal['Total_Avg'])

                # Prefer pre-computed per-segment figures straight from the source
                # file (e.g. "Microsoft List Projection": Expected CAAS/Local/Int
                # for headcount, CAAS/Local/Int Trainee Days for trainee-days) —
                # these are the source system's own estimate, not ours, so they're
                # used directly instead of the historical-mix estimation below.
                use_direct_segments = (_has_expected_counts if metric_choice == "Trainee Count" else _has_expected_days)
                calendar_used_direct_segments = use_direct_segments
                if use_direct_segments:
                    if metric_choice == "Trainee Count":
                        _cal['_seg_caas'], _cal['_seg_local'], _cal['_seg_int'] = _cal['_exp_caas'], _cal['_exp_local'], _cal['_exp_int']
                    else:
                        _cal['_seg_caas'], _cal['_seg_local'], _cal['_seg_int'] = _cal['_days_caas'], _cal['_days_local'], _cal['_days_int']
                    _cal['Total_Avg'] = _cal['_seg_caas'] + _cal['_seg_local'] + _cal['_seg_int']
                    if metric_choice == "Trainee Days" and colmap['_min_days'] and colmap['_max_days']:
                        _cal['Total_Min'] = _cal['_min_days'].where(_cal['_min_days'] > 0, _cal['Total_Avg'])
                        _cal['Total_Max'] = _cal['_max_days'].where(_cal['_max_days'] > 0, _cal['Total_Avg'])
                    elif colmap['_min'] and colmap['_max']:
                        _cal['Total_Min'] = _min.where(_min.notna(), _cal['Total_Avg'])
                        _cal['Total_Max'] = _max.where(_max.notna(), _cal['Total_Avg'])
                    else:
                        _cal['Total_Min'] = _cal['Total_Avg']
                        _cal['Total_Max'] = _cal['Total_Avg']

                if len(hist_df):
                    seg_mix_course = (hist_df.groupby(['Course Code', 'Segment'])['No of Trainee Days'].sum()
                                       .groupby(level=0).apply(lambda s: s / s.sum() if s.sum() > 0 else s))
                    seg_mix_school = (hist_df.groupby(['School', 'Segment'])['No of Trainee Days'].sum()
                                       .groupby(level=0).apply(lambda s: s / s.sum() if s.sum() > 0 else s))
                else:
                    seg_mix_course = pd.Series(dtype=float)
                    seg_mix_school = pd.Series(dtype=float)
                codes_with_hist = set(seg_mix_course.index.get_level_values(0)) if len(seg_mix_course) else set()
                schools_with_hist = set(seg_mix_school.index.get_level_values(0)) if len(seg_mix_school) else set()

                avg_rows, range_rows = [], []
                for _, r in _cal.iterrows():
                    total_avg = r['Total_Avg']
                    if pd.isna(total_avg) or total_avg <= 0:
                        continue
                    code, school = r['Course Code'], r['School']

                    if use_direct_segments:
                        seg_values = {'CAAS': r['_seg_caas'], 'Local': r['_seg_local'], 'International': r['_seg_int']}
                        for seg, val in seg_values.items():
                            if val <= 0:
                                continue
                            avg_rows.append({'Course Code': code, 'School': school, 'Year': r['Year'], 'Month': r['Month'],
                                              'Segment': seg, 'No of Trainee Days': val, 'Course End Date': r['_end']})
                    else:
                        if code in codes_with_hist:
                            mix = seg_mix_course.loc[code]
                        elif school in schools_with_hist:
                            mix = seg_mix_school.loc[school]
                        else:
                            mix = pd.Series({'Unknown': 1.0})
                        for seg, frac in mix.items():
                            if frac <= 0:
                                continue
                            avg_rows.append({'Course Code': code, 'School': school, 'Year': r['Year'], 'Month': r['Month'],
                                              'Segment': normalize_segment(seg), 'No of Trainee Days': total_avg * frac,
                                              'Course End Date': r['_end']})

                    range_rows.append({'School': school, 'Year': r['Year'], 'Month': r['Month'],
                                        'Total_Min': r['Total_Min'], 'Total_Max': r['Total_Max'],
                                        'Course End Date': r['_end']})

                if avg_rows:
                    calendar_avg_df = pd.DataFrame(avg_rows)
                    calendar_avg_df = calendar_avg_df[calendar_avg_df['Year'].isin(selected_years)]
                if range_rows:
                    calendar_range_df = pd.DataFrame(range_rows)
                    calendar_range_df = calendar_range_df[calendar_range_df['Year'].isin(selected_years)]
                calendar_ready = len(calendar_avg_df) > 0
    except Exception as e:
        st.sidebar.error(f"❌ Could not process the Course Calendar: {e}")
        with st.sidebar.expander("🔍 Full error detail"):
            st.code(traceback.format_exc())

if uploaded_calendar is not None:
    if calendar_ready:
        _seg_source_note = "using pre-computed Expected/Trainee-Days columns" if calendar_used_direct_segments else "estimated from historical Local/CAAS/Intl mix"
        st.sidebar.success(f"✅ Course Calendar loaded — {calendar_avg_df['Course Code'].nunique():,} projected course(s), segments {_seg_source_note}")
    else:
        st.sidebar.warning("⚠️ Course Calendar loaded but produced no usable projection rows.")


# ────────────────────────────────────────────────────────────────
# CHART / CARD BUILDING
# ────────────────────────────────────────────────────────────────
def get_monthly_segment_matrix(school_filter=None, year=None):
    year = year if year is not None else REPORT_YEAR
    h = hist_df[hist_df['Year'] == year]
    if school_filter is not None:
        h = h[h['School'] == school_filter]
    # Defensive: coerce to numeric here too, at the point of use, not just at
    # the source — a chokepoint against any bad non-numeric value reaching
    # .sum() and crashing the whole app (this exact crash has happened once
    # already, from a stray 'c' in a real source file's trainee-days column).
    h = h.copy()
    h['No of Trainee Days'] = pd.to_numeric(h['No of Trainee Days'], errors='coerce').fillna(0)
    # Calendar/projection data only ever applies to the current reporting
    # year — a past year like 2025 is entirely historical, and the
    # projection file has no bearing on it.
    if year == REPORT_YEAR:
        c_avg = calendar_avg_df if school_filter is None else calendar_avg_df[calendar_avg_df['School'] == school_filter]
        c_rng = calendar_range_df if school_filter is None else calendar_range_df[calendar_range_df['School'] == school_filter]
    else:
        c_avg = calendar_avg_df.iloc[0:0]
        c_rng = calendar_range_df.iloc[0:0]

    actual = {seg: [0.0] * 12 for seg in SEGMENTS}
    proj = {seg: [0.0] * 12 for seg in SEGMENTS}
    proj_min_total = [0.0] * 12
    proj_max_total = [0.0] * 12

    for m in range(1, 13):
        if month_is_elapsed(year, m, as_of_date):
            sub = h[h['Month'] == m]
            for seg in SEGMENTS:
                actual[seg][m - 1] = float(sub.loc[sub['Segment'] == seg, 'No of Trainee Days'].sum())
        elif calendar_ready:
            sub = c_avg[c_avg['Month'] == m]
            for seg in SEGMENTS:
                proj[seg][m - 1] = float(sub.loc[sub['Segment'] == seg, 'No of Trainee Days'].sum())
            rsub = c_rng[c_rng['Month'] == m]
            proj_min_total[m - 1] = float(rsub['Total_Min'].sum())
            proj_max_total[m - 1] = float(rsub['Total_Max'].sum())
    return actual, proj, proj_min_total, proj_max_total


def international_total(school_filter=None):
    """Actual + projected total for the International segment only."""
    actual, proj, _, _ = get_monthly_segment_matrix(school_filter)
    return sum(actual['International']) + sum(proj['International'])


def segments_total(school_filter, segments, year=None):
    """Actual + projected total across whichever segments are given."""
    actual, proj, _, _ = get_monthly_segment_matrix(school_filter, year=year)
    return sum(sum(actual[seg]) + sum(proj[seg]) for seg in segments)


def build_bar_chart(school_filter=None, height=260, show_cumulative_lines=False, segments=None, year=None):
    """Stacked bars (solid=actual, green shades=projected). By default plots
    all segments (Local/CAAS/International/Unknown) stacked — pass
    segments=['International'] to restrict to just that one, e.g. for a
    panel that's specifically labeled as International-only, so the chart
    and its label actually agree. Optionally overlaid with two continuous
    cumulative-total lines on a secondary axis: Expected (the projected-
    average scenario) and Max. Both lines run unbroken from January through
    December — the actual cumulative total for elapsed months, continuing
    directly into the projected scenario for future months, so there are
    no gaps. Pass year= to chart a specific year (e.g. 2025) instead of the
    current REPORT_YEAR — a past year is always fully 'actual'."""
    year = year if year is not None else REPORT_YEAR
    chart_segments = segments if segments is not None else SEGMENTS
    actual, proj, proj_min, proj_max = get_monthly_segment_matrix(school_filter, year=year)
    months = list(range(1, 13))
    month_labels = [month_map[m] for m in months]
    fig = go.Figure()

    for seg in chart_segments:
        y_actual, y_proj = [], []
        for i, m in enumerate(months):
            if month_is_elapsed(year, m, as_of_date):
                y_actual.append(actual[seg][i])
                y_proj.append(None)
            else:
                y_actual.append(None)
                y_proj.append(proj[seg][i])
        # Two separate traces per segment (Actual color, Projected color) so
        # every distinct shade actually used gets its own legend entry —
        # one ambiguous trace with per-point color switching can't do that.
        fig.add_trace(go.Bar(x=month_labels, y=y_actual, name=f"{seg} (Actual)",
                              marker_color=SEG_COLORS[seg], legendgroup=f"{seg}_actual", showlegend=any(v is not None for v in y_actual)))
        fig.add_trace(go.Bar(x=month_labels, y=y_proj, name=f"{seg} (Projected)",
                              marker_color=GREEN_SHADES[seg], legendgroup=f"{seg}_proj", showlegend=any(v is not None for v in y_proj)))

    # Y-axis buffer: leave headroom above the tallest bar so nothing sits
    # flush against the top of the chart.
    tops = []
    for i, m in enumerate(months):
        if month_is_elapsed(year, m, as_of_date):
            tops.append(sum(actual[seg][i] for seg in chart_segments))
        else:
            tops.append(sum(proj[seg][i] for seg in chart_segments))
    chart_max = max(tops) if tops else 0
    y_upper = chart_max * 1.15 if chart_max > 0 else 1

    # Headers demarcating the actual vs projected portions of the chart, plus
    # a vertical divider at the boundary — applied to every bar chart.
    elapsed_count = sum(1 for m in months if month_is_elapsed(year, m, as_of_date))
    if 0 < elapsed_count < 12:
        boundary_x = elapsed_count - 0.5
        fig.add_vline(x=boundary_x, line_width=1, line_dash="dash", line_color="rgba(130,130,130,0.55)")
        fig.add_annotation(x=(elapsed_count - 1) / 2, y=1.1, xref='x', yref='paper', text="ACTUAL",
                            showarrow=False, font=dict(size=10, color='#555555', family='Arial Black'))
        fig.add_annotation(x=elapsed_count + (12 - elapsed_count - 1) / 2, y=1.1, xref='x', yref='paper', text="PROJECTION",
                            showarrow=False, font=dict(size=10, color='#2E7D32', family='Arial Black'))
    elif elapsed_count >= 12:
        fig.add_annotation(x=5.5, y=1.1, xref='x', yref='paper', text="ACTUAL",
                            showarrow=False, font=dict(size=10, color='#555555', family='Arial Black'))
    elif elapsed_count <= 0:
        fig.add_annotation(x=5.5, y=1.1, xref='x', yref='paper', text="PROJECTION",
                            showarrow=False, font=dict(size=10, color='#2E7D32', family='Arial Black'))

    layout_kwargs = dict(
        barmode='stack', height=height, margin=dict(t=32, b=150, l=10, r=55 if show_cumulative_lines else 10),
        plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
        xaxis=dict(title=dict(text='Month', font=dict(size=11), standoff=10)),
        yaxis=dict(range=[0, y_upper], title=dict(
            text=(metric_label if set(chart_segments) == set(SEGMENTS) else f"{' + '.join(chart_segments)} {metric_label}"),
            font=dict(size=11))),
        legend=dict(orientation="h", yanchor="top", y=-0.5, xanchor="center", x=0.5, font=dict(size=9)),
    )

    if show_cumulative_lines:
        total_avg = [None if month_is_elapsed(year, m, as_of_date) else sum(proj[seg][i] for seg in chart_segments)
                     for i, m in enumerate(months)]
        monthly_actual_total = [sum(actual[seg][i] for seg in chart_segments) for i in range(12)]
        monthly_proj_avg_total = total_avg
        elapsed_idx = [i for i in range(12) if month_is_elapsed(year, i + 1, as_of_date)]
        last_actual_idx = max(elapsed_idx) if elapsed_idx else -1

        # One solid "Actual" line for the elapsed months, then two separate
        # dotted projection lines (Expected, Max) that both branch outward
        # from that same end point — rather than one line that silently
        # changes style partway through.
        cum_actual = [None] * 12
        running = 0.0
        for i in range(12):
            if i <= last_actual_idx:
                running += monthly_actual_total[i]
                cum_actual[i] = running
        base = cum_actual[last_actual_idx] if last_actual_idx >= 0 else 0.0

        def build_proj_segment(monthly_vals):
            """None for the actual months, starts exactly at the Actual
            line's end point (so it visually connects with no gap), then
            diverges through the projected months only."""
            out = [None] * 12
            if last_actual_idx >= 0:
                out[last_actual_idx] = base
            running = base
            for i in range(12):
                if i > last_actual_idx:
                    running += (monthly_vals[i] or 0)
                    out[i] = running
            return out

        cum_expected_proj = build_proj_segment(monthly_proj_avg_total)

        # Max is a whole-course capacity ceiling in the source file — it has
        # no per-segment breakdown at all. Only show this line for Overall
        # (all segments), where it's a real reported number. For any
        # narrower segment selection there's no honest way to derive it
        # (a proportional estimate was tried and rejected as inaccurate),
        # so the line is simply omitted rather than showing a guess.
        show_max_line = set(chart_segments) == set(SEGMENTS)
        if show_max_line:
            cum_max_proj = build_proj_segment(proj_max)

        fig.add_trace(go.Scatter(x=month_labels, y=cum_actual, mode='lines+markers', name='Cumulative (actual)',
                                  line=dict(color='#FF7A00', width=3, dash='solid'),
                                  marker=dict(size=6, color='#FF7A00', line=dict(color='#8A4200', width=1.5)),
                                  connectgaps=True, yaxis='y2'))
        fig.add_trace(go.Scatter(x=month_labels, y=cum_expected_proj, mode='lines+markers', name='Cumulative (expected)',
                                  line=dict(color='#1E70C4', width=2.5, dash='dot'),
                                  marker=dict(size=5, color='#1E70C4'),
                                  connectgaps=True, yaxis='y2'))
        if show_max_line:
            fig.add_trace(go.Scatter(x=month_labels, y=cum_max_proj, mode='lines+markers', name='Cumulative (max)',
                                      line=dict(color='#C0392B', width=2.5, dash='dot'),
                                      marker=dict(size=5, color='#C0392B'),
                                      connectgaps=True, yaxis='y2'))

        cum_top = max([v for v in (cum_max_proj if show_max_line else []) if v is not None]
                      + [v for v in cum_actual if v is not None]
                      + [v for v in cum_expected_proj if v is not None] + [0])
        cum_upper = cum_top * 1.15 if cum_top > 0 else 1
        layout_kwargs['yaxis2'] = dict(overlaying='y', side='right', range=[0, cum_upper], rangemode='tozero',
                                        showgrid=False, title=dict(text=f'Cumulative {metric_label.lower()}', font=dict(size=11)))

    fig.update_layout(**layout_kwargs)
    return fig


def build_yoy_bar_chart(years, school_filter=None, segments=None, height=420):
    """Grouped + stacked comparison across multiple years: each month shows
    one bar per year, side by side, each still internally stacked by
    segment. Uses Plotly's two-level category axis (month outer, year
    inner) so same-month bars from different years group together. The
    most recent year uses the normal segment colors (still split into
    solid=actual / green=projected as usual); every earlier year uses a
    genuinely different warm palette — not just a faded opacity version —
    so it's clearly distinct at a glance and gets its own legend entries
    (one per segment per year, e.g. 'Local — 2025' and 'Local — 2026'),
    rather than one ambiguous multi-color trace."""
    chart_segments = segments if segments is not None else SEGMENTS
    years = sorted(years)
    current_year = years[-1]
    month_year_pairs = [(m, y) for m in range(1, 13) for y in years]
    x_months = [month_map[m] for m, y in month_year_pairs]
    x_years = [str(y) for m, y in month_year_pairs]
    matrices = {y: get_monthly_segment_matrix(school_filter, year=y) for y in years}

    fig = go.Figure()
    tops = {}
    for yr in years:
        if yr == current_year:
            # Current year: split into Actual and Projected traces per
            # segment, same as the single-year chart, so both shades get
            # their own legend entry instead of one ambiguous trace.
            for seg in chart_segments:
                y_actual, y_proj = [], []
                for m, y2 in month_year_pairs:
                    if y2 != yr:
                        y_actual.append(None)
                        y_proj.append(None)
                        continue
                    actual, proj, _, _ = matrices[yr]
                    elapsed = month_is_elapsed(yr, m, as_of_date)
                    av = actual[seg][m - 1] if elapsed else None
                    pv = proj[seg][m - 1] if not elapsed else None
                    y_actual.append(av)
                    y_proj.append(pv)
                    tops[(m, yr)] = tops.get((m, yr), 0) + (av or pv or 0)
                fig.add_trace(go.Bar(x=[x_months, x_years], y=y_actual, name=f"{seg} (Actual) — {yr}",
                                      marker_color=SEG_COLORS[seg], legendgroup=f"{seg}_actual_{yr}",
                                      showlegend=any(v is not None for v in y_actual)))
                fig.add_trace(go.Bar(x=[x_months, x_years], y=y_proj, name=f"{seg} (Projected) — {yr}",
                                      marker_color=GREEN_SHADES[seg], legendgroup=f"{seg}_proj_{yr}",
                                      showlegend=any(v is not None for v in y_proj)))
        else:
            # Prior years are always fully actual — one consistent warm color.
            for seg in chart_segments:
                y_vals = []
                for m, y2 in month_year_pairs:
                    if y2 != yr:
                        y_vals.append(None)
                        continue
                    actual, proj, _, _ = matrices[yr]
                    val = actual[seg][m - 1]
                    y_vals.append(val)
                    tops[(m, yr)] = tops.get((m, yr), 0) + (val or 0)
                fig.add_trace(go.Bar(x=[x_months, x_years], y=y_vals, name=f"{seg} — {yr}",
                                      marker_color=PRIOR_YEAR_SEG_COLORS[seg], legendgroup=f"{seg}_{yr}",
                                      showlegend=any(v is not None for v in y_vals)))

    chart_max = max(tops.values()) if tops else 0
    y_upper = chart_max * 1.15 if chart_max > 0 else 1

    fig.update_layout(
        barmode='stack', height=height, margin=dict(t=20, b=170, l=10, r=10),
        plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
        yaxis=dict(range=[0, y_upper], title=dict(
            text=(metric_label if set(chart_segments) == set(SEGMENTS) else f"{' + '.join(chart_segments)} {metric_label}"),
            font=dict(size=11))),
        legend=dict(orientation="h", yanchor="top", y=-0.35, xanchor="center", x=0.5, font=dict(size=9)),
    )
    return fig


def render_school_cards(schools):
    cutoff_month = as_of_date.month if as_of_date.year == REPORT_YEAR else (12 if as_of_date.year > REPORT_YEAR else 0)
    actual_label = f"Actual (Jan–{month_map.get(cutoff_month, '')})" if cutoff_month >= 1 else "Actual"
    proj_label = f"Projection ({month_map.get(min(cutoff_month + 1, 12), '')}–Dec)" if cutoff_month < 12 else "Projection"

    overall_intl_total = sum(international_total(s) for s in schools) or 1

    cols = st.columns(len(schools))
    for col, sch in zip(cols, schools):
        actual, proj, _, _ = get_monthly_segment_matrix(sch)
        actual_total = sum(sum(actual[seg]) for seg in SEGMENTS)
        proj_total = sum(sum(proj[seg]) for seg in SEGMENTS)
        seg_totals = {seg: sum(actual[seg]) + sum(proj[seg]) for seg in SEGMENTS}
        intl_pct = seg_totals['International'] / overall_intl_total * 100
        color = SCHOOL_COLORS.get(sch, '#666666')
        icon = SCHOOL_ICONS.get(sch, '📁')
        with col:
            st.markdown(f"""
<div style="border:1px solid var(--border,#e3e3e3); border-radius:12px; padding:18px 20px; background:var(--surface-2,#fff);">
  <div style="display:flex; align-items:center; gap:12px; margin-bottom:14px;">
    <div style="width:52px; height:52px; border-radius:50%; background:{color}; display:flex; align-items:center; justify-content:center; font-size:11px; font-weight:700; color:#F5F5F5; letter-spacing:0.2px; text-align:center; line-height:1.1; flex-shrink:0;">{icon}</div>
    <div>
      <div style="font-weight:700; color:{color}; font-size:20px; line-height:1.1;">{sch}</div>
      <div style="font-size:12px; color:#888; letter-spacing:0.4px;">TOTAL {metric_label.upper()} ({REPORT_YEAR})</div>
    </div>
  </div>
  <div style="font-size:38px; font-weight:700; color:#222; margin-bottom:14px;">{actual_total + proj_total:,.0f}</div>
  <div style="display:flex; justify-content:space-between; font-size:13px; color:#666; margin-bottom:14px;">
    <div>{actual_label}<br><b style="font-size:19px; color:#222;">{actual_total:,.0f}</b></div>
    <div style="text-align:right;">{proj_label}<br><b style="font-size:19px; color:#222;">{proj_total:,.0f}</b></div>
  </div>
  <div style="border-top:1px solid #eee; padding-top:10px; display:flex; justify-content:space-between; font-size:13px; color:#666; margin-bottom:12px;">
    <span>Local<br><b style="font-size:16px; color:#222;">{seg_totals['Local']:,.0f}</b></span>
    <span>CAAS<br><b style="font-size:16px; color:#222;">{seg_totals['CAAS']:,.0f}</b></span>
    <span>Intl<br><b style="font-size:16px; color:#222;">{seg_totals['International']:,.0f}</b></span>
  </div>
  <div style="background:#EAF6EC; border-radius:6px; padding:6px 10px; text-align:center; font-size:12px; font-weight:700; color:#1E7B34;">
    {intl_pct:.1f}% of overall international {metric_label.lower()}
  </div>
</div>
            """, unsafe_allow_html=True)


def build_international_trend_chart(schools, height=380):
    """One line per school, International segment only — solid through the
    actual period, dashed for the projected period, sharing one legend
    entry per school. Matches the reference dashboard's trend chart."""
    months = list(range(1, 13))
    month_labels = [month_map[m] for m in months]
    fig = go.Figure()
    elapsed_count = sum(1 for m in months if month_is_elapsed(REPORT_YEAR, m, as_of_date))

    for sch in schools:
        actual, proj, _, _ = get_monthly_segment_matrix(sch)
        color = SCHOOL_COLORS.get(sch, '#666666')
        y_actual = [actual['International'][i] if month_is_elapsed(REPORT_YEAR, i + 1, as_of_date) else None for i in range(12)]
        y_proj = [None if month_is_elapsed(REPORT_YEAR, i + 1, as_of_date) else proj['International'][i] for i in range(12)]
        if 0 < elapsed_count < 12:
            y_proj[elapsed_count - 1] = y_actual[elapsed_count - 1]
        fig.add_trace(go.Scatter(x=month_labels, y=y_actual, mode='lines+markers', name=sch,
                                  line=dict(color=color, width=2.5), marker=dict(size=6, color=color),
                                  legendgroup=sch, showlegend=True))
        fig.add_trace(go.Scatter(x=month_labels, y=y_proj, mode='lines+markers', name=sch,
                                  line=dict(color=color, width=2.5, dash='dash'), marker=dict(size=6, color=color),
                                  legendgroup=sch, showlegend=False))

    if 0 < elapsed_count < 12:
        boundary_x = elapsed_count - 0.5
        fig.add_vline(x=boundary_x, line_width=1, line_dash="dash", line_color="rgba(130,130,130,0.55)")
        fig.add_annotation(x=(elapsed_count - 1) / 2, y=1.08, xref='x', yref='paper', text="ACTUAL",
                            showarrow=False, font=dict(size=11, color='#555555', family='Arial Black'))
        fig.add_annotation(x=elapsed_count + (12 - elapsed_count - 1) / 2, y=1.08, xref='x', yref='paper', text="PROJECTION",
                            showarrow=False, font=dict(size=11, color='#2E7D32', family='Arial Black'))

    fig.update_layout(
        height=height, margin=dict(t=36, b=60, l=10, r=10),
        plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
        xaxis=dict(title=dict(text='Month', font=dict(size=11), standoff=10)),
        yaxis=dict(title=dict(text=f'International {metric_label}', font=dict(size=11))),
        legend=dict(orientation="h", yanchor="top", y=-0.2, xanchor="center", x=0.5),
    )
    return fig


def render_executive_highlights(schools):
    totals = {s: international_total(s) for s in schools}
    overall = sum(totals.values()) or 1
    ranked = sorted(totals.items(), key=lambda x: x[1], reverse=True)
    lines = [f"Total international {metric_label.lower()} for {REPORT_YEAR} are <b>{overall:,.0f}</b>."]
    if ranked and ranked[0][1] > 0:
        top_sch, top_val = ranked[0]
        lines.append(f"<b>{top_sch}</b> leads with {top_val/overall*100:.1f}% of total international {metric_label.lower()} ({top_val:,.0f}).")
    for sch, val in ranked[1:]:
        if val <= 0:
            continue
        lines.append(f"<b>{sch}</b> contributes {val/overall*100:.1f}% ({val:,.0f}) of total international {metric_label.lower()}.")
    items_html = "".join(f'<div style="margin-bottom:10px; font-size:12px; color:#333; line-height:1.4;">🔹 {l}</div>' for l in lines)
    st.markdown(f"""
<div style="border:1px solid #e3e3e3; border-radius:10px; padding:14px 16px; background:#fff; height:100%;">
  <div style="font-weight:700; font-size:13px; color:#1B3A6B; margin-bottom:10px; letter-spacing:0.3px;">EXECUTIVE HIGHLIGHTS</div>
  {items_html}
</div>
    """, unsafe_allow_html=True)


# ────────────────────────────────────────────────────────────────
# EXCEL EXPORT HELPERS
# ────────────────────────────────────────────────────────────────
_NAVY = "192F55"


def _style_export_sheet(ws, df):
    for c in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=_NAVY, end_color=_NAVY, fill_type="solid")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def fig_to_png_bytes(fig, width=1000, height=550, scale=2):
    """Renders a Plotly figure to PNG bytes for embedding in Excel — legend
    included, since it's just part of the rendered image. Requires Chrome
    (via the kaleido package); returns None on any failure so callers can
    degrade gracefully instead of breaking the whole export."""
    try:
        return fig.to_image(format='png', width=width, height=height, scale=scale)
    except Exception:
        return None


def embed_chart_images(ws, chart_specs, start_row=1, col_letter='A', img_width_px=560):
    """Writes a title + image for each (title, fig) pair down one worksheet
    column, stacking vertically. Returns True if at least one image
    embedded successfully, False if the very first attempt failed (treated
    as 'this environment can't render charts at all' rather than retrying
    every single one)."""
    row = start_row
    any_success = False
    for i, (title, fig) in enumerate(chart_specs):
        png_bytes = fig_to_png_bytes(fig)
        if png_bytes is None:
            if i == 0:
                return False  # first failure -> stop trying entirely, let caller fall back
            continue
        any_success = True
        ws.cell(row=row, column=1, value=title).font = Font(name="Arial", bold=True, size=12, color=_NAVY)
        row += 1
        img = XLImage(io.BytesIO(png_bytes))
        img.width = img_width_px
        img.height = img_width_px * fig.layout.height / fig.layout.width if fig.layout.width else img_width_px * 0.55
        ws.add_image(img, f"{col_letter}{row}")
        row += int(img.height / 15) + 3  # roughly convert pixel height to row count, plus spacing
    return any_success


def build_school_summary_df(schools):
    overall_intl = sum(international_total(s) for s in schools) or 1
    rows = []
    for sch in schools:
        actual, proj, _, _ = get_monthly_segment_matrix(sch)
        actual_total = sum(sum(actual[seg]) for seg in SEGMENTS)
        proj_total = sum(sum(proj[seg]) for seg in SEGMENTS)
        seg_totals = {seg: sum(actual[seg]) + sum(proj[seg]) for seg in SEGMENTS}
        rows.append({
            'School': sch,
            f'Total {metric_label}': actual_total + proj_total,
            'Actual': actual_total,
            'Projection': proj_total,
            'Local': seg_totals['Local'],
            'CAAS': seg_totals['CAAS'],
            'International': seg_totals['International'],
            'Unknown': seg_totals['Unknown'],
            '% of Overall International': round(seg_totals['International'] / overall_intl * 100, 1),
        })
    return pd.DataFrame(rows)


def build_monthly_export_df(school_filter=None):
    actual, proj, proj_min, proj_max = get_monthly_segment_matrix(school_filter)
    rows = []
    for i, m in enumerate(range(1, 13)):
        elapsed = month_is_elapsed(REPORT_YEAR, m, as_of_date)
        status = 'Actual' if elapsed else 'Projected'
        for seg in SEGMENTS:
            rows.append({
                'Month': month_map[m], 'Segment': seg, 'Status': status,
                'Actual': actual[seg][i] if elapsed else None,
                'Projected (Expected)': None if elapsed else proj[seg][i],
                'Projected (Min)': None, 'Projected (Max)': None,
            })
        rows.append({
            'Month': month_map[m], 'Segment': 'TOTAL', 'Status': status,
            'Actual': sum(actual[seg][i] for seg in SEGMENTS) if elapsed else None,
            'Projected (Expected)': None if elapsed else sum(proj[seg][i] for seg in SEGMENTS),
            'Projected (Min)': None if elapsed else proj_min[i],
            'Projected (Max)': None if elapsed else proj_max[i],
        })
    return pd.DataFrame(rows)


def build_kpi_summary_df():
    _kpi1_actual, _kpi1_proj, _, _ = get_monthly_segment_matrix(None)
    _kpi1_total = sum(_kpi1_actual['International']) + sum(_kpi1_proj['International'])
    _kpi2_total = hist_df.loc[(hist_df['Segment'] == 'International') & (hist_df['Is Senior']), 'No of Trainee Days'].sum()
    return pd.DataFrame([
        {'KPI': f'KPI 1 — International {metric_label}', 'Actual (Actual+Projected)': _kpi1_total,
         'Target': KPI1_TARGET, '% of Target': round(_kpi1_total / KPI1_TARGET * 100, 1) if KPI1_TARGET > 0 else None},
        {'KPI': f'KPI 2 — International Senior {metric_label} (Actual only)', 'Actual (Actual+Projected)': _kpi2_total,
         'Target': KPI2_TARGET, '% of Target': round(_kpi2_total / KPI2_TARGET * 100, 1) if KPI2_TARGET > 0 else None},
    ])


def _lf_badge(pct):
    if pct >= 80:
        return 'GOOD', '#1E7B34', '#2E7D32'
    elif pct >= 65:
        return 'ACCEPTABLE', '#B8860B', '#F2B705'
    elif pct >= 40:
        return 'BELOW TARGET', '#C06A10', '#F2994A'
    return 'NEEDS ATTENTION', '#C0392B', '#E74C3C'


def compute_performance_data(school_scope):
    """Returns (merged, course_agg, summary_dict) for the given school scope
    ('Overall' or a specific school), or None if there's nothing to show —
    shared by both the Performance Dashboard tab and the top-level export,
    so the two never compute this differently."""
    _perf_hist = hist_df if school_scope == "Overall" else hist_df[hist_df['School'] == school_scope]
    if calendar_capacity_df.empty or len(_perf_hist) == 0:
        return None

    intake_headcount = _perf_hist.groupby(['Course Intake Number', 'Course Code (Performance)'], as_index=False)['Headcount'].sum()
    intake_month = _perf_hist.groupby('Course Intake Number', as_index=False)['Course End Date'].max()
    intake_title = _perf_hist.groupby('Course Intake Number', as_index=False)['Course Title'].first()
    intake_headcount = intake_headcount.merge(intake_month, on='Course Intake Number', how='left')
    intake_headcount = intake_headcount.merge(intake_title, on='Course Intake Number', how='left')

    cap_lookup = calendar_capacity_df[['Course Intake Number', 'MaxClassSize']].drop_duplicates(subset='Course Intake Number')
    merged = intake_headcount.merge(cap_lookup, on='Course Intake Number', how='inner')
    merged = merged[merged['MaxClassSize'] > 0]
    if len(merged) == 0:
        return None

    merged['Load Factor'] = merged['Headcount'] / merged['MaxClassSize'] * 100
    merged['Month-Year'] = merged['Course End Date'].dt.strftime('%b-%y')

    course_agg = merged.groupby('Course Code (Performance)', as_index=False).agg(
        Course_Title=('Course Title', lambda s: s.mode().iloc[0] if len(s.mode()) else s.iloc[0]),
        Course_Runs=('Course Intake Number', 'nunique'),
        Total_Trainees=('Headcount', 'sum'),
        Total_Capacity=('MaxClassSize', 'sum'),
    )
    course_agg['Load Factor'] = course_agg['Total_Trainees'] / course_agg['Total_Capacity'] * 100
    course_agg = course_agg.sort_values('Load Factor', ascending=False).reset_index(drop=True)
    course_agg.insert(0, 'Rank', range(1, len(course_agg) + 1))

    summary = dict(
        total_trainees=merged['Headcount'].sum(), total_capacity=merged['MaxClassSize'].sum(),
        overall_lf=(merged['Headcount'].sum() / merged['MaxClassSize'].sum() * 100) if merged['MaxClassSize'].sum() > 0 else 0,
        highest=merged.loc[merged['Load Factor'].idxmax()], lowest=merged.loc[merged['Load Factor'].idxmin()],
        courses_conducted=merged['Course Code (Performance)'].nunique(), n_runs=len(merged),
    )
    return merged, course_agg, summary


def build_performance_export_dfs(school_scope):
    """Summary + Load-Factor-by-Course export dataframes for one school
    scope, or (None, None) if there's nothing to export for it."""
    result = compute_performance_data(school_scope)
    if result is None:
        return None, None
    merged, course_agg, s = result
    export_df = course_agg.copy()
    export_df['Performance'] = export_df['Load Factor'].apply(lambda p: _lf_badge(p)[0])
    export_df = export_df.rename(columns={
        'Course_Title': 'Course Title', 'Course Code (Performance)': 'Course Code',
        'Course_Runs': 'Course Runs', 'Total_Trainees': 'Total Trainees', 'Total_Capacity': 'Total Capacity',
    })[['Rank', 'Course Title', 'Course Code', 'Course Runs', 'Total Trainees', 'Total Capacity', 'Load Factor', 'Performance']]
    summary_df = pd.DataFrame([{
        'School': school_scope, 'Years Included': ', '.join(str(y) for y in selected_years),
        'Overall Weighted Load Factor (%)': round(s['overall_lf'], 1), 'Total Trainees': s['total_trainees'],
        'Total Capacity': s['total_capacity'], 'Courses Conducted': s['courses_conducted'], 'Course Runs Matched': s['n_runs'],
        'Highest Load Factor (%)': round(s['highest']['Load Factor'], 1), 'Highest — Course': s['highest']['Course Title'],
        'Lowest Load Factor (%)': round(s['lowest']['Load Factor'], 1), 'Lowest — Course': s['lowest']['Course Title'],
    }])
    return summary_df, export_df


# ────────────────────────────────────────────────────────────────
# MAIN LAYOUT — two tabs
# ────────────────────────────────────────────────────────────────
active_schools = [
    s for s in SCHOOL_ORDER
    if len(hist_df[hist_df['School'] == s]) or (calendar_ready and len(calendar_avg_df[calendar_avg_df['School'] == s]))
]
if not active_schools:
    active_schools = SCHOOL_ORDER[:4]

# Fill the export button slot reserved earlier, now that everything it needs is ready.
_primary_year = REPORT_YEAR if REPORT_YEAR in selected_years else (max(selected_years) if selected_years else REPORT_YEAR)
_show_yoy_export = len(selected_years) >= 2 and any(y != _primary_year for y in selected_years)

_combined_export_buf = io.BytesIO()
with pd.ExcelWriter(_combined_export_buf, engine='openpyxl') as _writer:
    _export_sheets = {
        'School Summary': build_school_summary_df(active_schools),
        'KPI Summary': build_kpi_summary_df(),
        'Combined Monthly': build_monthly_export_df(None),
    }
    for _sch in active_schools:
        _export_sheets[f'{_sch} Monthly'[:31]] = build_monthly_export_df(_sch)
    _perf_summary_df, _perf_course_df = build_performance_export_dfs("Overall")
    if _perf_summary_df is not None:
        _export_sheets['Performance Summary'] = _perf_summary_df
        _export_sheets['Load Factor by Course'] = _perf_course_df
    for _sheet_name, _df in _export_sheets.items():
        _df.to_excel(_writer, sheet_name=_sheet_name, index=False)
        _style_export_sheet(_writer.sheets[_sheet_name], _df)

    _charts_embedded = False
    if include_chart_images:
        # Rebuild the same charts shown on-screen (using default/Overall
        # scope, since the tab's live segment-filter widgets haven't
        # rendered yet at this point in the script) for embedding.
        _chart_specs = [(f"Combined total — {_primary_year} (Overall)",
                          build_bar_chart(None, height=480, show_cumulative_lines=True, year=_primary_year))]
        if _show_yoy_export:
            _chart_specs.append((f"Combined total — Year on Year (Overall)",
                                  build_yoy_bar_chart(selected_years, None, height=520)))
        for _sch in active_schools:
            _chart_specs.append((f"{_sch} — International {metric_label}",
                                  build_bar_chart(_sch, height=340, segments=['International'], year=_primary_year)))
            if _show_yoy_export:
                _chart_specs.append((f"{_sch} — International {metric_label} — Year on Year",
                                      build_yoy_bar_chart(selected_years, _sch, segments=['International'], height=420)))
        _charts_ws = _writer.book.create_sheet("Charts")
        _charts_embedded = embed_chart_images(_charts_ws, _chart_specs)
        if not _charts_embedded:
            del _writer.book["Charts"]

with _export_button_slot.container():
    st.download_button(
        "⬇️ Download Full Report (Excel)", data=_combined_export_buf.getvalue(),
        file_name=f"SAA_Course_Performance_Projection_{REPORT_YEAR}_{datetime.date.today().strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    st.caption(
        "Includes both dashboards — Projection Dashboard sheets (School Summary, KPI Summary, monthly breakdowns) "
        "and Performance Dashboard sheets (Overall scope), regardless of which tab or school filter is currently selected below."
        + (" Chart images (with legends) are on the 'Charts' sheet."
           if include_chart_images and _charts_embedded else
           " Chart images were skipped — Chrome/kaleido isn't available in this environment; data sheets are unaffected."
           if include_chart_images else "")
    )

tab_projection, tab_performance = st.tabs(["📊 Projection Dashboard", "🎯 Performance Dashboard"])

with tab_projection:
    st.caption(
        f"Historical source: {hist_source_note} · "
        f"Projection: {'Course Calendar loaded' if calendar_ready else 'not uploaded — projected months show 0'} · "
        f"As of: {as_of_date:%d %b %Y}"
    )

    render_school_cards(active_schools[:5])

    st.divider()
    combined_segments = st.multiselect(
        "Segments to show", options=SEGMENTS, default=SEGMENTS, key="combined_segments",
        help="Pick any combination — e.g. just International, or International + CAAS.",
    )
    if not combined_segments:
        combined_segments = SEGMENTS
        st.info("No segments selected — showing all (Overall).")
    _combined_scope_label = "Overall" if set(combined_segments) == set(SEGMENTS) else " + ".join(combined_segments)

    _available_years = sorted(hist_df['Year'].dropna().astype(int).unique().tolist())
    # Which year powers the single "Combined total" chart: the current
    # reporting year if it's among the selected years, else the most
    # recent selected year (e.g. if only 2025 was picked).
    _primary_year = REPORT_YEAR if REPORT_YEAR in selected_years else (max(selected_years) if selected_years else REPORT_YEAR)
    _show_yoy = len(selected_years) >= 2 and any(y != _primary_year for y in selected_years)

    st.subheader(f"Combined total — {_primary_year} ({_combined_scope_label})")
    st.caption(
        "Bars: solid = actual · faded green = projected. "
        "Lines (right axis): cumulative running total — orange solid = actual, blue dotted = expected projection"
        + (". Red dotted = max scenario (only shown for Overall — Max has no per-segment breakdown in the source data)."
           if _combined_scope_label != "Overall" else ", red dotted = max scenario.")
    )
    st.plotly_chart(build_bar_chart(None, height=520, show_cumulative_lines=True, segments=combined_segments, year=_primary_year), use_container_width=True)

    if _show_yoy:
        st.divider()
        st.subheader(f"Combined total — Year on Year ({_combined_scope_label})")
        st.caption(
            f"Comparing {', '.join(str(y) for y in selected_years)} — solid = actual, faded green = projected "
            "(only the current reporting year has a projection; past years are fully actual). "
            f"Older years are shown at lower opacity ({selected_years[0]} is the faintest, {selected_years[-1]} is fully solid) "
            "so same-segment bars from different years are visually distinct. "
            "No cumulative lines on this view — those are specific to the current reporting year."
        )
        st.plotly_chart(build_yoy_bar_chart(selected_years, None, segments=combined_segments, height=600), use_container_width=True)

    _kpi1_actual, _kpi1_proj, _, _ = get_monthly_segment_matrix(None)
    _kpi1_total = sum(_kpi1_actual['International']) + sum(_kpi1_proj['International'])
    _kpi1_pct = (_kpi1_total / KPI1_TARGET * 100) if KPI1_TARGET > 0 else None
    k1, k2, k3 = st.columns(3)
    k1.metric(f"KPI 1 — International {metric_label}", f"{_kpi1_total:,.0f}")
    k2.metric("Target", f"{KPI1_TARGET:,.0f}")
    k3.metric("% of Target", f"{_kpi1_pct:,.0f}%" if _kpi1_pct is not None else "N/A")

    _kpi2_total = hist_df.loc[(hist_df['Segment'] == 'International') & (hist_df['Is Senior']), 'No of Trainee Days'].sum()
    _kpi2_pct = (_kpi2_total / KPI2_TARGET * 100) if KPI2_TARGET > 0 else None
    k4, k5, k6 = st.columns(3)
    k4.metric(f"KPI 2 — International Senior {metric_label}", f"{_kpi2_total:,.0f}")
    k5.metric("Target", f"{KPI2_TARGET:,.0f}")
    k6.metric("% of Target", f"{_kpi2_pct:,.0f}%" if _kpi2_pct is not None else "N/A")
    st.caption(
        f"KPI 2 source: {designation_source}. Reflects **actual data only** — the Course Calendar projection "
        "has no seniority data, so projected months aren't included in this figure."
    )

    st.divider()
    by_school_segments = st.multiselect(
        "Segments to show", options=SEGMENTS, default=['International'], key="by_school_segments",
        help="Pick any combination — e.g. just International, or International + CAAS.",
    )
    if not by_school_segments:
        by_school_segments = SEGMENTS
        st.info("No segments selected — showing all (Overall).")
    _by_school_scope_label = "Overall" if set(by_school_segments) == set(SEGMENTS) else " + ".join(by_school_segments)
    st.subheader(f"By school — {_by_school_scope_label} {metric_label.lower()}")
    grid_schools = SCHOOL_ORDER
    cols_row1 = st.columns(3)
    cols_row2 = st.columns(3)
    all_cols = cols_row1 + cols_row2
    for i, c in enumerate(all_cols):
        with c:
            if i < len(grid_schools):
                sch = grid_schools[i]
                st.markdown(f"**{sch} — {_by_school_scope_label} {metric_label}**")
                st.plotly_chart(build_bar_chart(sch, height=380, segments=by_school_segments, year=_primary_year), use_container_width=True)
                st.caption(f"{_by_school_scope_label} {metric_label.lower()} ({_primary_year}): **{segments_total(sch, by_school_segments, year=_primary_year):,.0f}**")
                if _show_yoy:
                    st.plotly_chart(build_yoy_bar_chart(selected_years, sch, segments=by_school_segments, height=500), use_container_width=True)
            else:
                render_executive_highlights(active_schools)

with tab_performance:
    st.header("🎯 Performance Dashboard")
    st.caption(f"Years: {', '.join(str(y) for y in selected_years)} (Actual)")

    st.markdown("""
<div style="display:flex; gap:12px; flex-wrap:wrap; align-items:stretch; margin-bottom:16px;">
  <div style="border:1px solid #e3e3e3; border-radius:10px; padding:10px 16px; display:flex; align-items:center; gap:8px; font-size:13px; color:#333;">
    <span>ℹ️ <b>Load Factor</b> =</span>
    <span style="display:inline-flex; flex-direction:column; align-items:center; font-size:12px; line-height:1.3;">
      <span style="padding:0 4px;">Total Trainees</span>
      <span style="border-top:1.5px solid #333; width:100%;"></span>
      <span style="padding:0 4px;">Max Class Size</span>
    </span>
    <span>× 100%</span>
  </div>
  <div style="border:1px solid #e3e3e3; border-radius:10px; padding:10px 16px; display:flex; align-items:center; gap:16px; font-size:12px; color:#333;">
    <span style="display:flex; align-items:center; gap:5px;"><span style="width:10px; height:10px; border-radius:2px; background:#2E7D32; display:inline-block;"></span>GOOD (≥80%)</span>
    <span style="display:flex; align-items:center; gap:5px;"><span style="width:10px; height:10px; border-radius:2px; background:#F2B705; display:inline-block;"></span>ACCEPTABLE (65–79%)</span>
    <span style="display:flex; align-items:center; gap:5px;"><span style="width:10px; height:10px; border-radius:2px; background:#F2994A; display:inline-block;"></span>BELOW TARGET (40–64%)</span>
    <span style="display:flex; align-items:center; gap:5px;"><span style="width:10px; height:10px; border-radius:2px; background:#E74C3C; display:inline-block;"></span>NEEDS ATTENTION (&lt;40%)</span>
  </div>
</div>
    """, unsafe_allow_html=True)

    perf_schools = sorted(hist_df['School'].dropna().unique().tolist())
    if not perf_schools:
        st.info("No historical (actual) data available yet for the Performance Dashboard.")
    elif calendar_capacity_df.empty:
        st.warning(
            "⚠️ No Course Calendar capacity data available — Load Factor needs Max Class Size per course intake, "
            "which comes from the Course Calendar / Projection List upload. Upload it in the sidebar to populate this tab."
        )
    else:
        perf_school = st.selectbox("School", options=["Overall"] + perf_schools, key="perf_school")
        _perf_result = compute_performance_data(perf_school)

        if _perf_result is None:
            _scope_text = "any school's" if perf_school == "Overall" else f"{perf_school}'s"
            st.warning(
                f"⚠️ None of {_scope_text} actual course intakes were found in the uploaded Course Calendar — "
                "Load Factor needs a Max Class Size match per intake, so there's nothing to show yet."
            )
        else:
            merged, course_agg, s = _perf_result
            total_trainees, total_capacity, overall_lf = s['total_trainees'], s['total_capacity'], s['overall_lf']
            highest, lowest, courses_conducted = s['highest'], s['lowest'], s['courses_conducted']

            c1, c2, c3, c4, c5 = st.columns(5)
            c1.metric("Overall Weighted Load Factor", f"{overall_lf:.1f}%")
            c1.caption(f"Trainees {total_trainees:,.0f} · Capacity {total_capacity:,.0f}")
            c2.metric("Highest Load Factor", f"{highest['Load Factor']:.0f}%")
            c2.caption(f"{highest['Course Title']} ({highest['Month-Year']})")
            c3.metric("Lowest Load Factor", f"{lowest['Load Factor']:.0f}%")
            c3.caption(f"{lowest['Course Title']} ({lowest['Month-Year']})")
            c4.metric("Courses Conducted", f"{courses_conducted}")
            c4.caption(f"Unique course codes ({s['n_runs']} course runs)")
            c5.metric("Total Trainees", f"{total_trainees:,.0f}")
            c5.caption("Across all matched courses")

            st.divider()
            st.subheader("Average Load Factor by Course (Weighted) — Highest to Lowest")
            max_bar_scale = max(150.0, float(course_agg['Load Factor'].max()) * 1.1)
            target_pct_pos = 80 / max_bar_scale * 100
            rows_html = ""
            for _, r in course_agg.iterrows():
                label, badge_color, bar_color = _lf_badge(r['Load Factor'])
                bar_pct = min(r['Load Factor'] / max_bar_scale * 100, 100)
                rows_html += f"""
<tr style="border-bottom:1px solid #eee;">
  <td style="padding:8px; text-align:center; font-weight:700; color:#1B3A6B;">{int(r['Rank'])}</td>
  <td style="padding:8px;"><b>{r['Course_Title']}</b><br><span style="font-size:11px; color:#888;">{r['Course Code (Performance)']}</span></td>
  <td style="padding:8px; text-align:center;">{int(r['Course_Runs'])}</td>
  <td style="padding:8px; text-align:center;">{int(r['Total_Trainees'])}</td>
  <td style="padding:8px; text-align:center;">{int(r['Total_Capacity'])}</td>
  <td style="padding:8px; width:240px;">
    <div style="position:relative; background:#f0f0f0; border-radius:4px; height:16px; width:100%;">
      <div style="position:absolute; left:{target_pct_pos:.1f}%; top:-2px; bottom:-2px; width:1px; background:#1E70C4;"></div>
      <div style="background:{bar_color}; height:16px; border-radius:4px; width:{bar_pct:.1f}%;"></div>
    </div>
  </td>
  <td style="padding:8px; text-align:right; font-weight:700; color:{bar_color};">{r['Load Factor']:.1f}%</td>
  <td style="padding:8px; text-align:center;"><span style="border:1px solid {badge_color}; color:{badge_color}; border-radius:4px; padding:2px 8px; font-size:11px; font-weight:700;">{label}</span></td>
</tr>
"""
            st.markdown(f"""
<table style="width:100%; border-collapse:collapse; font-size:13px;">
<thead><tr style="background:#192F55; color:#fff;">
<th style="padding:8px;">Rank</th><th style="padding:8px; text-align:left;">Course Title</th><th style="padding:8px;">Runs</th>
<th style="padding:8px;">Trainees</th><th style="padding:8px;">Capacity</th><th style="padding:8px;">Load Factor (80% target)</th><th></th><th style="padding:8px;">Performance</th>
</tr></thead>
<tbody>{rows_html}</tbody>
</table>
<p style="font-size:11px; color:#888; margin-top:8px;">Note: Weighted Load Factor = Total Trainees ÷ Total Capacity (Max Class Size). Capacity is sourced from the Course Calendar upload — course runs with no matching intake there are excluded from this table.</p>
            """, unsafe_allow_html=True)
